import os
import csv
import re
import openpyxl

class PKExtractorEngine:
    WEIGHT_REGEX = re.compile(r'[0-9]+\s*(?:oz|fl\s*oz|lb|lbs|ml|g|kg|fluid\s*ounce|ounce)s?', re.IGNORECASE)
    VALID_PATTERNS = [
        re.compile(r'\bpack of\s*([0-9]+)\b', re.IGNORECASE),
        re.compile(r'\bpacks of\s*([0-9]+)\b', re.IGNORECASE),
        re.compile(r'\b([0-9]+)\s*pack\b', re.IGNORECASE),
        re.compile(r'\b([0-9]+)-pack\b', re.IGNORECASE),
        re.compile(r'\bpk of\s*([0-9]+)\b', re.IGNORECASE),
        re.compile(r'\b([0-9]+)\s*pk\b', re.IGNORECASE),
        re.compile(r'\b([0-9]+)-pk\b', re.IGNORECASE),
    ]

    @classmethod
    def safe_stoi(cls, value):
        if value is None: return 0
        s = str(value).strip()
        if not s: return 0
        clean = "".join(ch for ch in s if ch.isdigit())
        if not clean: return 0
        try:
            val = int(clean)
            return val if val <= 2147483647 else 0
        except:
            return 0

    @classmethod
    def extract_true_pk(cls, title, original_pk):
        if not title: title = ""
        clean_title = cls.WEIGHT_REGEX.sub("", str(title).lower())
        extracted = -1
        for reg in cls.VALID_PATTERNS:
            match = reg.search(clean_title)
            if match:
                extracted = cls.safe_stoi(match.group(1))
                break
        if 0 < extracted < 100:
            return extracted
        return original_pk

    @classmethod
    def find_column(cls, columns, candidates):
        lower_map = {str(c).strip().lower(): i for i, c in enumerate(columns)}
        for cand in candidates:
            key = cand.lower()
            if key in lower_map: return lower_map[key]
        for cand in candidates:
            key = cand.lower()
            for name, idx in lower_map.items():
                if key in name: return idx
        return None

    @classmethod
    def process_file(cls, file_path, progress_callback=None):
        def log(msg):
            if progress_callback: progress_callback(msg)

        log("Dosya okunuyor, format kontrol ediliyor...")
        ext = os.path.splitext(file_path)[1].lower()
        out_path = os.path.splitext(file_path)[0] + "_pk_extracted" + ext
        
        if ext == ".xlsx":
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            data = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
            wb.close()
            if not data: raise ValueError("Dosya boş.")
            header, data_rows = data[0], data[1:]
        elif ext == ".csv":
            rows = []
            for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1254"):
                try:
                    with open(file_path, "r", encoding=enc, newline="") as f:
                        rows = list(csv.reader(f))
                    break
                except:
                    pass
            if not rows: raise ValueError("CSV okunamadı veya boş.")
            header, data_rows = rows[0], rows[1:]
        else:
            raise ValueError("Sadece .csv ve .xlsx desteklenir.")

        title_idx = cls.find_column(header, ["title"])
        pk_idx = cls.find_column(header, ["package quantity", "pk", "package qty"])

        if title_idx is None: raise ValueError("Başlık (Title) sütunu bulunamadı.")
        if pk_idx is None: raise ValueError("PK sütunu bulunamadı.")

        modified = 0
        out_rows = []
        total = len(data_rows)

        log(f"Veriler işleniyor... (Toplam Satır: {total})")

        for i, row in enumerate(data_rows):
            # Her 500 satırda bir iptal sinyali kontrolü ve log basımı yap
            if i % 500 == 0:
                log(f"Satırlar taranıyor: {i} / {total}")

            def get_val(idx): return row[idx] if idx < len(row) else ""
            original_pk = cls.safe_stoi(get_val(pk_idx))
            final_pk = cls.extract_true_pk(get_val(title_idx), original_pk)
            
            if final_pk != original_pk: modified += 1
            
            new_row = list(row)
            new_row.append(final_pk)
            out_rows.append(new_row)

        log("Sonuçlar yeni dosyaya yazılıyor...")
        new_header = list(header) + ["Extracted_PK"]

        if ext == ".xlsx":
            out_wb = openpyxl.Workbook()
            out_ws = out_wb.active
            out_ws.append(list(new_header))
            for r in out_rows: out_ws.append(list(r))
            out_wb.save(out_path)
            out_wb.close()
        else:
            with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(new_header)
                w.writerows(out_rows)

        return total, modified, out_path