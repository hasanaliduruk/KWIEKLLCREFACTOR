import os
import pandas as pd
import openpyxl

def check_column(df, possible_cols, file_name, col_type):
    for col in possible_cols:
        if col in df.columns:
            return col
    raise ValueError(f"Eksik Sütun Hatası: '{file_name}' dosyasında '{col_type}' sütunu bulunamadı. Beklenenler: {possible_cols}")


def process_order_create(
    restock_files: list,
    orderform_files: list,
    template_path: str,
    output_folder: str,
    settings_dict: dict,
    progress_callback=None,
) -> dict:
    if not restock_files:
        raise FileNotFoundError("Hata: Restock excel dosyası sağlanmadı.")
    if not orderform_files:
        raise FileNotFoundError("Hata: Order Form excel dosyası sağlanmadı.")
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Hata: Template dosyası bulunamadı -> {template_path}")

    sutunlar_dict = {
        "restock_upc": settings_dict.get("restock_columns", {}).get("upc", []),
        "restock_pcs": settings_dict.get("restock_columns", {}).get("pcs", []),
        "restock_suplier": settings_dict.get("restock_columns", {}).get("suplier", []),
        "restock_notes": settings_dict.get("restock_columns", {}).get("notes", []),
        "orderform_upc": settings_dict.get("orderform_columns", {}).get("upc", []),
        "orderform_pcs": settings_dict.get("orderform_columns", {}).get("pcs", []),
        "orderform_suplier": settings_dict.get("orderform_columns", {}).get("suplier", []),
    }

    # 1. RESTOCK İŞLEMLERİ (Vektörel)
    if progress_callback:
        progress_callback("Restock excel dosyası okunuyor...")
        
    # KRİTİK HATA DÜZELTİLDİ: Dosya önce okunmalı, sonra işlenmeli.
    df_restock = pd.read_excel(restock_files[0])

    upc_col = check_column(df_restock, sutunlar_dict["restock_upc"], "Restock", "UPC")
    pcs_col = check_column(df_restock, sutunlar_dict["restock_pcs"], "Restock", "PCS")
    suplier_col = check_column(df_restock, sutunlar_dict["restock_suplier"], "Restock", "SUPLIER")
    notes_col = check_column(df_restock, sutunlar_dict["restock_notes"], "Restock", "NOTES")

    # Veri Tiplerini Güvenceye Al
    df_restock[pcs_col] = pd.to_numeric(df_restock[pcs_col], errors="coerce").fillna(0)
    df_restock[suplier_col] = df_restock[suplier_col].fillna("").astype(str)
    df_restock[notes_col] = df_restock[notes_col].fillna("").astype(str)

    # PCS değeri 0 olanları at
    valid_restock = df_restock[df_restock[pcs_col] != 0].copy()

    # Tedarikçi (Supplier) bazlı gruplama ve toplama
    restock_sup = valid_restock.groupby([suplier_col, upc_col], as_index=False)[pcs_col].sum()
    restock_sup.rename(columns={suplier_col: "Target", upc_col: "UPC", pcs_col: "PCS"}, inplace=True)

    # Not (Note) bazlı gruplama ve toplama (İş Kuralı: Not varsa oraya da kopyalanır)
    valid_notes = valid_restock[(valid_restock[notes_col] != "") & (valid_restock[notes_col] != "0")].copy()
    restock_notes = valid_notes.groupby([notes_col, upc_col], as_index=False)[pcs_col].sum()
    restock_notes.rename(columns={notes_col: "Target", upc_col: "UPC", pcs_col: "PCS"}, inplace=True)


    # 2. ORDER FORM İŞLEMLERİ (Vektörel)
    if progress_callback:
        progress_callback("Order Form excel dosyası okunuyor...")
        
    df_order = pd.read_excel(orderform_files[0])

    o_upc_col = check_column(df_order, sutunlar_dict["orderform_upc"], "Order Form", "UPC")
    o_pcs_col = check_column(df_order, sutunlar_dict["orderform_pcs"], "Order Form", "PCS")
    o_sup_col = check_column(df_order, sutunlar_dict["orderform_suplier"], "Order Form", "SUPLIER")

    df_order[o_pcs_col] = pd.to_numeric(df_order[o_pcs_col], errors="coerce").fillna(0)
    df_order[o_sup_col] = df_order[o_sup_col].fillna("").astype(str)

    valid_order = df_order[df_order[o_pcs_col] != 0].copy()
    order_sup = valid_order.groupby([o_sup_col, o_upc_col], as_index=False)[o_pcs_col].sum()
    order_sup.rename(columns={o_sup_col: "Target", o_upc_col: "UPC", o_pcs_col: "PCS"}, inplace=True)


    # 3. VERİLERİ BİRLEŞTİR (Concat & GroupBy)
    # Tüm tablolar tek bir yapıda toplanıp nihai gruplama C hızında gerçekleştirilir.
    all_targets = pd.concat([restock_sup, restock_notes, order_sup], ignore_index=True)
    final_aggregation = all_targets.groupby(["Target", "UPC"], as_index=False)["PCS"].sum()


    # 4. ŞABLONA YAZDIRMA
    target_dir = os.path.join(output_folder, "ORDERS")
    os.makedirs(target_dir, exist_ok=True)

    if progress_callback:
        progress_callback("Bulunan değerler template dosyalarına yazdırılıyor...")

    # Dataframe'i hedef (Supplier veya Note) bazında ayırıp iterasyon yapıyoruz
    for target_name, group_df in final_aggregation.groupby("Target"):
        if not target_name:
            continue

        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        start_row = 2

        # Pandas üzerinden tuple array olarak çekerek O(1) erişim sağlıyoruz
        for i, (upc, pcs) in enumerate(group_df[["UPC", "PCS"]].itertuples(index=False)):
            ws.cell(row=start_row + i, column=1, value=upc)
            ws.cell(row=start_row + i, column=3, value=pcs)
            ws[f"A{start_row+i}"].number_format = "000000000000"

        safe_suplier = str(target_name).replace("/", "-").replace("\\", "-").upper()
        output_path = os.path.join(target_dir, f"{safe_suplier}.xlsx")
        wb.save(output_path)

    return {
        "status": "success",
        "message": "Order Create işlemi Vektörel Hızda başarıyla tamamlandı!",
        "output_path": target_dir,
    }