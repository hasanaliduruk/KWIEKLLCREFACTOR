
import os
import sys
import re
import sqlite3
from datetime import datetime, date
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional, Set, Any
from collections import Counter

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. CORE DOMAIN DATA MODELS
# ==========================================

@dataclass
class ShipmentHeader:
    shipment_id: str
    shipment_name: str
    created_date_raw: str
    created_date_formatted: str
    created_datetime: datetime

@dataclass
class ItemInfoRow:
    sku: str
    qty_shipped: int
    exp_date_usa: str
    exp_date_tur: str
    days_remaining: int

@dataclass
class FIFOResultRow:
    shipment_name: str
    shipment_id: str
    created_date: str
    created_dt: datetime
    sku: str
    qty_shipped: int
    exp_date_usa: str
    exp_date_tur: str
    days_remaining: int
    amz_stock_days: int
    amz_stock_allocated: int
    unfulfillable: int = 0
    note: str = ""


# ==========================================
# 2. UNIVERSAL POLYMORPHIC DATE & TEXT PARSERS
# ==========================================

class UniversalDateParser:
    MONTH_MAP = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }

    @classmethod
    def _strip_time(cls, val: Any) -> str:
        # '10-10-2026 10:57:34' gibi verilerden saat kısmını söküp atar
        val_str = str(val).strip()
        return re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?.*', '', val_str)

    @classmethod
    def parse_created_date(cls, date_val: Any) -> Tuple[str, datetime]:
        if pd.isna(date_val) or date_val is None:
            dt = datetime.now()
            return dt.strftime("%d %b %Y").lstrip("0"), dt

        val_str = cls._strip_time(date_val)

        try:
            parsed_dt = pd.to_datetime(val_str, errors='coerce', dayfirst=True)
            if pd.notnull(parsed_dt):
                dt = parsed_dt.to_pydatetime()
                return dt.strftime("%d %b %Y").lstrip("0"), dt
        except Exception:
            pass

        dt = datetime.now()
        return dt.strftime("%d %b %Y").lstrip("0"), dt

    @classmethod
    def parse_exp_date(cls, date_val: Any) -> Tuple[str, str, int]:
        if pd.isna(date_val) or date_val is None:
            return "", "", 0

        val_str = cls._strip_time(date_val)
        dt = None
        try:
            parsed_dt = pd.to_datetime(val_str, errors='coerce')
            if pd.notnull(parsed_dt):
                dt = parsed_dt.to_pydatetime()
        except Exception:
            pass

        if dt is None:
            return val_str, val_str, 0

        exp_usa = dt.strftime("%m-%d-%Y")
        exp_tur = dt.strftime("%d.%m.%Y")
        days_left = (dt.date() - datetime.now().date()).days
        return exp_usa, exp_tur, days_left

class ShipmentIDExtractor:
    """Extraction Engine for extracting FBA Shipment IDs from raw text or complex Worksheets."""

    @classmethod
    def extract_fba_shipment_ids_from_text(cls, text_val: Any) -> List[str]:
        if pd.isna(text_val) or text_val is None:
            return []
        
        val_str = str(text_val).strip()
        if not val_str:
            return []

        found_ids = []
        tokens = re.split(r'[,;\n\r\t\s]+', val_str)
        for token in tokens:
            token_clean = token.strip()
            match = re.search(r'\b(FBA[A-Z0-9]{8,12})\b', token_clean, re.IGNORECASE)
            if match:
                found_ids.append(match.group(1).upper())
        return found_ids

    @classmethod
    def extract_shipment_ids_from_file(cls, file_path: str) -> Set[str]:
        """Extracts FBA Shipment IDs from 2DWork, Picklist, or raw lists focusing on Column C first."""
        extracted_ids = set()
        filename = os.path.basename(file_path)

        fn_match = re.search(r'(FBA[A-Z0-9]{8,12})', filename, re.IGNORECASE)
        if fn_match:
            extracted_ids.add(fn_match.group(1).upper())

        try:
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names

            for sheet in sheet_names:
                df = pd.read_excel(xls, sheet_name=sheet, header=None)
                if df.empty:
                    continue

                # Check Column C (Index 2) specifically for 2DWork format
                if df.shape[1] > 2:
                    col_c_values = df.iloc[:, 2].dropna().tolist()
                    for val in col_c_values:
                        ids = cls.extract_fba_shipment_ids_from_text(val)
                        extracted_ids.update(ids)

                # Full Worksheet Scan if no IDs found in Column C
                if not extracted_ids:
                    for col_idx in range(df.shape[1]):
                        col_vals = df.iloc[:, col_idx].dropna().tolist()
                        for val in col_vals:
                            ids = cls.extract_fba_shipment_ids_from_text(val)
                            extracted_ids.update(ids)

        except Exception:
            pass

        return extracted_ids


class PicklistParser:
    @classmethod
    def parse_picklist_file(cls, file_path: str) -> Tuple[ShipmentHeader, List[ItemInfoRow]]:
        xls = pd.ExcelFile(file_path)
        
        df_shipment = pd.read_excel(xls, sheet_name='Shipment Info')
        shipment_id = str(df_shipment['Shipment Id'].iloc[0]).strip()
        shipment_name = str(df_shipment['Shipment Name'].iloc[0]).strip()
        created_raw = df_shipment['Created'].iloc[0] if 'Created' in df_shipment.columns else ""
        
        created_formatted, created_dt = UniversalDateParser.parse_created_date(created_raw)
        
        header = ShipmentHeader(
            shipment_id=shipment_id,
            shipment_name=shipment_name,
            created_date_raw=str(created_raw),
            created_date_formatted=created_formatted,
            created_datetime=created_dt
        )
        
        df_items = pd.read_excel(xls, sheet_name='Items Info')
        items = []
        for _, row in df_items.iterrows():
            sku = str(row['SKU']).strip() if pd.notnull(row['SKU']) else ""
            if not sku or sku.lower() == 'nan':
                continue
            qty = int(row['QTY (Shipped)']) if pd.notnull(row.get('QTY (Shipped)')) else 0
            exp_raw = row.get('Expiration date', '')
            exp_usa, exp_tur, days_left = UniversalDateParser.parse_exp_date(exp_raw)
            
            items.append(ItemInfoRow(
                sku=sku,
                qty_shipped=qty,
                exp_date_usa=exp_usa,
                exp_date_tur=exp_tur,
                days_remaining=days_left
            ))
            
        return header, items


# ==========================================
# 3. PERSISTENCE DATABASE LAYER (Auto-Migrating)
# ==========================================

class DatabaseManager:
    """
    Multi-User SQLite Connection Engine:
    Configured with WAL Mode, Busy Timeouts, Auto-Migration, and ANALİZ Sheet Note Ingestion.
    """
    def __init__(self, db_path: str = "fba_inventory.db"):
        self.db_path = db_path
        self.init_db()

    def get_connection(self):
        conn = sqlite3.connect(self.db_path, timeout=30.0)
        try:
            conn.execute("PRAGMA journal_mode=WAL;")
            conn.execute("PRAGMA busy_timeout=30000;")
        except Exception:
            pass
        return conn

    def init_db(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='shipment_items'")
            table_sql_row = cursor.fetchone()
            
            if table_sql_row and 'shipment_name' in table_sql_row[0] and 'UNIQUE(shipment_id, sku, exp_date_usa)' in table_sql_row[0]:
                cursor.execute("""
                    CREATE TABLE shipment_items_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        shipment_id TEXT NOT NULL,
                        shipment_name TEXT NOT NULL,
                        created_date TEXT NOT NULL,
                        created_timestamp TIMESTAMP NOT NULL,
                        sku TEXT NOT NULL,
                        qty INTEGER NOT NULL,
                        exp_date_usa TEXT,
                        exp_date_tur TEXT,
                        days_remaining INTEGER,
                        note TEXT NOT NULL DEFAULT '',
                        UNIQUE(shipment_name, shipment_id, created_date, sku, qty, exp_date_usa) ON CONFLICT REPLACE
                    )
                """)
                cursor.execute("""
                    INSERT INTO shipment_items_new 
                    (shipment_id, shipment_name, created_date, created_timestamp, sku, qty, exp_date_usa, exp_date_tur, days_remaining, note)
                    SELECT shipment_id, shipment_name, created_date, created_timestamp, sku, qty, exp_date_usa, exp_date_tur, days_remaining, note
                    FROM shipment_items
                """)
                cursor.execute("DROP TABLE shipment_items")
                cursor.execute("ALTER TABLE shipment_items_new RENAME TO shipment_items")
            else:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS shipment_items (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        shipment_id TEXT NOT NULL,
                        shipment_name TEXT NOT NULL,
                        created_date TEXT NOT NULL,
                        created_timestamp TIMESTAMP NOT NULL,
                        sku TEXT NOT NULL,
                        qty INTEGER NOT NULL,
                        exp_date_usa TEXT,
                        exp_date_tur TEXT,
                        days_remaining INTEGER,
                        note TEXT NOT NULL DEFAULT '',
                        UNIQUE(shipment_name, shipment_id, created_date, sku, qty, exp_date_usa) ON CONFLICT REPLACE
                    )
                """)

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS amazon_stock (
                    sku TEXT PRIMARY KEY,
                    total_units INTEGER NOT NULL,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            cursor.execute("CREATE INDEX IF NOT EXISTS idx_shipment_sku ON shipment_items(sku)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_created_timestamp ON shipment_items(created_timestamp)")
            try:
                cursor.execute("ALTER TABLE amazon_stock ADD COLUMN unfulfillable INTEGER DEFAULT 0")
            except Exception:
                pass
            conn.commit()

    def update_item_note(self, shipment_id: str, sku: str, exp_date_usa: str, note: str):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if exp_date_usa and exp_date_usa.strip():
                cursor.execute("""
                    UPDATE shipment_items 
                    SET note = ? 
                    WHERE UPPER(shipment_id) = UPPER(?) 
                      AND UPPER(sku) = UPPER(?) 
                      AND UPPER(exp_date_usa) = UPPER(?)
                """, (note.strip(), shipment_id.strip(), sku.strip(), exp_date_usa.strip()))
            else:
                cursor.execute("""
                    UPDATE shipment_items 
                    SET note = ? 
                    WHERE UPPER(shipment_id) = UPPER(?) 
                      AND UPPER(sku) = UPPER(?)
                """, (note.strip(), shipment_id.strip(), sku.strip()))
            conn.commit()

    def reset_all_data(self):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM shipment_items;")
            cursor.execute("DELETE FROM amazon_stock;")
            conn.commit()

    def is_shipment_exists(self, shipment_id: str) -> bool:
        if not shipment_id or shipment_id.lower() == 'nan' or shipment_id.strip().upper() == 'BOŞ':
            return False
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT 1 FROM shipment_items WHERE UPPER(shipment_id) = ? LIMIT 1", (shipment_id.strip().upper(),))
            return cursor.fetchone() is not None

    def get_all_registered_shipment_ids(self) -> Set[str]:
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT UPPER(shipment_id) FROM shipment_items WHERE UPPER(shipment_id) != 'BOŞ'")
            return {row[0].strip() for row in cursor.fetchall() if row[0]}

    def import_babil_master_excel(self, file_path: str) -> Tuple[bool, str]:
        import time  # Süre ölçümü için time modülü
        
        genel_baslangic = time.time()
        
        # --- AŞAMA 1: MEVCUT KAYITLARI KONTROL ETME ---
        candidate_ids = ShipmentIDExtractor.extract_shipment_ids_from_file(file_path)
        for s_id in candidate_ids:
            if self.is_shipment_exists(s_id):
                return False, f"Shipment ID {s_id} sistemde kayıtlıdır."

        # --- AŞAMA 2: EXCEL OKUMA (TARGET SHEET) ---
        xls = pd.ExcelFile(file_path)
        sheet_names = xls.sheet_names
        
        target_sheet = None
        for candidate in ['SIRALI', 'ANALİZ', 'FORMÜL', 'EXPRATION-DATE-REPORT']:
            for s in sheet_names:
                if candidate in s.upper():
                    target_sheet = s
                    break
            if target_sheet:
                break
        if not target_sheet:
            target_sheet = sheet_names[0]

        df = pd.read_excel(xls, sheet_name=target_sheet)
        cols = list(df.columns)

        col_name = next((c for c in cols if 'name' in str(c).lower()), cols[0])
        col_id = next((c for c in cols if 'id' in str(c).lower()), cols[1] if len(cols) > 1 else cols[0])
        
        col_date = next((c for c in cols if 'date' in str(c).lower() and 'exp' not in str(c).lower()), None)
        if not col_date and len(cols) > 2:
            col_date = cols[2]

        col_sku = next((c for c in cols if 'sku' in str(c).lower()), cols[3] if len(cols) > 3 else cols[0])
        col_qty = next((c for c in cols if any(k in str(c).lower() for k in ['qty', 'quantity', 'adet', 'miktar'])), cols[4] if len(cols) > 4 else cols[0])
        
        col_exp_usa = next((c for c in cols if 'exp' in str(c).lower() and 'usa' in str(c).lower()), None)
        if not col_exp_usa:
            col_exp_usa = next((c for c in cols if 'exp' in str(c).lower()), cols[5] if len(cols) > 5 else cols[0])

        # --- AŞAMA 3: ANALİZ SAYFASINDAN NOTLARI ÇEKME ---
        notes_dict = {}
        for s in sheet_names:
            if 'ANALİZ' in s.upper() or 'ANALIZ' in s.upper():
                df_analiz_sheet = pd.read_excel(xls, sheet_name=s)
                analiz_cols = list(df_analiz_sheet.columns)
                a_col_id = next((c for c in analiz_cols if 'id' in str(c).lower()), None)
                a_col_sku = next((c for c in analiz_cols if 'sku' in str(c).lower()), None)
                a_col_note = next((c for c in analiz_cols if 'not' in str(c).lower() or 'note' in str(c).lower() or 'açıklama' in str(c).lower()), None)
                
                if a_col_id and a_col_sku and a_col_note:
                    for _, a_row in df_analiz_sheet.iterrows():
                        a_id = str(a_row[a_col_id]).strip().replace('\t', '').replace('\n', '').upper() if pd.notnull(a_row.get(a_col_id)) else ""
                        a_sku = str(a_row[a_col_sku]).strip().replace('\t', '').replace('\n', '').upper() if pd.notnull(a_row.get(a_col_sku)) else ""
                        a_note = str(a_row[a_col_note]).strip() if pd.notnull(a_row.get(a_col_note)) and str(a_row.get(a_col_note)).lower() != 'nan' else ""
                        if a_id and a_sku and a_note:
                            notes_dict[(a_id, a_sku)] = a_note


        # --- AŞAMA 4: VEKTÖRİZASYON VE METİN TEMİZLİĞİ ---
        now = datetime.now()

        df[col_sku] = df[col_sku].astype(str).str.strip().str.replace('\t', '', regex=False).str.replace('\n', '', regex=False)
        df[col_id] = df[col_id].astype(str).str.strip().str.replace('\t', '', regex=False).str.replace('\n', '', regex=False)
        df[col_name] = df[col_name].astype(str).str.strip().str.replace('\t', '', regex=False).str.replace('\n', '', regex=False)
        df = df[(df[col_sku] != '') & (df[col_id] != '') & (df[col_sku].str.lower() != 'nan')].copy()
        
        df['__qty'] = pd.to_numeric(
            df[col_qty].astype(str).str.replace(',', '.').str.extract(r'(\d+\.?\d*)', expand=False),
            errors='coerce').fillna(0).astype(int)

        if col_date in df.columns:
            date_clean = df[col_date].astype(str).str.replace(r'\s+\d{1,2}:\d{2}(:\d{2})?.*', '', regex=True)
        else:
            date_clean = pd.Series([''], index=df.index)

        created_series = pd.to_datetime(date_clean, errors='coerce', dayfirst=True).fillna(pd.Timestamp(now))
        df['__created_dt'] = created_series
        df['__created_fmt'] = created_series.dt.strftime('%d %b %Y').str.lstrip('0')

        if col_exp_usa in df.columns:
            exp_clean = df[col_exp_usa].astype(str).str.replace(r'\s+\d{1,2}:\d{2}(:\d{2})?.*', '', regex=True)
        else:
            exp_clean = pd.Series([''], index=df.index)

        exp_series = pd.to_datetime(exp_clean, errors='coerce')

        df['__exp_usa'] = exp_series.dt.strftime('%m-%d-%Y').fillna("")
        df['__exp_tur'] = exp_series.dt.strftime('%d.%m.%Y').fillna("")

        df['__days_remaining'] = (exp_series.dt.normalize() - pd.Timestamp(now).normalize()).dt.days.fillna(0).astype(int)

        records = df[[col_id, col_name, col_sku, '__qty', '__created_dt', '__created_fmt', '__exp_usa', '__exp_tur', '__days_remaining']].to_dict('records')


        # --- AŞAMA 5: SQL INSERT DİZİSİNİ OLUŞTURMA ---
        insert_data = []
        for row in records:
            s_id = str(row[col_id])
            s_name = str(row[col_name])
            sku = str(row[col_sku])

            matched_note = notes_dict.get((s_id.upper(), sku.upper()), "")

            insert_data.append((
                s_id, s_name, row['__created_fmt'], row['__created_dt'].to_pydatetime(), sku, row['__qty'],
                row['__exp_usa'], row['__exp_tur'], row['__days_remaining'], matched_note, s_name, s_id, sku
            ))

        rows_imported = len(insert_data)


        # --- AŞAMA 5: SQL INSERT DİZİSİNİ OLUŞTURMA ---
        
        # SQL Subquery yerine mevcut notları RAM'e çekiyoruz (O(1) erişim)
        existing_notes = {}
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT shipment_id, sku, note FROM shipment_items WHERE note != '' AND note IS NOT NULL")
            for r_id, r_sku, r_note in cursor.fetchall():
                if r_id and r_sku:
                    existing_notes[(str(r_id).upper(), str(r_sku).upper())] = r_note

        insert_data = []
        for row in records:
            s_id = str(row[col_id])
            s_name = str(row[col_name])
            sku = str(row[col_sku])

            # 1. Excel'den (Analiz sayfasından) gelen not var mı?
            matched_note = notes_dict.get((s_id.upper(), sku.upper()), "")
            
            # 2. Yoksa veritabanındaki eski notu koru
            if not matched_note:
                matched_note = existing_notes.get((s_id.upper(), sku.upper()), "")

            # Tuple uzunluğu saf INSERT için 10 parametreye düşürüldü
            insert_data.append((
                s_id, s_name, row['__created_fmt'], row['__created_dt'].to_pydatetime(), sku, row['__qty'],
                row['__exp_usa'], row['__exp_tur'], row['__days_remaining'], matched_note
            ))

        rows_imported = len(insert_data)


        # --- AŞAMA 6 & 7: VERİTABANINA YAZMA (BULK INSERT) ---
        with self.get_connection() as conn:
            cursor = conn.cursor()
            
            # Saf, alt sorgusuz (subquery-free) yüksek hızlı INSERT
            cursor.executemany("""
                INSERT INTO shipment_items
                (shipment_id, shipment_name, created_date, created_timestamp, sku, qty, exp_date_usa,
                 exp_date_tur, days_remaining, note)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, insert_data)

            if 'Amazon' in sheet_names:
                df_amz = pd.read_excel(xls, sheet_name='Amazon')
                amz_insert_data = []
                sku_col = next((c for c in df_amz.columns if 'sku' in str(c).lower()), df_amz.columns[0])
                qty_col = next(
                    (c for c in df_amz.columns if any(k in str(c).lower() for k in ['qty', 'units', 'total'])),
                    df_amz.columns[1])

                unfill_col = next((c for c in df_amz.columns if 'unfulfillable' in str(c).lower()), None)

                for _, r in df_amz.iterrows():
                    s = str(r[sku_col]).strip().replace('\t', '').replace('\n', '')
                    try:
                        q = int(float(str(r[qty_col]).strip().replace(',', '.'))) if pd.notnull(r[qty_col]) else 0
                    except Exception:
                        q = 0

                    u_qty = 0
                    if unfill_col:
                        try:
                            u_qty = int(float(str(r[unfill_col]).strip().replace(',', '.'))) if pd.notnull(r[unfill_col]) else 0
                        except Exception:
                            pass

                    if s and s.lower() != 'nan':
                        amz_insert_data.append((s, q, u_qty))

                cursor.executemany(
                    "INSERT OR REPLACE INTO amazon_stock (sku, total_units, unfulfillable) VALUES (?, ?, ?)",
                    amz_insert_data)

            conn.commit()

        return True, f"Master Excel verileri başarıyla yüklendi. ({rows_imported} satır eksiksiz aktarıldı, {len(notes_dict)} adet açıklama notu işlendi)"

    def add_picklist_shipment(self, header: ShipmentHeader, items: List[ItemInfoRow]):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            for item in items:
                cursor.execute("""
                    INSERT INTO shipment_items 
                    (shipment_id, shipment_name, created_date, created_timestamp, sku, qty, exp_date_usa, exp_date_tur, days_remaining, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, COALESCE((SELECT note FROM shipment_items WHERE UPPER(shipment_name)=UPPER(?) AND UPPER(shipment_id)=UPPER(?) AND UPPER(sku)=UPPER(?)), ''))
                """, (
                    header.shipment_id,
                    header.shipment_name,
                    header.created_date_formatted,
                    header.created_datetime,
                    item.sku,
                    item.qty_shipped,
                    item.exp_date_usa,
                    item.exp_date_tur,
                    item.days_remaining,
                    header.shipment_name,
                    header.shipment_id,
                    item.sku
                ))
            conn.commit()

    def update_amazon_stock(self, stock_dict: Dict[str, Dict[str, int]]):
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM amazon_stock;")
            for sku, data in stock_dict.items():
                cursor.execute("""
                    INSERT INTO amazon_stock (sku, total_units, unfulfillable) VALUES (?, ?, ?)
                """, (sku, data['total'], data['unfulfillable']))
            conn.commit()

    def get_all_shipments_sorted_by_date(self) -> List[Dict]:
        with self.get_connection() as conn:
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM shipment_items ORDER BY created_timestamp ASC, id ASC")
            return [dict(r) for r in cursor.fetchall()]

    def get_amazon_stock_map(self) -> Dict[str, Dict[str, int]]:
        with self.get_connection() as conn:
            cursor = conn.cursor()
            try:
                cursor.execute("SELECT sku, total_units, unfulfillable FROM amazon_stock")
                return {r[0]: {'total': r[1], 'unfulfillable': r[2]} for r in cursor.fetchall()}
            except:
                cursor.execute("SELECT sku, total_units FROM amazon_stock")
                return {r[0]: {'total': r[1], 'unfulfillable': 0} for r in cursor.fetchall()}


# ==========================================
# 4. ALGORITHMIC FIFO CALCULATOR ENGINE
# ==========================================

class FIFOEngine:
    @staticmethod
    def calculate_fifo(raw_shipments: List[Dict], amazon_stock_map: Dict[str, int]) -> Tuple[List[Dict], List[FIFOResultRow]]:
        sirali_list = []
        sku_groups = {}
        today_dt = datetime.now().date()

        for item in raw_shipments:
            exp_usa = item['exp_date_usa']
            exp_tur = item['exp_date_tur']
            days_left = 0
            
            if exp_usa:
                _, exp_tur, days_left = UniversalDateParser.parse_exp_date(exp_usa)

            created_dt = datetime.strptime(str(item['created_timestamp'])[:10], "%Y-%m-%d") if item['created_timestamp'] else datetime.min
            
            amz_stok_gun = (today_dt - created_dt.date()).days - 30
            note_val = item.get('note', '') or ''

            formatted_item = {
                'shipment_name': item['shipment_name'],
                'shipment_id': item['shipment_id'],
                'created_date': item['created_date'],
                'created_dt': created_dt,
                'sku': item['sku'],
                'qty_shipped': item['qty'],
                'exp_date_usa': exp_usa,
                'exp_date_tur': exp_tur,
                'days_remaining': days_left,
                'amz_stock_days': amz_stok_gun,
                'note': note_val
            }

            sirali_list.append(formatted_item)
            sku_groups.setdefault(item['sku'], []).append(formatted_item)

        sirali_list.sort(key=lambda x: (x['created_dt'], x['sku']))

        analiz_list = []

        for sku in sorted(sku_groups.keys()):
            lots = sku_groups[sku]
            lots.sort(key=lambda x: x['created_dt'])

            total_amazon_stock = amazon_stock_map.get(sku, {}).get('total', 0)
            unfill_qty = amazon_stock_map.get(sku, {}).get('unfulfillable', 0)

            for i, lot in enumerate(lots):
                qty_shipped = lot['qty_shipped']
                sum_newer_qty = sum(l['qty_shipped'] for l in lots[i + 1:])

                if total_amazon_stock <= 0:
                    allocated = 0
                elif sum_newer_qty >= total_amazon_stock:
                    allocated = 0
                elif (sum_newer_qty + qty_shipped) > total_amazon_stock:
                    allocated = total_amazon_stock - sum_newer_qty
                else:
                    allocated = qty_shipped

                if allocated > 0:
                    analiz_list.append(FIFOResultRow(
                        shipment_name=lot['shipment_name'],
                        shipment_id=lot['shipment_id'],
                        created_date=lot['created_date'],
                        created_dt=lot['created_dt'],
                        sku=lot['sku'],
                        qty_shipped=qty_shipped,
                        exp_date_usa=lot['exp_date_usa'],
                        exp_date_tur=lot['exp_date_tur'],
                        days_remaining=lot['days_remaining'],
                        amz_stock_days=lot['amz_stock_days'],
                        amz_stock_allocated=allocated,
                        unfulfillable=unfill_qty,
                        note=lot['note']
                    ))

        return sirali_list, analiz_list


# ==========================================
# 5. EXCEL REPORT EXPORTER
# ==========================================

class ExcelReportExporter:
    @staticmethod
    def export_master_excel(file_path: str, db: DatabaseManager):
        import time
        genel_baslangic = time.time()

        # --- 1. Veri Okuma ve FIFO Hesaplaması ---
        raw_shipments = db.get_all_shipments_sorted_by_date()
        stock_map = db.get_amazon_stock_map()
        sirali_list, analiz_list = FIFOEngine.calculate_fifo(raw_shipments, stock_map)
        analiz_sku_counts = Counter(item.sku for item in analiz_list)

        # --- 2. Excel Still Tanımlamaları ---
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)

        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        alert_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        alert_font = Font(name="Segoe UI", size=10, bold=True, color="B71C1C")

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(name="Segoe UI", size=10, bold=True, color="006100")

        # --- 3. SIRALI Tablosu ---
        ws_sirali = wb.create_sheet(title="SIRALI")
        headers_sirali = ["SHIPMENT NAME", "SHIPMENT ID", "DATE", "SKU", "QTY", "EXP DATE USA", "EXP DATE TUR", "SKT GÜN", "AMZ Stok Gün"]
        ws_sirali.append(headers_sirali)

        for col in range(1, len(headers_sirali) + 1):
            cell = ws_sirali.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(sirali_list):
            row_vals = [
                item['shipment_name'], item['shipment_id'], item['created_date'],
                item['sku'], item['qty_shipped'], item['exp_date_usa'],
                item['exp_date_tur'], item['days_remaining'], item['amz_stock_days']
            ]
            ws_sirali.append(row_vals)
            curr_row = idx + 2  # max_row taraması iptal edildi, statik sayaç eklendi

            for col in range(1, 10):
                c = ws_sirali.cell(row=curr_row, column=col)
                c.font = regular_font
                c.border = border_thin
                if col in [1, 4]:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # --- 4. ANALİZ Tablosu ---
        ws_analiz = wb.create_sheet(title="ANALİZ")
        headers_analiz = ["SHIPMENT NAME", "SHIPMENT ID", "SKU", "UNFULFILLABLE", "QTY", "AMZ STOK", "AMZ STOK GÜN", "SKT GÜN", "NOT"]
        ws_analiz.append(headers_analiz)

        for col in range(1, len(headers_analiz) + 1):
            cell = ws_analiz.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, item in enumerate(analiz_list):
            row_vals = [
                item.shipment_name, item.shipment_id, item.sku, item.unfulfillable,
                item.qty_shipped, item.amz_stock_allocated,
                item.amz_stock_days, item.days_remaining, item.note
            ]
            ws_analiz.append(row_vals)
            curr_row = idx + 2

            sku_cell = ws_analiz.cell(row=curr_row, column=3)
            unfill_cell = ws_analiz.cell(row=curr_row, column=4)
            stock_cell = ws_analiz.cell(row=curr_row, column=6)
            days_cell = ws_analiz.cell(row=curr_row, column=8)

            if analiz_sku_counts[item.sku] > 1:
                sku_cell.fill = green_fill
                sku_cell.font = green_font
            else:
                sku_cell.font = regular_font

            if getattr(item, 'unfulfillable', 0) > 0:
                unfill_cell.fill = alert_fill
                unfill_cell.font = alert_font
            else:
                unfill_cell.font = regular_font

            if item.days_remaining <= 180:
                days_cell.fill = alert_fill
                days_cell.font = alert_font
            else:
                days_cell.font = regular_font

            stock_cell.fill = green_fill
            stock_cell.font = bold_font

            for col in range(1, 10):
                c = ws_analiz.cell(row=curr_row, column=col)
                if col not in [3, 4, 6, 8] or (col == 3 and analiz_sku_counts[item.sku] <= 1) or (
                        col == 4 and getattr(item, 'unfulfillable', 0) <= 0) or (col == 8 and item.days_remaining > 180):
                    c.font = regular_font
                c.border = border_thin

                if col in [1, 3, 9]:
                    c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # --- 5. AMAZON Tablosu ---
        ws_amz = wb.create_sheet(title="Amazon")
        ws_amz.append(["SKU", "Total Units"])
        for col in range(1, 3):
            c = ws_amz.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        for sku, data in stock_map.items():
            ws_amz.append([sku, data['total']])

        # --- 6. SÜTUN GENİŞLİKLERİ (ÖLÜMCÜL DARBOĞAZIN ÇÖZÜMÜ) ---
        # Dinamik 300.000 hücre taraması iptal edildi, sabit genişlikler atandı.
        column_widths = {
            "SIRALI": [40, 15, 12, 18, 8, 12, 12, 10, 12],
            "ANALİZ": [40, 15, 18, 15, 8, 12, 15, 10, 30],
            "Amazon": [18, 12]
        }
        
        for sheet_name, widths in column_widths.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for i, width in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width

        # --- 7. KAYIT ---
        wb.save(file_path)

        