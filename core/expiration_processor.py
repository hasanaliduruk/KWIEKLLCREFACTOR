import os
import traceback
import warnings
import requests
from bs4 import BeautifulSoup
from bs4 import XMLParsedAsHTMLWarning
import openpyxl

def ensure_settings():
    if not os.path.exists("Settings"):
        os.makedirs("Settings")
    settings_path = "Settings/expration_settings.txt"
    if not os.path.exists(settings_path):
        with open(settings_path, "w", encoding='utf-8') as file:
            file.write('login_button_id = mainForm:j_idt23, mainForm:j_idt13, mainForm:j_idt22\n'
                       'email_id = mainForm:email\n'
                       'password_id = mainForm:password\n'
                       'default_email = sales@buyable.net\n'
                       'default_password = hasali2603\n')

def read_settings():
    dictionary = {
        'login_button_id': [], 'email_id': [], 'password_id': [],
        'default_email': [], 'default_password': [],
    }
    with open("Settings/expration_settings.txt", "r", encoding='utf-8') as file:
        lines = file.readlines()
        for line in lines:
            parts = line.split('=')
            if len(parts) < 2: continue
            key = parts[0].strip()
            values = [v.strip() for v in parts[1].split(',')]
            if key in dictionary:
                dictionary[key].extend(values)
    return dictionary

def format_date(date_str):
    if not date_str:
        return None
    try:
        date_str = date_str.replace(' ', '')
        if '-' in date_str: parts = date_str.split('-')
        elif '/' in date_str: parts = date_str.split('/')
        elif '.' in date_str: parts = date_str.split('.')
        else: return None
        return f"{parts[1]}.{parts[0]}.{parts[2]}"
    except Exception:
        return None

def write_excel(main_dictionary, target_id, path):
    save_name = f"{target_id}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = save_name
    
    headers = ['NAME', 'SHIPMENT ID', 'SHIPMENT DATE', 'SKU', 'SHIPPED', 'DATE', 'TR DATE']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header

    a = 2
    for sku, data in main_dictionary[target_id].items():
        ws.cell(row=a, column=1).value = data.get("shipment_name", "")
        ws.cell(row=a, column=2).value = target_id
        ws.cell(row=a, column=3).value = data.get("created", "")
        ws.cell(row=a, column=4).value = sku
        ws.cell(row=a, column=5).value = data.get("shipped", "")
        ws.cell(row=a, column=6).value = data.get("date")[0] if data.get("date") else ""
        ws.cell(row=a, column=7).value = data.get("noktali", "")
        
        c = 8
        for date_val in data.get("date", [])[1:]:
            ws.cell(row=a, column=c).value = str(date_val)
            c += 1
        a += 1

    wb.save(os.path.join(path, save_name))

def write_combined_excel(main_dictionary, path):
    save_name = "combined.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined"

    headers = ['NAME', 'SHIPMENT ID', 'SHIPMENT DATE', 'SKU', 'SHIPPED', 'DATE', 'TR DATE']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col).value = header

    a = 2
    for target_id, skus in main_dictionary.items():
        for sku, data in skus.items():
            ws.cell(row=a, column=1).value = data.get("shipment_name", "")
            ws.cell(row=a, column=2).value = target_id
            ws.cell(row=a, column=3).value = data.get("created", "")
            ws.cell(row=a, column=4).value = sku
            ws.cell(row=a, column=5).value = data.get("shipped", "")
            ws.cell(row=a, column=6).value = data.get("date")[0] if data.get("date") else ""
            ws.cell(row=a, column=7).value = data.get("noktali", "")
            
            c = 8
            for date_val in data.get("date", [])[1:]:
                ws.cell(row=a, column=c).value = str(date_val)
                c += 1
            a += 1

    wb.save(os.path.join(path, save_name))

def process_expiration(username, password, item_ids_str, output_path, progress_callback=None):
    warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
    
    if not username or not password:
        raise ValueError("Kullanıcı adı veya şifre boş olamaz.")
    if not item_ids_str:
        raise ValueError("İşlenecek Item ID(leri) eksik.")
    if not os.path.exists(output_path):
        os.makedirs(output_path, exist_ok=True)
        
    ensure_settings()
    settings = read_settings()
    
    if progress_callback: progress_callback("2D Workflow'a giriş yapılıyor...", "white")
    
    url = "https://app.2dworkflow.com/login.jsf"
    session = requests.Session()
    response = session.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    
    try:
        javax = soup.find('input', {'name': 'javax.faces.ViewState'})['value']
        button = soup.find('button')['name']
    except TypeError:
        raise RuntimeError("Giriş sayfası yapısı değişmiş veya sunucu yanıt vermiyor.")
        
    payload = {
        'mainForm:email': username,
        'mainForm:password': password,
        'mainForm': 'mainForm',
        'javax.faces.ViewState': javax,
        button: ''
    }
    
    response = session.post(url, data=payload)
    if response.status_code != 200 or "login" in response.url.lower():
        raise PermissionError("Giriş başarısız! Lütfen kullanıcı adı ve şifrenizi kontrol edin.")
        
    if progress_callback: progress_callback("Giriş başarılı. Shipment verileri çekiliyor...", "#90EE90")
    
    id_list = [i.strip() for i in item_ids_str.split(',') if i.strip()]
    id_dict = {}
    main_dict = {}
    
    shipments_url = "https://app.2dworkflow.com/shipped.jsf"
    response = session.get(shipments_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    tbody = soup.find("tbody", id="mainForm:shipments_data")
    
    fba_date = ""
    for target_id in id_list:
        main_dict[target_id] = {}
        href = ""
        if tbody:
            trs = tbody.findAll("tr")
            for index, tr in enumerate(trs):
                a_tag = tr.find("a")
                if a_tag and target_id in a_tag.get("title", ""):
                    href = a_tag.get("href", "")
                    try:
                        fba_date = trs[index-1].findAll("span")[2].text.split(",")[0]
                    except IndexError:
                        pass
        if not href:
            if progress_callback: progress_callback(f"Uyarı: '{target_id}' için sevkiyat linki bulunamadı.", "yellow")
        id_dict[target_id] = href
        
    for target_id in id_list:
        if not id_dict[target_id]: continue
        
        target_url = f"https://app.2dworkflow.com/{id_dict[target_id]}"
        response = session.get(target_url)
        soup = BeautifulSoup(response.text, 'html.parser')
        
        tbody_items = soup.find("tbody", {"id": "mainForm:shipmentItems_data"})
        tbody_info = soup.find("tbody", {"id": "mainForm:shipmentInfo_data"})
        
        if not tbody_items or not tbody_info:
            continue
            
        info_tr = tbody_info.find("tr")
        shipment_name = info_tr.findAll("td")[3].text if info_tr else "Unknown"
        
        trler = tbody_items.findAll("tr")
        for tr in trler:
            sku = tr.find("span").text
            main_dict[target_id][sku] = {
                "item_id": tr.get('data-rk', ''),
                "shipped": tr.findAll("td")[3].text if len(tr.findAll("td")) > 3 else "",
                "created": fba_date,
                "shipment_name": shipment_name,
                "date": []
            }
            
        if progress_callback: progress_callback(f"{target_id} için {len(main_dict[target_id])} adet ürün bulundu, işleniyor...", "white")
        javax = soup.find('input', {'name': 'javax.faces.ViewState'})['value']
        
        skus = list(main_dict[target_id].keys())
        for a, sku in enumerate(skus):
            if progress_callback: progress_callback(f"İşleniyor: {target_id} -> Ürün {a+1}/{len(skus)}", "white")
            
            payload = {
                "mainForm": "mainForm",
                'javax.faces.ViewState': javax,
                'mainForm:shipmentItems_instantSelectedRowKey': main_dict[target_id][sku]["item_id"],
                'mainForm:shipmentItems_selection': main_dict[target_id][sku]["item_id"],
                'javax.faces.partial.ajax': 'true',
                'javax.faces.source': 'mainForm:shipmentItems',
                'javax.faces.partial.execute': 'mainForm:shipmentItems',
                'javax.faces.partial.render': 'mainForm:boxContentsPanel mainForm:boxContents',
                'javax.faces.partial.event': 'rowSelect',
            }
            
            res = session.post("https://app.2dworkflow.com/items.jsf", data=payload)
            item_soup = BeautifulSoup(res.text, "lxml")
            item_tbody = item_soup.find("tbody", {"id": "mainForm:boxContents_data"})
            
            if item_tbody:
                for tr in item_tbody.findAll("tr"):
                    tds = tr.findAll("td")
                    if len(tds) > 3:
                        main_dict[target_id][sku]["date"].append(f" {tds[3].text}")
                        
            # Tekrarlayan tarihleri temizle ve dönüştür
            unique_dates = []
            for d in main_dict[target_id][sku]["date"]:
                if d not in unique_dates: unique_dates.append(d)
            main_dict[target_id][sku]["date"] = unique_dates
            main_dict[target_id][sku]["noktali"] = format_date(unique_dates[0]) if unique_dates else None

        if progress_callback: progress_callback(f"'{target_id}' Excel'e yazılıyor...", "white")
        write_excel(main_dict, target_id, output_path)

    if progress_callback: progress_callback("Combined Excel dosyası yazılıyor...", "white")
    write_combined_excel(main_dict, output_path)
    
    return {"status": "success", "message": "Expiration işlemi başarıyla tamamlandı!", "output_path": output_path}