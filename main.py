"""
This project made by HASAN ALI DURUK
Duruk/'s Software LLC
"""

from utils.gui_helpers import open_folder_in_explorer, dark_title_bar, Error_box, text_print, hata_print
from utils.file_operations import browse_directory, browse_excel, placeholder_finder, placeholder_saver, save_location_saver, path_text_function
from utils.event_handlers import on_focus_in, on_focus_out, on_click_outside
from gui.components.animated_image import AnimatedImage
from gui.components.choosers import ConvertChooser, PathAdressGroup
from gui.components.custom_buttons import MyButton, SwitchButton
from gui.components.option_menu import CustomOptionMenu
from gui.components.scrollbar import MyScrollbar
from core.cost_updater import process_costupdater, process_costupdater2

import socket
from threading import Thread
import tkinter
import tempfile
import subprocess
import signal

from screeninfo import get_monitors
from tkinter import messagebox
import os
import sys
import traceback
import pandas as pd
import numpy as np
import openpyxl
from multiprocessing.pool import ThreadPool as Pool
from multiprocessing import Pool, freeze_support


from pathlib import Path
from PIL import Image as HASAN
from PIL import ImageTk, ImageDraw

from tkinter import *
from tkinter import filedialog

from tkinter import Tk, Canvas, Entry, Text, Button, PhotoImage
import tkinter as tk
import requests
from bs4 import BeautifulSoup



from tkinterdnd2 import DND_FILES, TkinterDnD
from tkinter import ttk
from xlsxwriter import Workbook
import csv
import math
import ctypes as ct
import shutil
from openpyxl import load_workbook

import lxml
import warnings
from bs4 import XMLParsedAsHTMLWarning
import platform

def on_button_click(option_menu):
    option_menu.toggle()
def create_round_button(canvas, lenght, height, radius, text, fill_color, text_color, corners= [1, 1, 1, 1], fonksiyon=None, onenter_color='', args=(), outline_widht=1, outline_color='black', active_color='white',canvas_bg='white', active_text_color='black', font=("Helvetica", 12)):
    button_canvas = tk.Canvas(canvas, height=height+(outline_widht*2)+2, width=lenght+(outline_widht*2)+2, border=0, bg=canvas_bg, highlightthickness=0)
    if outline_widht == 0:
        outline_color = ""
    if onenter_color == '':
        onenter_color = fill_color
    if fonksiyon == None:
        def donothing():
            pass
        fonksiyon = donothing
    button_canvas.pack_propagate(False)
    x = ((outline_widht*2)+2)/2
    y = ((outline_widht*2)+2)/2
    horizontal_rectangle_lenght = lenght-(2*radius)
    vertical_rectangle_lenght = height-(2*radius)
    a = horizontal_rectangle_lenght/2
    # Buton arka planını çiz
    items = []
    if corners[0] == 1:
        oval1 =button_canvas.create_arc(x, y, x+(radius*2), y+(radius*2),start=90, extent=90,style=tk.PIESLICE, fill=fill_color, outline="")
        oval1_out =button_canvas.create_arc(x, y, x+(radius*2), y+(radius*2),start=90, extent=90,style=tk.ARC, fill=fill_color, outline=outline_color, width=outline_widht)#sol ust oval
        button_canvas.addtag_withtag("button", oval1_out)
        button_canvas.addtag_withtag("button", oval1)
        items.append(oval1)
    else:
        rectnw = button_canvas.create_rectangle(x, y, x+(radius*2), y+(radius*2), fill=fill_color, outline="")
        items.append(rectnw)
    if corners[1] == 1:
        oval2 =button_canvas.create_arc(x+lenght-(2*radius)-1, y, x+lenght-1, y+(radius*2),start=360, extent=90,style=tk.PIESLICE, fill=fill_color, outline="") #sag ust oval
        oval2_out =button_canvas.create_arc(x+lenght-(2*radius), y, x+lenght, y+(radius*2),start=360, extent=90,style=tk.ARC, fill=fill_color, outline=outline_color, width=outline_widht)
        button_canvas.addtag_withtag("button", oval2_out)
        button_canvas.addtag_withtag("button", oval2)
        items.append(oval2)
    else:
        rectne= button_canvas.create_rectangle(x+lenght-(2*radius), y, x+lenght, y+(radius*2), fill=fill_color, outline="")
        items.append(rectne)
    if corners[2] == 1:
        oval3 = button_canvas.create_arc(x, y+height, x+(radius*2), y+height-(2*radius),start=180, extent=90,style=tk.PIESLICE, fill=fill_color, outline="") #sol alt oval
        oval3_out = button_canvas.create_arc(x, y+height, x+(radius*2), y+height-(2*radius),start=180, extent=90,style=tk.ARC, fill=fill_color, outline=outline_color, width=outline_widht) #sol alt oval
        button_canvas.addtag_withtag("button", oval3_out)
        button_canvas.addtag_withtag("button", oval3)
        items.append(oval3)
    else:
        rectsw = button_canvas.create_rectangle(x, y+height+1, x+(radius*2), y+height-(2*radius),fill=fill_color, outline="")
        items.append(rectsw)
    if corners[3] == 1:
        oval4 = button_canvas.create_arc(x+lenght-(2*radius)-1, y+height, x+lenght-1, y+height-(2*radius),start=270, extent=90,style=tk.PIESLICE, fill=fill_color, outline="") #sag alt oval
        oval4_out = button_canvas.create_arc(x+lenght-(2*radius)-1, y+height, x+lenght-1, y+height-(2*radius),start=270, extent=90,style=tk.ARC, fill=fill_color, outline=outline_color, width=outline_widht) #sag alt oval
        button_canvas.addtag_withtag("button", oval4_out)
        button_canvas.addtag_withtag("button", oval4)
        items.append(oval4)
    else:
        rectse = button_canvas.create_rectangle(x+lenght-(2*radius), y+height+1, x+lenght, y+height-(2*radius),fill=fill_color, outline="")
        items.append(rectse)
    # Dikdörtgenin üst ve alt kenarlarını çiz
    rectangle = button_canvas.create_rectangle(x+radius, y, x+lenght-radius, y+height+1, outline="", fill=fill_color)
    line1 = button_canvas.create_line(x+radius, y, x+lenght-radius, y, fill=outline_color, width=outline_widht)  # Üst kenar
    line2 = button_canvas.create_line(x+radius, y+height, x+lenght-radius, y+height, fill=outline_color, width=outline_widht)  # Alt kenar

    rectangle_horizontal = button_canvas.create_rectangle(x, y+radius, x+lenght, y+height-radius, fill=fill_color, outline="")
    line3 = button_canvas.create_line(x, y+radius, x, y+height-radius, fill=outline_color, width=outline_widht) # Sol kenar
    line4 = button_canvas.create_line(x+lenght, y+radius, x+lenght, y+height-radius, fill=outline_color, width=outline_widht) # Sag kenar


    button_canvas.addtag_withtag("button", rectangle)
    button_canvas.addtag_withtag("button", line1)
    button_canvas.addtag_withtag("button", line2)
    button_canvas.addtag_withtag("button", rectangle_horizontal)
    button_canvas.addtag_withtag("button", line3)
    button_canvas.addtag_withtag("button", line4)

    items.append(rectangle)
    items.append(rectangle_horizontal)
    def on_enter(e):

        for item in button_canvas.find_all():
            if item in items:
                button_canvas.itemconfig(item, fill=onenter_color)
    def on_leave(e):
        for item in button_canvas.find_all():
            if item in items:
                button_canvas.itemconfig(item, fill=fill_color)
    def start_function(fonksiyon, args):
        return fonksiyon(*args)
    def on_button_click(e):
        start_function(fonksiyon, args)
        for i in button_canvas.find_all():
            if i in items:
                button_canvas.itemconfig(i, fill=active_color)
        button_canvas.itemconfig(text_widget, fill=active_text_color)

    def on_button_release(e):
        for i in button_canvas.find_all():
            if i in items:
                button_canvas.itemconfig(i, fill=fill_color)

        button_canvas.itemconfig(text_widget, fill=text_color)


    #button_canvas.create_rectangle(x-rectangle_lenght, y-radius, x+rectangle_lenght, y+radius, fill="lightblue", outline="black")
    # Buton metnini ekle
    text_widget = button_canvas.create_text(x+(lenght/2), y+(height/2), text=text, font=font, fill=text_color)

    button_canvas.addtag_withtag("button", text_widget)
    button_canvas.tag_bind("button", "<Button-1>", on_button_click)
    button_canvas.tag_bind("button", "<ButtonRelease-1>", on_button_release)
    button_canvas.tag_bind("button", "<Enter>", on_enter)
    button_canvas.tag_bind("button", "<Leave>", on_leave)
    return button_canvas




def color_change(old_color, new_color, window):
    global color
    widgets = window.winfo_children()
    for widget in widgets:
        if widget.cget("background") == old_color:
            widget.config(bg=new_color)
    color = new_color

def on_text_enter(event):
    canvas2.unbind_all("<MouseWheel>")
def on_text_leave(event):
    canvas2.bind_all("<MouseWheel>", on_mouse_wheel)

def show_menu(event, options, var, button, window):
    # Menü oluşturma
    menu = Menu(window, borderwidth=0, activeborderwidth=0, relief="flat", tearoff=2, background=color, fg=canvas2_text_color, activebackground=line_color, cursor="hand2")

    for option in options:
        menu.add_command(label=option, command=lambda opt=option: var.set(opt))
    menu.post(button.winfo_rootx(), button.winfo_rooty()+button.winfo_height())
def start_expration_thread(username_entry, password_entry, output_text, path, item_ids):
    t = Thread(target=expration, args=(username_entry, password_entry, output_text, path, item_ids), daemon=True)
    t.start()

def start_excel_editor_thread(ham_liste,export_liste,restock_liste,path,islem, restock_output, save_name, progress):
    t = Thread(target=rest, args=(path, ham_liste, export_liste, restock_liste, islem, progress, restock_output, save_name), daemon=True)
    t.start()


def find_column(df, possible_columns, Error):
    """
    DataFrame'de verilen olası UPC sütun adlarını arar ve ilk bulduğunu döner.
    Eğer hiçbiri bulunamazsa, None döner.
    """
    for column in possible_columns:
        if column in df.columns:
            return column

    messagebox.showerror('Error', 'Error: {}'.format(Error))
    return None
def file_reader(file):
    row_df = pd.read_excel(file, engine='openpyxl')
    return [file, row_df]
def export(liste): #liste must include (path, row, export_files, columns_dict, dataframe_dictionary)
    path = liste[0]
    row = liste[1]
    export_files = liste[2]
    columns_dict = liste[3]
    dataframe_dictionary = liste[4]

    row_code = row.split('/')[-1].split('-')[0]
    row_df = dataframe_dictionary[row]
    colrow = find_column(row_df, columns_dict['upc_sutunlari_olabilir'], f'{row} ham dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
    row_upcs = set(row_df[colrow].tolist())
    export_file = next((file for file in export_files if file.split('/')[-1].split('-')[0] == row_code), None)
    export_df = pd.read_excel(export_file, engine='openpyxl')
    colexp = find_column(export_df, columns_dict['upc_sutunlari_olabilir'], f'{row} export dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
    upcs = export_df[colexp].tolist()
    qtycol = find_column(export_df, columns_dict['quantity_sutunlari_olabilir'], f'{row} export dosyası için Quantity sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
    qtyonhand = export_df[qtycol].tolist()
    upcs_unique, idx = np.unique(upcs, return_index=True)
    qtyonhand_unique = [qtyonhand[i] for i in idx]

    qty_dict = pd.Series(qtyonhand_unique, index=upcs_unique)

    upcs_set = set(upcs)

    # silinecek değerleri belirlemek
    silinecek_degerler = row_upcs - upcs_set
    kosul = ~row_df[colrow].isin(silinecek_degerler)
    row_df = row_df[kosul]
    price_sutun = find_column(row_df, columns_dict['price_sutunlari_olabilir'], f'{row} ham dosyası için Price sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
    savename = row.split('/')[-1]
    uzunluk = row_df.shape[1]
    quantity_list = row_df[colrow].map(qty_dict).fillna('#YOK')
    try:
        price_index = row_df.columns.get_loc(price_sutun)
        row_df.insert(price_index+1, 'Qty on Hand', quantity_list, True)
    except:
        if uzunluk < 21:
            row_df.insert(uzunluk, 'Qty on Hand', quantity_list, True)
        elif uzunluk >= 21:
            row_df.insert(21, 'Qty on Hand', quantity_list, True)


    row_df.to_excel(r"{}/sonuclar/{}".format(path, savename), index=False, sheet_name='export sonuc', engine='xlsxwriter')
    return [row, row_df]
def birbirinden_dusme_remove(liste): #liste must include [file_name, remove_upc, dataframe_dictionary, path, columns_dictionary]
    file_name = liste[0]
    remove_upc = liste[1]
    dataframe_dictionary = liste[2]
    path = liste[3]
    columns_dictionary = liste[4]
    islem = liste[5]
    upc_column = find_column(dataframe_dictionary[file_name], columns_dictionary['upc_sutunlari_olabilir'], f'{file_name} dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
    kosul = ~dataframe_dictionary[file_name][upc_column].isin(remove_upc[file_name])
    dataframe_dictionary[file_name] = dataframe_dictionary[file_name][kosul]
    save_name = os.path.basename(file_name)
    save_path = os.path.join(path, 'sonuclar', save_name)
    save_path = str(save_path)
    if islem == 1:
        with pd.ExcelWriter(save_path, engine='openpyxl', mode='a') as writer:
            dataframe_dictionary[file_name].to_excel(writer, sheet_name='dusulmus liste', index=False)
    else:
        dataframe_dictionary[file_name].to_excel(save_path, sheet_name='dusulmus liste', index=False)
    return [file_name, dataframe_dictionary[file_name]]
def rest(path, row_files, export_files, restock_files, islem, progress_bar, output_text, save_name):
    def folder_creater(path):
        try:
            os.mkdir(f'{path}/sonuclar')
        except FileExistsError:
            pass

    def settings():
        with open('Settings/restock_settings.txt', 'w', encoding='utf-8') as file:
            file.write('41 cost\n'
                       '41 standart\n'
                       '45 standart\n'
                       '45 cost\n'
                       '19 cost\n'
                       '19 standart\n'
                       '27 cost\n'
                       '27 standart\n'
                       '18 cost\n'
                       '18 standart\n'
                       '01 cost\n'
                       '01 standart\n'
                       'NF\n'
                       '======================================\n'
                       "upc = UPC, upc, Upc, UPC #\n"
                       "brand = BRAND, Brand, brand\n"
                       "price = NET_AMOUNT, Price, price\n"
                       "case = CASEPACK, Size, Case, case, size\n"
                       "Quantity on hand = Qty on Hand, Quantity Available\n"
                       "pk = PK\n"
                       "======================================\n"
                       "41 cost: 0.78\n"
                       "41 standart: 0.78\n"
                       "45 cost: 0.78\n"
                       "45 standart: 0.78\n"
                       "19 cost: 0.78\n"
                       "19 standart: 0.78\n"
                       "27 cost: 1.10\n"
                       "27 standart: 1.10\n"
                       "18 cost: 1.10\n"
                       "18 standart: 1.10\n"
                       "01 cost: 1.10\n"
                       "01 standart: 1.10\n"
                       "NF: 0.78")

            file.close()

    def settings_reader():
        with open('Settings/restock_settings.txt', 'r', encoding='utf-8') as file:
            sutun_dictionary = {
                'upc_sutunlari_olabilir': [],
                'brand_sutunlari_olabilir': [],
                'price_sutunlari_olabilir': [],
                'case_sutunlari_olabilir': [],
                'quantity_sutunlari_olabilir': [],
                'pk_sutunlari_olabilir': []
            }

            satirlar = file.readlines()

            for satir in satirlar:
                ayrilmis = satir.split('=')
                if ayrilmis[0] == 'upc ' or ayrilmis[0] == 'upc':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '',1)
                        i = i.replace('\n', '')
                        sutun_dictionary['upc_sutunlari_olabilir'].append(i)
                elif ayrilmis[0] == 'brand ' or ayrilmis[0] == 'brand':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '',1)
                        i = i.replace('\n', '')
                        sutun_dictionary['brand_sutunlari_olabilir'].append(i)
                elif ayrilmis[0] == 'price ' or ayrilmis[0] == 'price':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '',1)
                        i = i.replace('\n', '')
                        sutun_dictionary['price_sutunlari_olabilir'].append(i)
                elif ayrilmis[0] == 'case ' or ayrilmis[0] == 'case':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '',1)
                        i = i.replace('\n', '')
                        sutun_dictionary['case_sutunlari_olabilir'].append(i)
                elif ayrilmis[0] == 'Quantity on hand ' or ayrilmis[0] == 'Quantity on hand':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '', 1)
                        i = i.replace('\n', '')
                        sutun_dictionary['quantity_sutunlari_olabilir'].append(i)
                elif ayrilmis[0] == 'pk' or ayrilmis[0] == 'pk ':
                    for i in ayrilmis[1].split(','):
                        i = i.replace(' ', '', 1)
                        i = i.replace('\n', '')
                        sutun_dictionary['pk_sutunlari_olabilir'].append(i)
            a = 0
            maliyet_dict = {}
            for satir in satirlar:
                if '=====' in satir:
                    a += 1
                    continue
                if a == 1:
                    satir = satir.split(':')
                    satir[1] = satir[1].replace(' ', '')
                    maliyet_dict[satir[0]] = float(satir[1].replace('\n', ''))
        return [sutun_dictionary, maliyet_dict]

    #def export()

    def birbirinden_dusme(row_files, dataframe_dictionary, columns_dict):
        remove_upc = {file: [] for file in row_files}
        for i, file in enumerate(row_files):
            this_file_df = dataframe_dictionary[file]
            upc_column = find_column(this_file_df, columns_dict['upc_sutunlari_olabilir'], f'{file} dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            price_column = find_column(this_file_df, columns_dict['price_sutunlari_olabilir'], f'{file} dosyası için Price sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            this_file_upc = this_file_df.set_index(upc_column)[price_column].to_dict()
            a = i+1
            while a < len(row_files):
                next_file_name = row_files[a]
                next_file_df = dataframe_dictionary[next_file_name]
                upc_column = find_column(next_file_df, columns_dict['upc_sutunlari_olabilir'], f'{next_file_name} dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
                price_column = find_column(next_file_df, columns_dict['price_sutunlari_olabilir'], f'{next_file_name} dosyası için Price sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
                next_file_upc = next_file_df.set_index(upc_column)[price_column].to_dict()
                for upc in this_file_upc.keys():
                    if upc in next_file_upc.keys():
                        if this_file_upc[upc] < next_file_upc[upc]:
                            remove_upc[next_file_name].append(upc)
                        elif this_file_upc[upc] > next_file_upc[upc]:
                            remove_upc[file].append(upc)
                        elif this_file_upc[upc] == next_file_upc[upc]:
                            remove_upc[next_file_name].append(upc)
                a+=1
        return remove_upc

    #def birbirinden_dusme_remove()

    def restock(path, row_dataframe_dictionary, export_dataframe_dictionary, file_names, main_excel, columns_dict, maliyet_dict, save_name):
        yazilacak_dictionary = {}
        main_excel_df = pd.read_excel(main_excel, engine='openpyxl')
        lenght = main_excel_df.shape[1]
        main_upc_col = find_column(main_excel_df, columns_dict['upc_sutunlari_olabilir'], f'{main_excel} dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
        main_upc_list = main_excel_df[main_upc_col].tolist()
        main_pk_col = find_column(main_excel_df, columns_dict['pk_sutunlari_olabilir'], f'{main_excel} dosyası için PK sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
        main_pk_list = main_excel_df[main_pk_col].tolist()
        main_dict = {}
        for i, upc in enumerate(main_upc_list):
            main_dict[i] = {}
            main_dict[i]['upc'] = main_upc_list[i]
            main_dict[i]['brand'] = '#YOK'
            main_dict[i]['suplier'] = '#YOK'
            main_dict[i]['price'] = '#YOK'
            main_dict[i]['case'] = '#YOK'
            main_dict[i]['qtyonhand'] = '#YOK'
            main_dict[i]['PK'] = main_pk_list[i]
            main_dict[i]['maliyet'] = '#YOK'
        progress_bar['maximum'] = len(row_files)
        progress_bar['value'] = 0
        for i, file in enumerate(file_names):
            row_upc_col = find_column(row_dataframe_dictionary[file], columns_dict['upc_sutunlari_olabilir'], f'{file} ham dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            row_case_col = find_column(row_dataframe_dictionary[file], columns_dict['case_sutunlari_olabilir'], f'{file} ham dosyası için Case sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            row_quantity_col = find_column(row_dataframe_dictionary[file], columns_dict['quantity_sutunlari_olabilir'], f'{file} ham dosyası için Quantity sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            export_upc_col = find_column(export_dataframe_dictionary[file], columns_dict['upc_sutunlari_olabilir'], f'{file} export dosyası için UPC sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            export_price_col = find_column(export_dataframe_dictionary[file], columns_dict['price_sutunlari_olabilir'], f'{file} export dosyası için Price sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            export_brand_col = find_column(export_dataframe_dictionary[file], columns_dict['brand_sutunlari_olabilir'], f'{file} export dosyası için Brand sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')
            export_quantity_col = find_column(export_dataframe_dictionary[file], columns_dict['quantity_sutunlari_olabilir'], f'{file} export dosyası için Quantity sütunu bulunamadı, sütunların isimlerini kontrol edip tekrar deneyiniz!')

            row_upc_list = row_dataframe_dictionary[file][row_upc_col].tolist()
            row_case_list = row_dataframe_dictionary[file][row_case_col].tolist()
            row_quantity_list = row_dataframe_dictionary[file][row_quantity_col].tolist()
            export_upc_list = export_dataframe_dictionary[file][export_upc_col].tolist()
            export_price_list = export_dataframe_dictionary[file][export_price_col].tolist()
            export_brand_list = export_dataframe_dictionary[file][export_brand_col].tolist()
            export_quantity_list = export_dataframe_dictionary[file][export_quantity_col].tolist()

            export_dict = {}
            for i, upc in enumerate(export_upc_list):
                export_dict[upc] = {}
                export_dict[upc]['price'] = export_price_list[i]
                export_dict[upc]['quantity'] = export_quantity_list[i]
                export_dict[upc]['brand'] = export_brand_list[i]
            row_dict = {}
            for i, upc in enumerate(row_upc_list):
                row_dict[upc] = {}
                row_dict[upc]['case'] = row_case_list[i]
                row_dict[upc]['quantity'] = row_quantity_list[i]
            yazilacak_dictionary[file] = {
                'price': [],
                'quantity': []
            }
            for index in main_dict.keys():
                upc = main_dict[index]['upc']
                if upc in export_upc_list:
                    x = True
                    if main_dict[index].keys() != []:
                        for key in main_dict[index].keys():
                            if key.endswith('.xlsx'):
                                if main_dict[index][key]['price'] != '#YOK':
                                    if main_dict[index][key]['price'] > export_dict[upc]['price']:
                                        pass
                                    elif main_dict[index][key]['price'] < export_dict[upc]['price']:
                                        x = False
                                        break
                    main_dict[index][file] = {}
                    main_dict[index][file]['quantity'] = export_dict[upc]['quantity']
                    main_dict[index][file]['price'] = export_dict[upc]['price']
                    main_dict[index]['brand'] = export_dict[upc]['brand']
                    if x == True:
                        main_dict[index]['price'] = export_dict[upc]['price']
                else:
                    main_dict[index][file] = {}
                    main_dict[index][file]['quantity'] = '#YOK'
                    main_dict[index][file]['price'] = '#YOK'
                yazilacak_dictionary[file]['price'].append(main_dict[index][file]['price'])
                yazilacak_dictionary[file]['quantity'].append(main_dict[index][file]['quantity'])
                filesplit = file.split('/')[-1].split('-')
                filename = filesplit[0]
                if upc in row_upc_list:
                    main_dict[index]['suplier'] = filename
                    main_dict[index]['case'] = row_dict[upc]['case']
                    main_dict[index]['qtyonhand'] = row_dict[upc]['quantity']
            progress_bar['value'] = i+1
        text_print(output_text, 'Restock dosyası kaydediliyor...')
        progress_bar['maximum'] = 1
        progress_bar['value'] = 0
        brand_list = []
        suplier_list = []
        price_list = []
        case_list = []
        quantity_list = []
        maliyet_list = []
        for index in main_dict.keys():
            brand_list.append(main_dict[index]['brand'])
            suplier_list.append(main_dict[index]['suplier'])
            price_list.append(main_dict[index]['price'])
            case_list.append(main_dict[index]['case'])
            quantity_list.append(main_dict[index]['qtyonhand'])
            if main_dict[index]['PK'] != '#YOK' and main_dict[index]['price'] != '#YOK':
                try:
                    pk = int(main_dict[index]['PK'].replace('PK', ''))
                    maliyet_list.append((pk * float(main_dict[index]['price'])) + float(maliyet_dict[main_dict[index]['suplier']]))
                except:
                    maliyet_list.append(main_dict[index]['price'])
            else:
                maliyet_list.append(main_dict[index]['price'])


        main_excel_df.insert(lenght, 'Brand', brand_list, True)
        main_excel_df.insert(lenght+1, 'Price', price_list, True)
        main_excel_df.insert(lenght+2, 'Maliyet', maliyet_list, True)
        main_excel_df.insert(lenght+3, 'Case', case_list, True)
        a = 4
        for file in yazilacak_dictionary.keys():
            filesplit = file.split('/')[-1].split('-')
            filename = filesplit[0]
            main_excel_df.insert(lenght+a, filename + ' price', yazilacak_dictionary[file]['price'], True)
            a+=1
        main_excel_df.insert(lenght+a, 'Qty on Hand', quantity_list, True)
        a = 5
        for file in yazilacak_dictionary.keys():
            filesplit = file.split('/')[-1].split('-')
            filename = filesplit[0]

            main_excel_df.insert(lenght + len(yazilacak_dictionary.keys())+a, filename + ' quantity', yazilacak_dictionary[file]['quantity'], True)
            a+=1
        main_excel_df.insert(lenght + len(yazilacak_dictionary.keys())+a, 'suplier', suplier_list, True)
        try:
            silme_kosul = ~main_excel_df['Price'].isin(['#YOK', '#YOK'])
            main_excel_df = main_excel_df[silme_kosul]
        except:
            pass

        try:
            os.mkdir('{}/restock'.format(path))
        except FileExistsError:
            pass

        main_excel_df.to_excel('{}/restock/{}.xlsx'.format(path, save_name), index=False, sheet_name='restock', engine='xlsxwriter')
        progress_bar['value'] = 1
        return main_upc_list

    def main():
        folder_creater(path)

        dataframe_dictionary = {}
        if 'restock_settings.txt' not in os.listdir('Settings'):
            settings()
        sets = settings_reader()
        columns_dict = sets[0]
        maliyet_dict = sets[1]

        text_print(output_text, 'Ham dosyalar okunuyor...')

        ###    DOSYALARI OKUMA    ###
        progress_bar['maximum'] = len(row_files)
        progress_bar['value'] = 0
        p0 = Pool()
        for i, a in enumerate(p0.imap_unordered(file_reader, row_files)):
            dataframe_dictionary[a[0]] = a[1]
            progress_bar['value'] = i + 1
        p0.close()
        p0.join()
        ###################################################

        if islem['export'] == 1:
            text_print(output_text, 'Export işlemi yapılıyor...')
            ###    EXPORT POOL    ###
            export_multiprocess_list = []
            for row in row_files:
                tmp = [path, row, export_files, columns_dict, dataframe_dictionary]
                export_multiprocess_list.append(tmp)

            progress_bar['maximum'] = len(export_multiprocess_list)
            progress_bar['value'] = 0
            p1 = Pool()
            for i, a in enumerate(p1.imap_unordered(export, export_multiprocess_list)):
                dataframe_dictionary[a[0]] = a[1]
                progress_bar['value'] = i + 1
            p1.close()
            p1.join()
            ################################################

        #Silinecek UPC degerleri
        text_print(output_text, 'Silinecek UPC değerleri tespit ediliyor...')
        remove_upc = birbirinden_dusme(row_files, dataframe_dictionary, columns_dict)
        ################################################


        ###    BIRBIRINDEN DUSME POOL    ###
        text_print(output_text, 'Tespit edilen UPC değerleri siliniyor...')
        row_dataframe_dictionary = {}
        dusme_multiprocess_list = []
        for row in row_files:
            tmp = [row, remove_upc, dataframe_dictionary, path, columns_dict, islem['export']]
            dusme_multiprocess_list.append(tmp)

        progress_bar['maximum'] = len(dusme_multiprocess_list)
        progress_bar['value'] = 0
        p2 = Pool()
        for i, a in enumerate(p2.imap_unordered(birbirinden_dusme_remove, dusme_multiprocess_list)):
            row_dataframe_dictionary[a[0]] = a[1]
            progress_bar['value'] = i + 1
        p2.close()
        p2.join()
        ###################################################
        if islem['restock'] == 1:
            text_print(output_text, 'Restock işlemi yapılıyor...')
            main_excel = restock_files[0]
            restock(path, row_dataframe_dictionary, dataframe_dictionary, row_files, main_excel, columns_dict, maliyet_dict, save_name)
    try:
        main()
        text_print(output_text, 'İşlem başarıyla tamamlandı!', color='#90EE90')
    except:
        text_print(output_text, 'Beklenmeyen bir hata meydana geldi!', color='red')
        text_print(output_text, traceback.format_exc(), color='red')

def expration(username, password, tex: tkinter.Text, path, item_ids):
    def settings_creater():
        with open("Settings/expration_settings.txt", "w", encoding='utf-8') as file:
            file.write('login_button_id = mainForm:j_idt23, mainForm:j_idt13, mainForm:j_idt22\n'
                       'email_id = mainForm:email\n'
                       'password_id = mainForm:password\n'
                       'default_email = sales@buyable.net\n'
                       'default_password = hasali2603\n')
            file.close()
    def settings_reader():
        dictionary = {
            'login_button_id': [],
            'email_id': [],
            'password_id': [],
            'default_email': [],
            'default_password': [],
        }
        with open("Settings/expration_settings.txt", "r", encoding='utf-8') as file:
            lines = file.readlines()
            for line in lines:
                line = line.split('=')
                if line[0] == 'login_button_id' or line[0] == 'login_button_id ':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace('\n', '')
                        deger = deger.replace(' ', '', 1)
                        dictionary['login_button_id'].append(deger)
                if line[0] == 'email_id' or line[0] == 'email_id ':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace('\n', '')
                        deger = deger.replace(' ', '', 1)
                        dictionary['email_id'].append(deger)
                if line[0] == 'password_id' or line[0] == 'password_id ':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace('\n', '')
                        deger = deger.replace(' ', '', 1)
                        dictionary['password_id'].append(deger)
                if line[0] == 'default_email' or line[0] == 'default_email ':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace('\n', '')
                        deger = deger.replace(' ', '', 1)
                        dictionary['default_email'].append(deger)
                if line[0] == 'default_password' or line[0] == 'default_password ':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace('\n', '')
                        deger = deger.replace(' ', '', 1)
                        dictionary['default_password'].append(deger)
        return dictionary

    def login():

        url = "https://app.2dworkflow.com/login.jsf"
        session = requests.Session()
        response = session.get(url)
        main_dict = {}
        soup = BeautifulSoup(response.text, 'html.parser')
        javax = soup.find('input', {'name': 'javax.faces.ViewState'})['value']
        button = soup.find('button')['name']
        payload = {'mainForm:email': username,
                   'mainForm:password': password,
                   'mainForm': 'mainForm',
                   'javax.faces.ViewState': javax,
                   button: ''}
        response = session.post(url, data=payload)
        if response.status_code == 200:
            text_print(tex, "Login Successful")
            id_list = []
            x = item_ids.split(',')
            for id in x:
                id = id.replace(' ', '')
                id_list.append(id)

            id_dict = {}
            shipments_url = "https://app.2dworkflow.com/shipped.jsf"
            response = session.get(shipments_url)
            soup = BeautifulSoup(response.text, 'html.parser')
            tbody = soup.find("tbody", id="mainForm:shipments_data")
            fba_date = ""
            for id in id_list:
                main_dict[id] = {}
                href = ""


                if tbody:
                    trs = tbody.findAll("tr")
                    if trs:
                        for index, tr in enumerate(trs):
                            a = tr.find("a")
                            if a and id in a.get("title", ""):
                                href = a.get("href", "")
                                fba_date = trs[index-1].findAll("span")[2].text.split(",")[0]
                                print(fba_date)
                id_dict[id] = href
            for id in id_list:
                url_payload = 'https://app.2dworkflow.com/'
                response = session.get(f"{url_payload}{id_dict[id]}")
                print(id, f"{url_payload}{id_dict[id]}")
                soup = BeautifulSoup(response.text, 'html.parser')
                tbody = soup.find("tbody", {"id": "mainForm:shipmentItems_data"})
                tbody_info = soup.find("tbody", {"id": "mainForm:shipmentInfo_data"})
                info_tr = tbody_info.find("tr")
                trler = tbody.findAll("tr")
                shipment_name = ""
                tds = info_tr.findAll("td")
                shipment_name = tds[3].text
                print(shipment_name)

                for tr in trler:
                    sku = tr.find("span").text
                    main_dict[id][sku] = {}

                    main_dict[id][sku]["item_id"] = tr['data-rk']
                    main_dict[id][sku]["shipped"] = tr.findAll("td")[3].text
                    main_dict[id][sku]["created"] = fba_date
                    main_dict[id][sku]["shipment_name"] = shipment_name
                text_print(tex, f"{len(list(main_dict[id].keys()))} adet urun bulundu!")
                javax = soup.find('input', {'name': 'javax.faces.ViewState'})['value']

                for a, sku in enumerate(list(main_dict[id].keys())):
                    main_dict[id][sku]["date"] = []
                    text_print(tex, f"({id_list.index(id) + 1} / {len(id_list)}) {id}: {str(a+1)} / {str(len(list(main_dict[id].keys())))}")

                    payload = {
                        "mainForm": "mainForm",
                        'javax.faces.ViewState': javax,
                        'mainForm:shipmentItems_instantSelectedRowKey': main_dict[id][sku]["item_id"],
                        'mainForm:shipmentItems_selection': main_dict[id][sku]["item_id"],
                        'javax.faces.partial.ajax': 'true',
                        'javax.faces.source': 'mainForm:shipmentItems',
                        'javax.faces.partial.execute': 'mainForm:shipmentItems',
                        'javax.faces.partial.render': 'mainForm:boxContentsPanel mainForm:boxContents',
                        'javax.faces.partial.event': 'rowSelect',
                    }
                    response = session.post("https://app.2dworkflow.com/items.jsf", data=payload)
                    soup = BeautifulSoup(response.text, "lxml")
                    tbody = soup.find("tbody", {"id": "mainForm:boxContents_data"})
                    trler = tbody.findAll("tr")
                    if trler != None:
                        for tr in trler:
                            tdler = tr.findAll("td")
                            if tdler != None and len(tdler) > 3:
                                main_dict[id][sku]["date"].append(f" {tdler[3].text}")
                date_converter(main_dict, id)
                writer(main_dict, id)
            print("combined.xlsx dosyasi yazdiriliyor...")
            combined_writer(main_dict)
            print("combined.xlsx dosyasi basariyla yazdirildi")
            return main_dict
        else:
            print('giris basarisiz')
            return None

    def date_converter(main_dict, id):
        for sku in main_dict[id].keys():
            empty_list = []
            for date in main_dict[id][sku]["date"]:
                if date not in empty_list:
                    empty_list.append(date)
            main_dict[id][sku]["date"] = empty_list
            noktalidate = ''
            try:
                date = main_dict[id][sku]["date"][0]
                if '-' in date:
                    a = date.replace(' ', '')
                    x = a.split('-')
                    noktalidate = (x[1] + '.' + x[0] + '.' + x[2])
                elif '/' in date:
                    a = date.replace(' ', '')
                    x = a.split('/')
                    noktalidate = (x[1] + '.' + x[0] + '.' + x[2])
                elif '.' in date:
                    a = date.replace(' ', '')
                    x = a.split('.')
                    noktalidate = (x[1] + '.' + x[0] + '.' + x[2])
                elif date == []:
                    noktalidate = None
            except:
                noktalidate = None
            main_dict[id][sku]["noktali"] = noktalidate
    def writer(main_dictionary, id):
        save_name = f"{id}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = save_name
        a = 2
        ws.cell(row=1, column=1).value = 'NAME'
        ws.cell(row=1, column=2).value = 'SHIPMENT ID'
        ws.cell(row=1, column=3).value = 'SHIPMENT DATE'
        ws.cell(row=1, column=4).value = 'SKU'
        ws.cell(row=1, column=5).value = 'SHIPPED'
        ws.cell(row=1, column=6).value = 'DATE'
        ws.cell(row=1, column=7).value = 'TR DATE'

        for sku in main_dictionary[id].keys():
            ws.cell(row=a, column=1).value = main_dictionary[id][sku]["shipment_name"]
            ws.cell(row=a, column=2).value = id
            ws.cell(row=a, column=3).value = main_dictionary[id][sku]["created"]
            ws.cell(row=a, column=4).value = sku
            ws.cell(row=a, column=5).value = main_dictionary[id][sku]["shipped"]
            ws.cell(row=a, column=6).value = main_dictionary[id][sku]["date"][0]
            ws.cell(row=a, column=7).value = main_dictionary[id][sku]["noktali"]
            c = 8
            for date in main_dictionary[id][sku]["date"][1:]:
                ws.cell(row=a, column=c).value = str(date)
                c = c + 1
            a = a + 1

        wb.save(f"{path}/{save_name}")
        text_print(tex, f"{id} {save_name} olarak belirtilen dizine kaydedildi!")


    def combined_writer(main_dictionary):
        save_name = f"combined.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = save_name

        a = 2
        for id in main_dictionary.keys():
            ws.cell(row=1, column=1).value = 'NAME'
            ws.cell(row=1, column=2).value = 'SHIPMENT ID'
            ws.cell(row=1, column=3).value = 'SHIPMENT DATE'
            ws.cell(row=1, column=4).value = 'SKU'
            ws.cell(row=1, column=5).value = 'SHIPPED'
            ws.cell(row=1, column=6).value = 'DATE'
            ws.cell(row=1, column=7).value = 'TR DATE'

            for sku in main_dictionary[id].keys():
                ws.cell(row=a, column=1).value = main_dictionary[id][sku]["shipment_name"]
                ws.cell(row=a, column=2).value = id
                ws.cell(row=a, column=3).value = main_dictionary[id][sku]["created"]
                ws.cell(row=a, column=4).value = sku
                ws.cell(row=a, column=5).value = main_dictionary[id][sku]["shipped"]
                ws.cell(row=a, column=6).value = main_dictionary[id][sku]["date"][0]
                ws.cell(row=a, column=7).value = main_dictionary[id][sku]["noktali"]
                c = 8
                for date in main_dictionary[id][sku]["date"][1:]:
                    ws.cell(row=a, column=c).value = str(date)
                    c = c + 1
                a = a + 1

        wb.save(f"{path}/{save_name}")
        text_print(tex, f"{id} {save_name} olarak belirtilen dizine kaydedildi!")

    def main():
        try:
            warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
            if 'expration_settings.txt' not in os.listdir('Settings'):
                settings_creater()
            dictionary = settings_reader()
            text_print(tex, str(dictionary))
            main_dict = login()
            if main_dict:
                text_print(tex, "islem basariyla tamamlandi.", "green")
                open_folder_in_explorer(path)
            else:
                text_print(tex, "bir sorun meydana geldi!")
        except:
            text_print(tex, "bir hata meydana geldi! Hata kodu: ")
            e = traceback.format_exc()
            text_print(tex, e, "red")
    main()

def relative_to_assets(path: str) -> Path:
    return ASSETS_PATH / path

def silici():
    items = canvas2.find_all()

    # Her bir öğeyi yok eder (Sadece Canvas üzerindeki şekiller)
    for item in items:
        canvas2.delete(item)

    for widget in canvas2.winfo_children():
        widget.destroy()

    except_list = [canvas, canvas2]
    for widget in window.winfo_children():
        if widget not in except_list:
            widget.destroy()

def width_f(widtha):
    deneme_label = Label(canvas2, text="0", bg='black', border=0)
    deneme_label.place(x=50000, y=0)
    deneme_label.update()
    a = deneme_label.winfo_width()
    deneme_label.destroy()
    widtha = round(widtha / a)
    return widtha










def settings(isim, settings_var):
    with open(isim, 'w', encoding='utf-8') as file:
        file.write(settings_var)

        file.close()


def button_hover(event, button):
    if dictionary[button] == 0 and button != button_5:
        button.config(background='#3C4043', image=program_icon_hover)
    elif dictionary[button] == 0 and button == button_5:
        button.config(background='#3C4043', image=home_icon_hover)


def button_leave(event, button):
    if dictionary[button] == 0 and button != button_5:
        button.config(background=color, image=program_icon_notselected)
    elif dictionary[button] == 0 and button == button_5:
        button.config(background=color, image=home_icon_notselected)


def smooth_scroll(delta, canvas, count):
    """Mouse wheel scroll hareketini smooth hale getirir."""
    if delta > 0:
        canvas.yview_scroll(-1, "units")  # Yukarı kaydırma
    else:
        canvas.yview_scroll(1, "units")  # Aşağı kaydırma
    count-=1
    if count >=0:
        window.after(10, lambda: smooth_scroll(delta, canvas, count))  # 10 ms sonra tekrar çağır
def on_mouse_wheel(event):
    canvas2.yview_scroll(int(-1*(event.delta/120)), "units")


def tsv_script(path, output_text):
    def xlsx_converter():
        try:
            os.mkdir(path + '/excel')
        except:
            pass
        text_print(output_text, 'ayarlar çekiliyor...')
        settings = open('Settings/tsv_settings.txt', 'r')
        columns = []
        for line in settings.readlines():
            line = line.split('=')
            line = line[1]
            for i in line.split(','):
                i = i.replace(' ', '')
                columns.append(i)
        text_print(output_text, 'columns= {}'.format(columns))

        dosyalar = dosyalar_dictionary['tsv']
        if dosyalar == []:
            text_print(output_text, 'hicbir dosya saglanilmadigi algilanildi lutfen tsv dosyalarinizi belirtilen yere surukleyip birakiniz ardindan yeniden deneyiniz.')
            sys.exit()
        for dosya in dosyalar:
            if dosya.endswith('.tsv'):
                dosya_name = dosya.split('/')
                dosya_name = dosya_name[-1]
                workbook = Workbook('{}/excel/{}'.format(path, dosya_name.replace('.tsv', '') + '.xlsx'))
                worksheet = workbook.add_worksheet()
                tsv_reader = csv.reader(open(dosya, 'r'), delimiter='\t')
                a = 0
                row1 = 0
                for row, data in enumerate(tsv_reader):
                    if a != 1:                         #'Merchant SKU' in data or 'Title' in data or 'ASIN' in data or 'FNSKU' in data:
                        for col in columns:
                            if col in data:
                                a = 1
                    if a == 1:
                        worksheet.write_row(row1, 0, data)
                        row1 += 1
                workbook.close()
    def compare_and_write():
        dosyalar = os.listdir(path + "/excel")
        new_path = path + '/excel/'
        dictionary = {}
        for dosya in dosyalar:
            if dosya.endswith('.xlsx'):
                text_print(output_text, dosya)
                df = pd.read_excel(new_path + dosya)
                skular= df['Merchant SKU'].tolist()
                shipped = df['Shipped'].tolist()
                index = 0
                a = 0
                for sku in skular:
                    try:
                        a = dictionary[sku]
                    except:
                        dictionary[sku] = 0
                for sku in skular:
                    dictionary[sku] = dictionary[sku] + shipped[index]
                    index += 1
        excel_dictionary = {
            'Merchant SKU': dictionary.keys(),
            'Shipped': dictionary.values()
        }
        son = pd.DataFrame(excel_dictionary)
        son.to_excel(path + '/son.xlsx', index=False)
    def settings():
        with open('Settings/tsv_settings.txt', 'w', encoding='utf-8') as file:
            file.write('columns = Merchant SKU, Title, ASIN, FNSKU, external-id, Condition, Shipped')
            file.close()
    def main():
        try:
            if 'tsv_settings.txt' not in os.listdir('Settings'):
                settings()
            xlsx_converter()
            compare_and_write()
            text_print(output_text, 'Operasyon başarıyla tamamlandı!')
            open_folder_in_explorer(path)
            sys.exit()
        except Exception as e:
            text_print(output_text, 'bir hata meydana geldi...')
            e = traceback.format_exc()
            text_print(output_text, e, color='red')
            sys.exit()
    main()

def converter_script(path: str, output_text: tk.Text, input_type: str, output_type: str):
    class ConvertPdf:
        def __init__(self, images, save_path, combine=False):
            self.images = images
            self.save_path = save_path
            self.combine = combine
        def converter(self, images, save_path, combine):
            image_list = []
            if combine:
                for image in images:
                    image_var = HASAN.open(image)
                    rgb = image_var.convert('RGB')
                    image_list.append(rgb)
                image_list[0].save(save_path, save_all=True, append_images=image_list[1:])
            else:
                image_var = HASAN.open(images)
                rgb = image_var.convert('RGB')
                rgb.save(save_path)



    objects = {
        'csv': pd,
        'xlsx': pd,
        'txt': pd
    }

    read_functions = {
        'csv': 'read_csv',
        'xlsx': 'read_excel',
        'txt': 'read_table'
    }
    write_functions = {
        'csv': 'to_csv',
        'xlsx': 'to_excel',
        'txt': 'to_csv'
    }

    read_function = getattr(objects[input_type], read_functions[input_type])
    def dir_creater():
        try:
            os.mkdir(path + '/sonuc dosyalari')
        except FileExistsError:
            pass
    def noktavirgul(df):
        df = df.map(lambda x: str(x).replace('.', ',') if isinstance(x, (str, float)) and x != 0 and x != 0.0 else x)
        return df
    def main():
        try:
            dir_creater()
            for file in dosyalar_dictionary['convert']:
                if file.endswith('.'+input_type):
                    if input_type == 'txt':
                        df = read_function(file, encoding="latin-1", dtype=str)
                        df = df.fillna("")
                    else:
                        df = read_function(file, dtype=str)
                        df = df.fillna("")
                    if input_type == 'csv':
                        df = noktavirgul(df)
                    write_function = getattr(df, write_functions[output_type])
                    save_name = file.split('/')
                    save_name = save_name[-1]
                    if output_type == 'txt':
                        write_function(path + '/sonuc dosyalari/'+save_name.replace("."+input_type, '.'+output_type), sep='\t', index=False, na_rep='')
                    else:
                        write_function(path + '/sonuc dosyalari/'+save_name.replace("."+input_type, '.'+output_type), index=False, na_rep='')
                    text_print(output_text, '{} dosyasi cevrildi'.format(save_name))
            text_print(output_text, 'cevirme islemi tamamlandi!')
            open_folder_in_explorer(path)
            sys.exit()
        except Exception:
            text_print(output_text, 'bir hata meydana geldi. Hata kodu:')
            text_print(output_text, traceback.format_exc(), color='red')
            sys.exit()
    main()


def future_price_script(path, name, restock_excel, future_excel, output_text):
    def restock_reader():
        restock_df = pd.read_excel(restock_excel)
        restock_dictionary = {}
        price_columns_list = []
        all_asins = restock_df['ASIN'].tolist()

        for i in restock_df.columns:
            if 'price' in i:
                price_columns_list.append(i)
        for i, asin in enumerate(all_asins):
            restock_dictionary[asin] = {}
            restock_dictionary[asin]['Price'] = restock_df['Price'][i]
            restock_dictionary[asin]['Maliyet'] = restock_df['Maliyet'][i]
            for name in price_columns_list:
                restock_dictionary[asin][name] = restock_df[name][i]
        return [restock_dictionary, restock_df, price_columns_list]
    def future_reader(restock_dictionary, restock_df: pd.DataFrame, restock_price_columns_list):
        future_df = pd.read_excel(future_excel)
        future_dictionary = {}
        price_columns_list = []
        future_name_list = []
        all_asins = future_df['ASIN'].tolist()
        for i in future_df.columns:
            if 'price' in i:
                price_columns_list.append(i)
        for i, asin in enumerate(all_asins):
            future_dictionary[asin] = {}
            future_dictionary[asin]['Price'] = future_df['Price'][i]
            future_dictionary[asin]['Maliyet'] = future_df['Maliyet'][i]
            for name in price_columns_list:
                future_dictionary[asin][name] = future_df[name][i]
        for asin in restock_dictionary.keys():
            if asin in future_dictionary.keys():
                restock_dictionary[asin]['Future Price'] = future_dictionary[asin]['Price']
                restock_dictionary[asin]['Future Maliyet'] = future_dictionary[asin]['Maliyet']
                for name in price_columns_list:
                    if name in restock_dictionary[asin].keys():
                        future_name = name.replace('price', 'future price')
                        restock_dictionary[asin][future_name] = future_dictionary[asin][name]
                    else:
                        future_name = name.replace('price', 'future price')
                        restock_dictionary[asin][future_name] = future_dictionary[asin][name]
            else:
                restock_dictionary[asin]['Future Price'] = '#YOK'
                restock_dictionary[asin]['Future Maliyet'] = '#YOK'
                for name in price_columns_list:
                    future_name = name.replace('price', 'future price')
                    restock_dictionary[asin][future_name] = '#YOK'
        future_price_list = []
        future_maliyet_list = []
        write_dict = {}
        for name in price_columns_list:
            future_name = name.replace('price', 'future price')
            write_dict[future_name] = []
        for i, asin in enumerate(restock_df['ASIN'].tolist()):
            future_price_list.append(restock_dictionary[asin]['Future Price'])
            future_maliyet_list.append(restock_dictionary[asin]['Future Maliyet'])
            for name in price_columns_list:
                future_name = name.replace('price', 'future price')
                write_dict[future_name].append([i, restock_dictionary[asin][future_name]])

        column_index = restock_df.columns.get_loc('Maliyet') + 1
        restock_df.insert(column_index, 'Future Price', future_price_list, allow_duplicates = False)
        restock_df.insert(column_index + 1, 'Future Maliyet', future_maliyet_list, allow_duplicates = False)
        dont_exist_list = []
        for future_name in write_dict.keys():
            liste = []
            for price in write_dict[future_name]:
                liste.append(price[1])
            name = future_name.replace('future price', 'price')
            try:
                column_index = restock_df.columns.get_loc(name) + 1
                restock_df.insert(column_index, future_name, liste, allow_duplicates = False)
            except:
                dont_exist_list.append(future_name)
        price_columns_indices = [restock_df.columns.get_loc(col) for col in restock_df.columns if 'price' in col]
        max_index = max(price_columns_indices)
        for name in dont_exist_list:
            liste = []
            for price in write_dict[name]:
                liste.append(price[1])
            max_index += 1
            restock_df.insert(max_index, name, liste, allow_duplicates = False)
        return restock_df
    def writer(path, name, df:pd.DataFrame):
        df.to_excel(f'{path}/{name}.xlsx', index=False)

    def main():
        text_print(output_text, 'restock dosyasi okunuyor')
        restock_return = restock_reader()
        restock_dictionary = restock_return[0]
        restock_df = restock_return[1]
        price_columns_list = restock_return[2]
        text_print(output_text, 'future price dosyasi okunuyor ve gerekli islemler yapiliyor')
        restock_df = future_reader(restock_dictionary, restock_df, price_columns_list)
        text_print(output_text, 'sonuc dosyasi istenilen konuma yazdiriliyor')
        writer(path, name, restock_df)
        text_print(output_text, 'İşlem başarıyla tamamlandı!', color='green')
        open_folder_in_explorer(path)

    main()

def future_price_window(future_price_button):
    canvas2.unbind_all("<MouseWheel>")
    def color_change(e,c,t, b):
        b.config(background=c, text_color=t)
    color_change(1,'#8AB4F8','black', future_price_button)
    f_window = Toplevel(window)
    f_window.title('Future Price')
    f_window.geometry('1000x860')
    f_window.config(bg=color)
    #f_window = TkinterDnD.Tk()
    try:
        f_window.iconbitmap('assets/icon.ico')
    except:pass
    content_canvas = Canvas(
        f_window,
        bg=color,
        border=0,
        highlightthickness=0,
    )
    content_canvas.place(x=0,y=0)
    content_canvas.grid_columnconfigure(0, weight=1)
    top_frame = Frame(
        content_canvas,
        background=color,
    )
    bottom_frame =Frame(
        content_canvas,
        background=color,
    )
    title = Label(
        top_frame,
        fg=canvas2_text_color,
        bg=color,
        text='Future Price',
        font=("JetBrainsMonoRoman Regular", 24*-1),
    )
    title_line = Frame(
        top_frame,
        bg=line_color,
        height=2
    )
    save_path_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        top_frame,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_name_label = Label(
        top_frame,
        text= 'Aşağıya sonucun kaydedilmesini istediginiz ismi giriniz:',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    save_name_text = Text(
        top_frame,
        height=1,
        border=0,
        fg=canvas2_text_color,
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')
    baslat_button = MyButton(
        bottom_frame,
        round=12,
        width=100,
        height=40,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    output_text = Text(
        f_window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )


    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=f_window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('ftr', save_path)
    def baslat_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        path = save_path.get(1.0, tk.END)
        name = save_name_text.get(1.0, tk.END)
        path = path.strip('\n')
        name = name.strip('\n')
        output(path, name)
    def future_price_script_starter(path, name, future_restock, future_future, output_text):
        t = Thread(target=future_price_script, args=(path, name, future_restock, future_future, output_text), daemon=True)
        t.start()
    def output(path, name):
        output_text.pack(side=BOTTOM, fill=X, anchor='w')
        if path == "Example: C:/Users/Username/Desktop/sonuc":
            text_print(output_text, "Maalesef path degeri algilanamadi! Dogru bir deger girdiginizden emin olup tekrar deneyiniz.")

        else:
            if len(dosyalar_dictionary['future_restock']) == 0 or len(dosyalar_dictionary['future_future']) == 0:
                text_print(output_text, 'Maalesef sağlanan dosyalar doğru bir şekilde algılanamadı, dosyaları istenilen şekilde sürükleyip bıraktığınızdan emin olunuz...')
            else:
                future_price_script_starter(path, name, dosyalar_dictionary['future_restock'][0], dosyalar_dictionary['future_future'][0], output_text)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))
    baslat_button.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black', baslat_button))
    baslat_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', baslat_button))
    baslat_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, baslat_button))
    baslat_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', baslat_button))

    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('ftr', save_path, placeholder, save_name_text)
    f_window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    f_window.bind("<Button-1>", lambda e: on_click_outside(e, save_path, placeholder, canvas2_text_color))

    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    save_path.pack(side=LEFT, fill=X, expand=True)

    top_frame.grid(column=0, row=0, sticky='we', padx=(25,0), pady=(25,0))
    bottom_frame.grid(column=0, row=1, sticky='we', padx=(25,0))
    top_frame.grid_columnconfigure(0, weight=1)
    bottom_frame.grid_columnconfigure(0, weight=1)
    title.grid(column=0, row=0, sticky='w')
    title_line.grid(column=0, row=1, sticky='we')
    save_path_label.grid(column=0, row=2, sticky='w')
    path_frame.grid(column=0, row=3, sticky='we')
    save_name_label.grid(column=0, row=4, sticky='w')
    save_name_text.grid(column=0, row=5, sticky='we')
    return_items_res = drag_drop(row1=6, row=7, column=0, dict_name='future_restock',
                                 text='Restock excel dosyasini asagiya surukleyip birakiniz:',
                                 parent=bottom_frame, padx=0, pady=0, win=f_window)
    return_items_ftr = drag_drop(row1=8, row=9, column=0, dict_name='future_future',
                                 text='Future Price excel dosyasini asagiya surukleyip birakiniz:',
                                 parent=bottom_frame, padx=0, pady=0, win=f_window)
    baslat_button.grid(column=0, row=10, sticky='e', pady=(10,0))
    def on_close():
        canvas2.bind_all("<MouseWheel>", lambda e: on_mouse_wheel(e))
        f_window.destroy()
    f_window.protocol('WM_DELETE_WINDOW', on_close)
    f_window.mainloop()


def restock(canvas2):
    scale = 1
    liste_restock= [canvas, canvas2,button_1, button_2, button_3, button_4, button_5]
    global restock_submit_button
    global restock_inner_frame
    global restock_main_scrollbar
    global restock_output

    def on_resize(e):
        scale = main_frame_resize()
        canvas2.config(scrollregion=canvas2.bbox('all'))
        item_list = [ham_surukle_text, export_surukle_text, restock_surukle_text]
        ust_list = [restock_question, export_question, file_path]

        for item in item_list:
            item.config(font=("JetBrainsMonoRoman Regular", round(9*scale)))
        ham_main_canvas.config(height=175*scale)
        export_main_canvas.config(height=175*scale)
        restock_main_canvas.config(height=175*scale)
        alt_canvas_uzaklik = alt_canvas.winfo_y()
        p = alt_canvas_uzaklik + alt_canvas.winfo_height()
        if canvas2.winfo_width() < resize_dictionary[restock_inner_frame]['width']*scale:
            frame_width = canvas2.winfo_width()
        else:
            frame_width = resize_dictionary[restock_inner_frame]['width']*scale
        if frame_width < 1000:
            frame_width = 1000
        if p+55 > canvas2.winfo_height():
            restock_inner_frame.configure(height=p+55, width=frame_width)
        else:
            restock_inner_frame.configure(height=canvas2.winfo_height(), width=frame_width)



    settings_height = 250
    if 'restock_settings.txt' not in os.listdir('Settings'):
        settings('Settings/restock_settings.txt', settings_var)


    def restock_builder(a,height):
        active_dictionary['restock'] = 1
        if active_dictionary['export'] == 0:
            active_dictionary['export'] = 1
            export_question_switch.pasif()
        a = active_dictionary['restock'] + active_dictionary['export']
        updater()


    def restock_destroyer():
        active_dictionary['restock'] = 0
        a = active_dictionary['restock'] + active_dictionary['export']
        #alt_canvas.configure(height=(a+1)*height+25*(a+1))
        #alt_canvas.update()
        try:
            restock_surukle_text.grid_forget()
            restock_main_canvas.grid_forget()
        except:
            pass
        updater()

    def export_builder(a,height):
        active_dictionary['export'] = 1
        a = active_dictionary['restock'] + active_dictionary['export']
        updater()

    def export_destroyer():
        active_dictionary['export'] = 0
        if active_dictionary['restock'] == 1:
            active_dictionary['restock'] = 0
            restock_question_switch.active()
        a = active_dictionary['restock'] + active_dictionary['export']

        #alt_canvas.configure(height=(a+1)*height+25*(a+1))
        #alt_canvas.update()
        try:
            export_surukle_text.grid_forget()
            export_main_canvas.grid_forget()
            restock_surukle_text.grid_forget()
            restock_main_canvas.grid_forget()
        except:
            pass
        updater()
    def updater():
        a = active_dictionary['export'] + active_dictionary['restock']
        restock_submit_button.grid(column=1, row=11, padx=(10,25), pady=10, sticky='e')
        future_price_button.grid(column=0, row=11, pady=10, sticky='e')

        if active_dictionary['export'] == 1:
            export_main_canvas.grid(column = 0, row = 5, columnspan=2, sticky='nwes', padx=25)
            export_surukle_text.grid(column=0, row=4, columnspan=2, padx=25, pady=10, sticky='w')


        if active_dictionary['restock'] == 1:
            restock_main_canvas.grid(column = 0, row = 7, columnspan=2, sticky='nwes', padx=25)
            restock_surukle_text.grid(column=0, row=6, columnspan=2, padx=25, pady=10, sticky='w')


        alt_canvas.config(height=restock_settings.winfo_y()+restock_settings.winfo_height()+100)
        alt_canvas.update()
        p = alt_canvas.winfo_y() + alt_canvas.winfo_height()

        if p+55 > canvas2.winfo_height():
            restock_inner_frame.configure(height=p+55)
        else:
            restock_inner_frame.configure(height=canvas2.winfo_height())

    restock_inner_frame = Frame(canvas2, bg=color, height= canvas2.winfo_height(), width=0)
    restock_inner_frame.grid_columnconfigure(0, weight=1)
    restock_inner_frame.grid_propagate(False)
    canvas2.create_window((0, 0), window=restock_inner_frame, anchor="nw")
    #resize_liste.append([0,0,600,900,restock_inner_frame])
    restock_main_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    restock_main_scrollbar.pack(side= RIGHT, fill=Y)
    canvas2.configure(yscrollcommand=restock_main_scrollbar.set)

    canvas2.bind_all("<MouseWheel>", on_mouse_wheel)
    #bg='#ADD8E6'
    alt_canvas = Canvas(
        restock_inner_frame,
        height=int(0),
        width=int(canvas2.winfo_width()),
        borderwidth=0, highlightthickness=0,
        bg=color, highlightbackground=color,
    )
    alt_canvas.grid_columnconfigure(0, weight=1)
    height = int(150)

    width = int(650)
    ham_liste = ham_drag_drop2(row1=0,row=1,column=0,dict_name='ham_dosyalar_liste',text="Ham dosyalarin excellerini asagiya surukleyip birakiniz:", parent=alt_canvas)
    ham_main_canvas= ham_liste[0]
    ham_surukle_text = ham_liste[1]
    export_liste = drag_drop(row1=2,row=3,column=0,dict_name='export_dosyalar_liste',text="Export dosyalarin excellerini asagiya surukleyip birakiniz:", parent=alt_canvas)
    export_main_canvas= export_liste[0]
    export_surukle_text = export_liste[1]
    restock_liste = drag_drop(row1=4,row=5,column=0,dict_name='restock_dosyalar_liste',text="Restock excelini asagiya surukleyip birakiniz:", parent=alt_canvas)
    restock_main_canvas= restock_liste[0]
    restock_surukle_text = restock_liste[1]
    a = 1
    alt_canvas.configure(height= height*4+settings_height+25)
    alt_canvas.update()

    settings_label = Label(alt_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)
    settings_label.grid(column=0, row=9, columnspan=2, sticky = 'w', padx=25, pady=3)
    restock_settings = Text(alt_canvas,insertbackground='#c0c0c0', border=0, wrap= WORD,width=int(width_f(650)), bg=line_color, fg='#c0c0c0', height = int(settings_height/15),font=("JetBrainsMonoRoman Regular", 10))
    restock_settings.grid(column=0, row=10, columnspan=2, sticky = 'we', padx=25, pady=5)
    #restock_settings.place(x=25, y=(0+1)*height+25*(0+1))
    restock_settings.bind('<Enter>',lambda e: on_text_enter(e))
    restock_settings.bind('<Leave>',lambda e: on_text_leave(e))




    with open('Settings/restock_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        restock_settings.insert(tk.END, readed)
        restock_settings.see(tk.END)
    alt_canvas.update()
    ust_canvas = Canvas(restock_inner_frame, background=color, highlightthickness=0)
    pad = 35
    options_frame = Frame(
        ust_canvas,
        bg=line_color,
        width = 450,
        height=142,
    )
    options_frame.grid_propagate(False)
    options_frame.grid_columnconfigure(0, weight=1)
    restock_option_frame = Frame(
        options_frame,
        bg=line_color,
        height=70
    )
    restock_option_frame.pack_propagate(False)
    restock_option_frame.grid(column=0, row=0, sticky='we')
    restock_question = Label(
        restock_option_frame,
        text='Restock',
        background=line_color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    restock_question_switch = SwitchButton(
        parent = restock_option_frame,
        active_function= lambda: restock_builder(a, height),
        pasif_function= lambda: restock_destroyer(),
        border=0,
        highlightthickness=0,
        f='red',
        s='green',
        status=True
    )
    restock_question.pack(side='left', padx=(15, 0))
    restock_question_switch.pack(side='right', padx=(0, 15))

    line_1 = Frame(
        options_frame,
        height=2,
        bg='#79918B'
    )
    line_1.grid(column=0, row=1, sticky='we')

    export_option_frame = Frame(
        options_frame,
        bg=line_color,
        height=70
    )
    export_option_frame.pack_propagate(False)
    export_option_frame.grid(column=0, row=2, sticky='we')
    export_question = Label(
        export_option_frame,
        text='Export',
        background=line_color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    export_question_switch = SwitchButton(
        parent = export_option_frame,
        active_function= lambda: export_builder(a, height),
        pasif_function= lambda: export_destroyer(),
        border=0,
        highlightthickness=0,
        f='red',
        s='green',
        status=True
    )
    export_question.pack(side='left', padx=(15, 0))
    export_question_switch.pack(side='right', padx=(0, 15))

    welcome_text = Label(
        ust_canvas,
        background=color,
        text="RESTOCK PROGRAMI",
        font=("JetBrainsMonoRoman Regular", 24 * -1),
        fg=canvas2_text_color,
    )
    welcome_line = Frame(
        ust_canvas,
        background=line_color,
        height=2
    )
    file_path = Label(
        ust_canvas,
        text= 'Aşağıya sonuçların kaydedilmesini istediğiniz dosya yolunu giriniz:',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    browse_frame = Frame(ust_canvas, bg=color, height=30)
    save_path_text = Text(
        browse_frame,
        height=1,
        border=0,
        fg='#747474',
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')
    browse_button = MyButton(
        browse_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )

    save_button = MyButton(
        browse_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_name_label = Label(
        ust_canvas,
        text= 'Aşağıya sonucun kaydedilmesini istediginiz ismi giriniz:',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    save_name_text = Text(
        ust_canvas,
        height=1,
        border=0,
        fg=canvas2_text_color,
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')

    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('res', save_path_text)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path_text, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))


    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('res', save_path_text, placeholder, save_name_text)
    window.unbind("<Button-1>")
    save_path_text.bind("<Button-1>", lambda e: on_focus_in(e, save_path_text, placeholder, canvas2_text_color))
    save_path_text.bind("<FocusOut>", lambda e: on_focus_out(e, save_path_text, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, save_path_text, placeholder, canvas2_text_color))
    browse_frame.pack_propagate(False)
    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    save_path_text.pack(side=LEFT, fill=X, expand=True)


    ust_canvas.grid_columnconfigure(0, weight=1)
    welcome_text.grid(column=0, row=0, sticky='w', padx=(25,0), pady=(20,0))
    welcome_line.grid(column=0, row=1, sticky='we', padx=(25,0))
    options_frame.grid(column=0, row=2, sticky='w', padx=(25,0), pady=(20,15))
    file_path.grid(column=0, row=6, sticky='w', padx=(25,0), pady=(12,3))
    browse_frame.grid(column=0, row=7, sticky='we', padx=(25, 275))
    save_name_label.grid(column=0, row=8, sticky='w', padx=(25,0), pady=(12,3))
    save_name_text.grid(column=0, row=9, sticky='we', padx=(25, 275))
    ust_canvas.grid(column=0,row=0,sticky='we')




    restock_inner_frame.bind("<Configure>", lambda e: canvas2.config(scrollregion=canvas2.bbox("all")))
    restock_inner_frame.configure(height=canvas2.winfo_height())
    restock_inner_frame.update()

    restock_output = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    restock_output.bind('<Enter>', on_text_enter)
    restock_output.bind('<Leave>', on_text_leave)

    global progress
    progress = ttk.Progressbar(window, orient=HORIZONTAL, mode='determinate')

    def print_liste(restock_settings, path):
        color_change(1,'#8AB4F8','black', restock_submit_button)
        def submit_resize(event):
            scale = main_frame_resize()
            item_list = [ham_surukle_text, export_surukle_text, restock_surukle_text]
            ust_list = [save_path_text, restock_question, export_question, file_path]

            for item in item_list:
                item.config(font=("JetBrainsMonoRoman Regular", round(9*scale)))
            ham_main_canvas.config(height=175*scale)
            export_main_canvas.config(height=175*scale)
            restock_main_canvas.config(height=175*scale)
            for item in ust_list:
                item.config(font=("JetBrainsMonoRoman Regular", round(9*scale)))
            restock_output.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0))
            restock_output.config(font=("JetBrainsMonoRoman Regular", round(9*scale)))
            restock_output.update_idletasks()
            try:
                progress.pack_configure(padx=(canvas.winfo_width(),0))
            except:pass
            alt_canvas_uzaklik = alt_canvas.winfo_y() + 250
            p = alt_canvas_uzaklik + alt_canvas.winfo_height()
            if canvas2.winfo_width() < resize_dictionary[restock_inner_frame]['width']*scale:
                frame_width = canvas2.winfo_width()
            else:
                frame_width = resize_dictionary[restock_inner_frame]['width']*scale
            if frame_width < 1000:
                frame_width = 1000
            if p+55 > canvas2.winfo_height():
                restock_inner_frame.configure(height=p+55, width=frame_width)
            else:
                restock_inner_frame.configure(height=canvas2.winfo_height(), width=frame_width)

        n_ham_dosyalar_liste = []
        n_export_dosyalar_liste = []
        n_restock_dosyalar_liste = []
        try:
            for i in dosyalar_dictionary['ham_dosyalar_liste']:
                if i[0] == ' ':
                    i = i.replace(' ','',1)
                n_ham_dosyalar_liste.append(i)
        except:pass
        try:
            for i in dosyalar_dictionary['export_dosyalar_liste']:
                if i[0] == ' ':
                    i = i.replace(' ','',1)
                n_export_dosyalar_liste.append(i)
        except:pass
        try:
            for i in dosyalar_dictionary['restock_dosyalar_liste']:
                if i[0] == ' ':
                    i = i.replace(' ','',1)
                n_restock_dosyalar_liste.append(i)
        except:pass
        restock_ayarlar = restock_settings.get("1.0", tk.END)
        restock_ayarlar = restock_ayarlar.rstrip("\n")
        settings('Settings/restock_settings.txt', restock_ayarlar)
        save_name = save_name_text.get(1.0, tk.END).strip('\n')
        save_location_saver('res', save_name_text)
        path = path.replace('\n','')

        restock_output.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0))

        progress.pack_forget()
        progress.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0))

        if path != "Example: C:/Users/Username/Desktop/sonuc":
            start_excel_editor_thread(n_ham_dosyalar_liste,n_export_dosyalar_liste,n_restock_dosyalar_liste,path,active_dictionary,restock_output, save_name, progress)
        else:
            restock_output.insert(END, 'dosya yolunu dogru girdiginizden emin olun ve tekrar deneyin...\n')
            restock_output.see(END)
        window.unbind('<Configure>')
        submit_resize(1)
        window.bind('<Configure>', submit_resize)



    restock_submit_button = MyButton(
        alt_canvas,
        round=15,
        width=100,
        height=50,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    future_price_button = MyButton(
        alt_canvas,
        round=15,
        width=150,
        height=50,
        text='Future Price',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    def color_change(e,c,t, b):
        b.config(background=c, text_color=t)
    restock_submit_button.bind('<Button-1>', lambda e: print_liste(restock_settings, save_path_text.get(1.0, END)))
    restock_submit_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', restock_submit_button))
    restock_submit_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, restock_submit_button))
    restock_submit_button.bind("<Leave>", lambda e: color_change(e,line_color,'white', restock_submit_button))





    future_price_button.bind('<Button-1>', lambda e: future_price_window(future_price_button))
    future_price_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', future_price_button))
    future_price_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, future_price_button))
    future_price_button.bind("<Leave>", lambda e: color_change(e,line_color,'white', future_price_button))

    restock_inner_frame.update()
    alt_canvas.grid(column=0, row=1, sticky='we', padx=(0,300))

    window.update_idletasks()

    #window.update_idletasks()
    updater()
    resize_dictionary[restock_inner_frame] = {'width': canvas2.winfo_width(), 'height': restock_inner_frame.winfo_height()}
    resize_dictionary[alt_canvas] = {'width': alt_canvas.winfo_width(), 'height': alt_canvas.winfo_height(), 'x': alt_canvas.winfo_x(), 'y': alt_canvas.winfo_y()}







    window.unbind('<Configure>')
    on_resize(1)
    window.bind('<Configure>', lambda e: on_resize(e))


def button_expration(canvas2):
    def color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def baslat_click(e,c,t,b):
        color_change(e,c,t,b)
        baslat_button.pack_forget()
        baslat_button.pack(side=RIGHT, padx=(5,0))

        path = save_path.get(1.0, tk.END)
        path = path.strip("\n")
        output(path)

    def output(path):
        window.unbind("<Configure>")
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0), anchor='w')
        expration_ayarlar = expration_settings_text.get("1.0", tk.END)
        expration_ayarlar = expration_ayarlar.rstrip("\n")
        settings('Settings/expration_settings.txt', expration_ayarlar)
        item_ids = item_ids_text.get(1.0, tk.END).strip('\n')
        username = expration_username_entry.get().strip('\n')
        password = expration_password_entry.get().strip('\n')
        print([username, password])

        if path == "Example: C:/Users/Username/Desktop/sonuc":
            text_print(output_text, "Maalesef path degeri algilanamadi! Dogru bir deger girdiginizden emin olup tekrar deneyiniz.")
        elif item_ids == '':
            text_print(output_text, "Lütfen düzgün bir shipment id değeri giriniz.")
        else:
            start_expration_thread(username, password, output_text, path, item_ids)
        window.bind("<Configure>", lambda e: expration_resize(e, 1))
    def expration_resize(e, isactive):
        scale = main_frame_resize()
        height = expration_bottom_canvas.winfo_y()+expration_bottom_canvas.winfo_height()+20 + isactive*output_text.winfo_height()
        if height < canvas2.winfo_height():
            block_frame.config(width=750*scale, height=canvas2.winfo_height())
        else:
            block_frame.config(width=750*scale, height=height)
        expration_login_main_frame.grid_columnconfigure(0, weight=1, minsize=300*scale)
        if scale >= 1.3:
            expration_settings_text.config(font=("JetBrainsMonoRoman Regular", round(10*(scale-0.3))))
            if scale >= 1.4:
                expration_username_entry.config(font=("JetBrainsMonoRoman Regular", round(12*(scale-0.4))))
                expration_password_entry.config(font=("JetBrainsMonoRoman Regular", round(12*(scale-0.4))))
        if isactive == 1:
            output_text.pack_configure(padx=(canvas.winfo_width(), 0))
        canvas2.config(scrollregion=canvas2.bbox('all'))


    block_frame = Frame(
        canvas2,
        background=color,
        width=750,
        height=canvas2.winfo_height()
    )
    canvas2.create_window((0, 0), window=block_frame, anchor='nw')
    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    expration_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=expration_scrollbar.set, scrollregion=canvas2.bbox('all'))
    expration_scrollbar.pack(side=RIGHT, fill=Y)
    expration_top_canvas = Canvas(
        block_frame,
        background=color,
        border=0,
        highlightthickness=0
    )
    expration_bottom_canvas = Canvas(
        block_frame,
        background=color,
        border=0,
        highlightthickness=0
    )
    expration_title = Label(
        expration_top_canvas,
        background=color,
        foreground=canvas2_text_color,
        text='Expration Date',
        font=("JetBrainsMonoRoman Regular", 24*-1)
    )
    expration_title_line = Frame(
        expration_top_canvas,
        background=line_color,
        height=2
    )
    save_path_label = Label(
        expration_top_canvas,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        expration_top_canvas,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('exp', save_path)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))

    expration_login_main_frame = Frame(
        expration_bottom_canvas,
        background=color,
    )
    expration_username_label = Label(
        expration_login_main_frame,
        text="Kullanici Adi:",
        background=color,
        foreground=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    expration_username_entry = Entry(
        expration_login_main_frame,
        border=0,
        highlightthickness=3,
        highlightcolor=line_color,
        highlightbackground=line_color,
        background=line_color,
        insertbackground=canvas2_text_color,
        foreground=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    expration_password_label = Label(
        expration_login_main_frame,
        text="Şifre:",
        background=color,
        foreground=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    expration_password_entry = Entry(
        expration_login_main_frame,
        border=0,
        highlightthickness=3,
        highlightcolor=line_color,
        highlightbackground=line_color,
        background=line_color,
        insertbackground=canvas2_text_color,
        foreground=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12),
    )
    button_group = Frame(
        expration_login_main_frame,
        background=color,
        width=300,
    )
    baslat_button = MyButton(
        button_group,
        round=12,
        width=100,
        height=40,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )

    baslat_button.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black', baslat_button))
    baslat_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', baslat_button))
    baslat_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, baslat_button))
    baslat_button.bind("<Leave>", lambda e: color_change(e,line_color,'white', baslat_button))


    baslat_button.pack(side=RIGHT, padx=(10,0))
    settings_height=150
    if "expration_settings.txt" not in os.listdir('Settings'):
        settings("Settings/expration_settings.txt", expration_settings_var)
    expration_settings_text = Text(
        expration_bottom_canvas,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )

    login_dictionary = {
        'default_email': [],
        'default_password': []
    }
    with open('Settings/expration_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        expration_settings_text.insert(tk.END, readed)
        expration_settings_text.see(tk.END)
        lines = readed.split('\n')
        for line in lines:
            line = line.split('=')
            if line[0] == 'default_email' or line[0] == 'default_email ':
                degerler = line[1].split(',')
                for deger in degerler:
                    deger = deger.replace('\n', '')
                    deger = deger.replace(' ', '', 1)
                    login_dictionary['default_email'].append(deger)
            elif line[0] == 'default_password' or line[0] == 'default_password ':
                degerler = line[1].split(',')
                for deger in degerler:
                    deger = deger.replace('\n', '')
                    deger = deger.replace(' ', '', 1)
                    login_dictionary['default_password'].append(deger)
        if expration_password_entry.get() == "":
            expration_password_entry.insert(0, login_dictionary['default_password'][0])
        if expration_username_entry.get() == "":
            expration_username_entry.insert(0, login_dictionary['default_email'][0])
    output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )

    item_ids_label = Label(
        expration_top_canvas,
        text= 'Aşağıya shipment id\'lerini giriniz.(birden fazla id girilecek ise virgül ile ayırınız.):',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    item_ids_text = Text(
        expration_top_canvas,
        height=1,
        border=0,
        fg=canvas2_text_color,
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')


    block_frame.grid_propagate(False)
    block_frame.grid_columnconfigure(0, weight=1)
    expration_top_canvas.grid_columnconfigure(0, weight=1)
    expration_bottom_canvas.grid_columnconfigure(0, weight=1)
    expration_top_canvas.grid(column=0, row=0, sticky="ew", padx=(25,0), pady=(25,0))
    expration_bottom_canvas.grid(column=0, row=1, sticky="ew", padx=(25,0))
    expration_title.grid(column=0, row=0, sticky='w')
    expration_title_line.grid(column=0, row=1, sticky='we')


    expration_login_main_frame.grid_columnconfigure(0, weight=1, minsize=300)
    expration_login_main_frame.grid(column=0, row=0, sticky='w')
    expration_settings_text.grid(column=0, row=1, sticky='we', pady=(25,0))
    expration_username_label.grid(column=0, row=0, sticky='w', pady=(25,5))
    expration_username_entry.grid(column=0, row=1, sticky='we')
    expration_password_label.grid(column=0, row=2, sticky='w', pady=(10,5))
    expration_password_entry.grid(column=0, row=3, sticky='we')
    button_group.grid(column=0, row=4, sticky='ew', pady=(15,0))

    save_path_label.grid(column=0, row=3, sticky='w', pady=(15,0))
    path_frame.grid(column=0, row=4, sticky='we')
    item_ids_label.grid(column=0, row=7, sticky='w', pady=(15,0))
    item_ids_text.grid(column=0, row=8, sticky='we')
    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('exp', save_path, placeholder)
    window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, save_path, placeholder, canvas2_text_color))
    save_path.pack(side=LEFT, fill=X, expand=True)




    canvas2.config(scrollregion=canvas2.bbox('all'))
    window.bind("<Configure>", lambda e: expration_resize(e, 0))


def shipmentcreater_script(path, output_text, save_name, dc_code):

    def settings():
        if 'shipment_settings.txt' not in os.listdir('Settings'):
            with open('Settings/shipment_settings.txt', 'w', encoding='utf-8') as settings:
                text_print(output_text, 'ayarlar dosyası oluşturuluyor...')
                settings.write('RESTOCK:\n'
                               'upc = Upc\n'
                               'pcs = PCS\n'
                               'asin = ASIN\n'
                               'pk = PK\n'
                               'price = Price\n'
                               'suplier = suplier\n'
                               '=====================================================\n'
                               'ORDER FORM:\n'
                               'upc = UPC\n'
                               'pcs = PCS\n'
                               'asin = ASIN 1, ASIN 2, ASIN 3, ASIN 4\n'
                               'SKU = ASIN1_SKU, ASIN2_SKU, ASIN3_SKU, ASIN4_SKU\n'
                               'pk = PK\n'
                               'price = price\n'
                               'suplier = suplier\n'
                               '=====================================================\n'
                               'INVOICE:\n'
                               'shipquantity = ShipQuantity\n'
                               'upc = Upc\n'
                               'price = NetEach2\n'
                               'packsize = PackSize\n'
                               'brand = Brand\n'
                               'description = Description\n')
                settings.close()
    def getSettings():
        text_print(output_text, 'ayarlar yükleniyor...')
        sutunlar_dict = {
            'restock_upc': [],
            'restock_pcs': [],
            'restock_asin': [],
            'restock_pk': [],
            'restock_price': [],
            'restock_suplier': [],
            'orderform_upc': [],
            'orderform_pcs': [],
            'orderform_asin': [],
            'orderform_sku': [],
            'orderform_pk': [],
            'orderform_price': [],
            'orderform_suplier': [],
            'invoice_shipquantity': [],
            'invoice_upc': [],
            'invoice_price': [],
            'invoice_packsize':[],
            'invoice_brand':[],
            'invoice_description':[]
        }

        with open('Settings/shipment_settings.txt', 'r', encoding='utf-8') as settings:
            settings = settings.readlines()
            #RESTOCK SETTINGS
            for line in settings:
                if '=====' in line:
                    break
                line = line.replace('\n', '')
                line = line.split('=')
                if line[0] == 'upc ' or line[0] == 'upc':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_upc'].append(deger)
                elif line[0] == 'pcs ' or line[0] == 'pcs':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_pcs'].append(deger)
                elif line[0] == 'asin ' or line[0] == 'asin':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_asin'].append(deger)
                elif line[0] == 'pk ' or line[0] == 'pk':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_pk'].append(deger)
                elif line[0] == 'price ' or line[0] == 'price':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_price'].append(deger)
                elif line[0] == 'suplier ' or line[0] == 'suplier':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_suplier'].append(deger)


                #ORDER FORM SETTINGS
            a = 0
            for line in settings:
                if '=====' in line:
                    a+=1
                if '=====' in line and a == 2:
                    break
                if a == 1:
                    line = line.replace('\n', '')
                    line = line.split('=')
                    if line[0] == 'upc ' or line[0] == 'upc':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_upc'].append(deger)
                    elif line[0] == 'pcs ' or line[0] == 'pcs':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_pcs'].append(deger)
                    elif line[0] == 'asin ' or line[0] == 'asin':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_asin'].append(deger)
                    elif line[0] == 'SKU ' or line[0] == 'SKU':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_sku'].append(deger)
                    elif line[0] == 'pk ' or line[0] == 'pk':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_pk'].append(deger)
                    elif line[0] == 'price ' or line[0] == 'price':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_price'].append(deger)
                    elif line[0] == 'suplier ' or line[0] == 'suplier':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_suplier'].append(deger)


            #INVOICE SETTINGS
            a = 0
            for line in settings:
                if '=====' in line:
                    a+=1
                if a == 2:
                    line = line.replace('\n', '')
                    line = line.split('=')
                    if line[0] == 'shipquantity ' or line[0] == 'shipquantity':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_shipquantity'].append(deger)
                    elif line[0] == 'upc ' or line[0] == 'upc':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_upc'].append(deger)
                    elif line[0] == 'price ' or line[0] == 'price':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_price'].append(deger)
                    elif line[0] == 'packsize ' or line[0] == 'packsize':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_packsize'].append(deger)
                    elif line[0] == 'brand ' or line[0] == 'brand':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_brand'].append(deger)
                    elif line[0] == 'description ' or line[0] == 'description':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['invoice_description'].append(deger)
        for key in sutunlar_dict.keys():
            text_print(output_text, str(key) + ': ' + str(sutunlar_dict[key]))
        text_print(output_text, 'Ayarlar başarıyla çekİldİ.'.upper())
        return sutunlar_dict
    def invoiceFormReader(sutunlar_dict):
        files = dosyalar_dictionary['invoice']
        invoice_form_dict = {}
        for file in files:
            if file.endswith('.xlsx'):
                df = pd.read_excel(file)
                invoice_form_dict['ShipQuantity'] = df[sutunlar_dict['invoice_shipquantity'][0]].tolist()
                invoice_form_dict['Upc'] = df[sutunlar_dict['invoice_upc'][0]].tolist()
                invoice_form_dict['Price'] = df[sutunlar_dict['invoice_price'][0]].tolist()
                invoice_form_dict['PackSize'] = df[sutunlar_dict['invoice_packsize'][0]].tolist()
                invoice_form_dict['Brand'] = df[sutunlar_dict['invoice_brand'][0]].tolist()
                invoice_form_dict['Description'] = df[sutunlar_dict['invoice_description'][0]].tolist()
        return invoice_form_dict
    def orderFormReader(sutunlar_dict):
        files = dosyalar_dictionary['order_form']
        order_form_dict = {}
        dictionary = {}
        dictionary['UPC'] = []
        dictionary['Price'] = []
        dictionary['ShipQuantity'] = []
        for file in files:
            if file.endswith('.xlsx'):
                df = pd.read_excel(file)
                order_form_dict['Upc'] = df[sutunlar_dict['orderform_upc'][0]].tolist()
                order_form_dict['Price'] = df[sutunlar_dict['orderform_price'][0]].tolist()
                order_form_dict['Suplier'] = df[sutunlar_dict['orderform_suplier'][0]].tolist()
                for i in range(1, len(sutunlar_dict['orderform_asin'])+1):
                    if i != 1:
                        order_form_dict['Pcs '+str(i)] = df[sutunlar_dict['orderform_pcs'][0]+'.'+str(i-1)].tolist()
                    else:
                        order_form_dict['Pcs '+str(i)] = df[sutunlar_dict['orderform_pcs'][0]].tolist()
                    #dictionary['Asin ' + str(i)] = []
                    #dictionary['Pcs ' + str(i)] = []
                    #dictionary['ASIN' + str(i)+'_SKU'] = []
                a = 1
                for name in sutunlar_dict['orderform_asin']:
                    order_form_dict['Asin '+ str(a)] = df[name].tolist()
                    a += 1
                a = 1
                for name in sutunlar_dict['orderform_sku']:
                    order_form_dict['ASIN'+str(a)+'_SKU'] = df[name].tolist()
                    liste = []
                    for i in order_form_dict['ASIN'+str(a)+'_SKU']:
                        if type(i) == str and i.count('_') >= 3:
                            print(i)
                            i = i.split('_')

                            i = i[2]
                            liste.append(i)
                        else:
                            liste.append('#YOK')
                    order_form_dict[f'PK {a}'] = liste
                    a += 1

        for key in order_form_dict.keys():
            print(f"{key}: {order_form_dict[key]}")
        liste = [order_form_dict, dictionary]
        return liste
    def restockFormReader(sutunlar_dict):
        files = dosyalar_dictionary['restock']
        restock_form_dict = {}
        for file in files:
            if file.endswith('.xlsx'):
                df = pd.read_excel(file)
                restock_form_dict['Asin'] = df[sutunlar_dict['restock_asin'][0]].tolist()
                restock_form_dict['Upc'] = df[sutunlar_dict['restock_upc'][0]].tolist()
                restock_form_dict['Pcs'] = df[sutunlar_dict['restock_pcs'][0]].tolist()
                restock_form_dict['PK'] = df[sutunlar_dict['restock_pk'][0]].tolist()
                restock_form_dict['Price'] = df[sutunlar_dict['restock_price'][0]].tolist()
                restock_form_dict['Suplier'] = df[sutunlar_dict['restock_suplier'][0]].tolist()
        return restock_form_dict
    def indexFinder(item, liste):
        a = 0
        index_list = []
        for z in liste:
            if z == item:
                index_list.append(a)
            a += 1
        return index_list
    def match(invoice_form_dict, order_form_dict, restock_form_dict, dictionary):
        liste = []
        text_print(output_text, 'UPC değerleri eşleniyor...')
        dictionary = {}
        dictionary['UPC'] = []
        dictionary['Price'] = []
        dictionary['Price Check'] = []
        dictionary['Suplier'] = []
        dictionary['ShipQuantity'] = []
        dictionary['Asin'] = []
        dictionary['Pcs'] = []
        dictionary['Yeni Pcs'] = []
        dictionary['PK'] = []
        dictionary['SKU'] = []
        dictionary['PackSize'] = []
        dictionary['Brand'] = []
        dictionary['Description'] = []
        dictionary['DOSYA'] = []
        dictionary['SKU2'] = []
        dictionary['PK EACH'] = []
        dictionary['Kalan'] = []

        for upc in invoice_form_dict['Upc']:
            dc = dc_code
            restock_kontrol = 0
            order_kontrol = 0
            if upc in restock_form_dict['Upc']:
                restock_kontrol = 1
                index_list = []
                z = 0
                for i in restock_form_dict['Upc']:
                    if i == upc:
                        index_list.append(z)
                    z+=1
                for index in index_list:
                    index_invoice = invoice_form_dict['Upc'].index(upc)
                    Price = invoice_form_dict['Price'][index_invoice]
                    ShipQuantity = invoice_form_dict['ShipQuantity'][index_invoice]
                    PackSize = invoice_form_dict['PackSize'][index_invoice]
                    Brand = invoice_form_dict['Brand'][index_invoice]
                    Description = invoice_form_dict['Description'][index_invoice]
                    Price_check = restock_form_dict['Price'][index]
                    Suplier = restock_form_dict['Suplier'][index]
                    Asin = restock_form_dict['Asin'][index]
                    Pcs = restock_form_dict['Pcs'][index]
                    Pk = restock_form_dict['PK'][index]
                    if Pk != '#YOK':
                        Pkint = Pk.replace('PK', '')
                        Pkint = int(Pkint)
                    else:
                        Pkint = '#YOK'
                    Sku = '#YOK'
                    excel = 'Restock'
                    upcstr = str(upc)
                    if len(upcstr) < 12:
                        upcstr = upcstr.zfill(12)

                    if False == math.isnan(Pcs):
                        dictionary['UPC'].append(upc)
                        dictionary['Price'].append(Price)
                        dictionary['Price Check'].append(Price_check)
                        dictionary['Suplier'].append(Suplier)
                        dictionary['ShipQuantity'].append(ShipQuantity)
                        dictionary['PackSize'].append(PackSize)
                        dictionary['Brand'].append(Brand)
                        dictionary['Description'].append(Description)
                        dictionary['Asin'].append(Asin)
                        dictionary['Pcs'].append(Pcs)
                        dictionary['Yeni Pcs'].append(0)
                        dictionary['PK'].append(Pk)
                        dictionary['SKU'].append(Sku)
                        dictionary['SKU2'].append(f"{dc}_{upcstr}_{Pk}_{format(Pkint * Price, '.2f')}")
                        dictionary['PK EACH'].append(0)
                        dictionary['Kalan'].append(0)
                        #a = 2
                        #while True:
                        #try:
                        #dictionary['Asin '+str(a)].append(' ')
                        #dictionary['Pcs '+str(a)].append(' ')
                        #dictionary['ASIN'+str(a)+'_SKU'].append(' ')
                        #a+=1
                        #except:
                        #break
                        dictionary['DOSYA'].append(excel)

            if upc in order_form_dict['Upc']:
                index_list = []
                z = 0
                for i in order_form_dict['Upc']:
                    order_kontrol = 1
                    if i == upc:
                        index_list.append(z)
                    z+=1
                for index in index_list:
                    index_invoice = invoice_form_dict['Upc'].index(upc)
                    Price = invoice_form_dict['Price'][index_invoice]
                    ShipQuantity = invoice_form_dict['ShipQuantity'][index_invoice]
                    PackSize = invoice_form_dict['PackSize'][index_invoice]
                    Brand = invoice_form_dict['Brand'][index_invoice]
                    Description = invoice_form_dict['Description'][index_invoice]
                    excel = 'Order Form'
                    upcstr = str(upc)
                    if len(upcstr) < 12:
                        upcstr = upcstr.zfill(12)
                    a = 1
                    while True:
                        try:
                            Asin = order_form_dict['Asin '+str(a)][index]
                            Pcs = order_form_dict['Pcs '+str(a)][index]
                            Sku = order_form_dict['ASIN'+str(a)+'_SKU'][index]
                            PK = order_form_dict[f'PK {a}'][index]
                            Price_check = order_form_dict['Price'][index]
                            Suplier = order_form_dict['Suplier'][index]
                            if PK != '#YOK':
                                Pkint = PK.replace('PK', '')
                                Pkint = int(Pkint)
                            else:
                                Pkint = '#YOK'
                            if type(Asin) == float and math.isnan(Asin) == False:
                                dictionary['UPC'].append(upc)
                                dictionary['Price'].append(Price)
                                dictionary['Price Check'].append(Price_check)
                                dictionary['Suplier'].append(Suplier)
                                dictionary['ShipQuantity'].append(ShipQuantity)
                                dictionary['PackSize'].append(PackSize)
                                dictionary['Brand'].append(Brand)
                                dictionary['Description'].append(Description)
                                dictionary['Asin'].append(Asin)
                                dictionary['Pcs'].append(Pcs)
                                dictionary['Yeni Pcs'].append(0)
                                dictionary['PK'].append(PK)
                                dictionary['SKU'].append(Sku)
                                dictionary['DOSYA'].append(excel)
                                dictionary['SKU2'].append(f"{dc}_{upcstr}_{PK}_{format(Pkint * Price, '.2f')}")
                                dictionary['PK EACH'].append(0)
                                dictionary['Kalan'].append(0)
                            if type(Asin) == str or type(Asin) == int:
                                dictionary['UPC'].append(upc)
                                dictionary['Price'].append(Price)
                                dictionary['Price Check'].append(Price_check)
                                dictionary['Suplier'].append(Suplier)
                                dictionary['ShipQuantity'].append(ShipQuantity)
                                dictionary['PackSize'].append(PackSize)
                                dictionary['Brand'].append(Brand)
                                dictionary['Description'].append(Description)
                                dictionary['Asin'].append(Asin)
                                dictionary['Pcs'].append(Pcs)
                                dictionary['Yeni Pcs'].append(0)
                                dictionary['PK'].append(PK)
                                dictionary['SKU'].append(Sku)
                                dictionary['DOSYA'].append(excel)
                                dictionary['SKU2'].append(f"{dc}_{upcstr}_{PK}_{format(Pkint * Price, '.2f')}")
                                dictionary['PK EACH'].append(0)
                                dictionary['Kalan'].append(0)
                            a+=1
                        except:
                            break
            if (restock_kontrol == 1 or order_kontrol == 1 or (restock_kontrol == 1 or order_kontrol == 1)) and upc not in dictionary['UPC']:
                index_invoice = invoice_form_dict['Upc'].index(upc)
                Price = invoice_form_dict['Price'][index_invoice]
                ShipQuantity = invoice_form_dict['ShipQuantity'][index_invoice]
                PackSize = invoice_form_dict['PackSize'][index_invoice]
                Brand = invoice_form_dict['Brand'][index_invoice]
                Description = invoice_form_dict['Description'][index_invoice]
                dictionary['Asin'].append('#YOK')
                dictionary['Pcs'].append('#YOK')
                dictionary['Yeni Pcs'].append(0)
                dictionary['PK'].append('#YOK')
                dictionary['SKU'].append('#YOK')
                dictionary['UPC'].append(upc)
                dictionary['Price'].append(Price)
                dictionary['ShipQuantity'].append(ShipQuantity)
                dictionary['PackSize'].append(PackSize)
                dictionary['Brand'].append(Brand)
                dictionary['Description'].append(Description)
                dictionary['Price Check'].append('#YOK')
                dictionary['Suplier'].append('#YOK')
                dictionary['SKU2'].append('#YOK')
                dictionary['PK EACH'].append('#YOK')
                dictionary['Kalan'].append('#YOK')
                if restock_kontrol == 1 and order_kontrol == 0:
                    dictionary['DOSYA'].append('restock')
                elif order_kontrol == 1 and restock_kontrol == 0:
                    dictionary['DOSYA'].append('order')
                elif restock_kontrol == 1 and order_kontrol == 1:
                    dictionary['DOSYA'].append('BOTH')
            if upc not in restock_form_dict['Upc'] and upc not in order_form_dict['Upc']:
                index_invoice = invoice_form_dict['Upc'].index(upc)
                Price = invoice_form_dict['Price'][index_invoice]
                ShipQuantity = invoice_form_dict['ShipQuantity'][index_invoice]
                PackSize = invoice_form_dict['PackSize'][index_invoice]
                Brand = invoice_form_dict['Brand'][index_invoice]
                Description = invoice_form_dict['Description'][index_invoice]
                dictionary['Asin'].append('#YOK')
                dictionary['Pcs'].append('#YOK')
                dictionary['Yeni Pcs'].append(0)
                dictionary['PK'].append('#YOK')
                dictionary['SKU'].append('#YOK')
                dictionary['UPC'].append(upc)
                dictionary['Price'].append(Price)
                dictionary['ShipQuantity'].append(ShipQuantity)
                dictionary['PackSize'].append(PackSize)
                dictionary['Brand'].append(Brand)
                dictionary['Description'].append(Description)
                dictionary['Price Check'].append('#YOK')
                dictionary['Suplier'].append('#YOK')
                dictionary['DOSYA'].append('#YOK')
                dictionary['SKU2'].append('#YOK')
                dictionary['PK EACH'].append('#YOK')
                dictionary['Kalan'].append('#YOK')

        letter_dictionary = {
            0: "",
            1: "_A",
            2: "_B",
            3: "_C",
            4: "_D",
            5: "_E",
            6: "_F",
            7: "_G",
            8: "_H",
            9: "_I",
            10: "_J",
            11: "_K",
            12: "_L",
            13: "_M",
            14: "_N",
            15: "_O",
            16: "_P",
            17: "_Q",
            18: "_R",
            19: "_S",
            20: "_T",
            21: "_U",
            22: "_V",
            23: "_Y",
            24: "_Z",
            25: "_Y",
            26: "_Z",
        }
        index_list = []
        for sku in dictionary['SKU2']:
            indexes = indexFinder(sku, dictionary['SKU2'])
            for i, index in enumerate(indexes):
                if index not in index_list and dictionary['SKU2'][index] != '#YOK':
                    dictionary['SKU2'][index] = dictionary['SKU2'][index]+letter_dictionary[i]
                    index_list.append(index)

        return dictionary
    def stock_allocater(dictionary):
        Upcler = dictionary['UPC']
        complated_upc = []
        for upc in Upcler:
            if upc not in complated_upc:
                complated_upc.append(upc)
                index_list = indexFinder(upc, dictionary['UPC'])
                pcs = 0
                ShipQuantity = 0
                oldpk=9999999
                smallest = []
                for index in index_list:
                    ShipQuantity = dictionary['ShipQuantity'][index]
                    if dictionary['Pcs'][index] != '#YOK' and math.isnan(dictionary['Pcs'][index]) != True:
                        pcs = float(dictionary['Pcs'][index]) + pcs
                    if dictionary['PK'][index] != '#YOK':
                        nowpk = dictionary['PK'][0].replace('PK', '')
                        nowpk = int(nowpk)
                        if nowpk <= oldpk:
                            oldpk = nowpk
                            smallest = [nowpk, index]
                for index in index_list:
                    if dictionary['Pcs'][index] != '#YOK' and math.isnan(dictionary['Pcs'][index]) != True:
                        pcs2 = dictionary['Pcs'][index]
                        new_pcs = float(pcs2)/float(pcs)*float(ShipQuantity)
                        new_pcs = round(new_pcs)
                        pk = dictionary['PK'][index]
                        if pk != '#YOK':
                            pk = pk.replace('PK', '')
                            pk = int(pk)
                            kalan = new_pcs % pk
                            if index != smallest[1]:
                                new_pcs = new_pcs - kalan
                                dictionary['Yeni Pcs'][index] = new_pcs
                                dictionary['Yeni Pcs'][smallest[1]] = dictionary['Yeni Pcs'][smallest[1]] + kalan
                            else:
                                dictionary['Yeni Pcs'][smallest[1]] = dictionary['Yeni Pcs'][smallest[1]] + new_pcs
                        else:
                            dictionary['Yeni Pcs'][index] = new_pcs

                for index in index_list:
                    pk = dictionary['PK'][index]
                    if pk != '#YOK':
                        pk = pk.replace('PK', '')
                        pk = int(pk)
                        yenipcs = dictionary['Yeni Pcs'][index]
                        pcseach = int(yenipcs/pk)
                        kalan = yenipcs % pk
                        dictionary['PK EACH'][index] = pcseach
                        dictionary['Kalan'][index] = kalan

        text_print(output_text, 'hazırlanmış excel dosyası kaydediliyor.')
        for key in dictionary.keys():
            print(f"{key}: {dictionary[key]}")
        excel_dosya = pd.DataFrame(dictionary)
        excel_dosya.to_excel(path+f'/{save_name}.xlsx', index=False)
    def main():
        try:
            try:
                settings()
            except Exception:
                text_print(output_text, 'settings.txt dosyası oluşturulurken bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            try:
                sutunlar_dict = getSettings()
            except Exception:
                text_print(output_text, 'Ayarlar settings.txt dosyasından çekilirken bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            try:
                liste = orderFormReader(sutunlar_dict)
            except Exception:
                text_print(output_text, 'Order Form dosyası okunurken bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            try:
                invoice_form_dict = invoiceFormReader(sutunlar_dict)
            except Exception:
                text_print(output_text, 'Invoice dosyası okunurken bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            order_form_dict = liste[0]
            try:
                restock_form_dict = restockFormReader(sutunlar_dict)
            except Exception:
                text_print(output_text, 'Restock dosyası okunurken bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            dictionary = liste[1]
            try:
                dictionary = match(invoice_form_dict, order_form_dict, restock_form_dict, dictionary)
            except Exception:
                text_print(output_text, 'UPC eşleme işlemi sırasında bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            try:
                stock_allocater(dictionary)
            except Exception:
                text_print(output_text, 'Yeni Pcs\'leri hesaplama işlemi sırasında bir hata meydana geldi.')
                text_print(output_text, traceback.format_exc(), color='red')
                sys.exit()
            text_print(output_text, 'İşlem başarıyla tamamlandı!', color='green')
            open_folder_in_explorer(path)
            sys.exit()
        except Exception:
            text_print(output_text, 'Tanımlanamayan bir hata meydana geldi.')
            text_print(output_text, traceback.format_exc(), color='red')
            sys.exit()
    main()


def shipmentCreater(canvas2):
    #MOUSE SCROLL
    def on_mouse_wheel(event):
        canvas2.yview_scroll(int(-1*(event.delta/120)), "units")
    #SCROLLBAR VE FRAME OLUSUMU
    shipment_inner_frame = Frame(canvas2, width = 0, height=0, bg=color)
    canvas2.create_window((0, 0), anchor='nw', window=shipment_inner_frame)
    shipment_scrollbar_y = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.configure(yscrollcommand=shipment_scrollbar_y.set)
    shipment_scrollbar_y.pack(side=RIGHT, fill=Y)


    #INNER FRAME GRID SETTINGS
    shipment_inner_frame.grid_propagate(False)
    shipment_inner_frame.grid_columnconfigure(0, weight=1)

    #OGELERI GURUPLAMAK ICIN CANVAS OLUSUMU

    title_canvas = Canvas(shipment_inner_frame,bg=color,highlightthickness=0)
    title_canvas.grid(column=0, row=0, sticky='nwes')
    items_canvas = Canvas(shipment_inner_frame,bg=color,highlightthickness=0)

    items_canvas.grid(column=0, row=1, sticky='nwes')


    title_canvas.grid_columnconfigure(0, weight=1)
    items_canvas.grid_columnconfigure(0, weight=1)
    #CANVASLAR ICINEKI OGELERIN OLUSUMU

    Shipment_Title = Label(
        title_canvas,
        text="Shipment Creater",
        font=("JetBrainsMonoRoman Regular", 24 * -1),
        bg=color,
        fg=canvas2_text_color
    )
    title_line = Frame(title_canvas, height=2, bg=line_color)
    shipment_output = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    shipment_output.bind('<Enter>',lambda e: on_text_enter(e))
    shipment_output.bind('<Leave>',lambda e: on_text_leave(e))
    shipment_output_line = Frame(
        window,
        height=2,
        bg='#787a7e'
    )

    path_label = Label(
        title_canvas,
        text="Aşağıya sonuçların kaydedilmesini istediğiniz dosya yolunu giriniz:",
        background=color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )


    path_frame = Frame(title_canvas, bg=color, height=30)
    path_text = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg="#747474",
        border=0,
        pady=4,
        bg=line_color,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_name_label = Label(
        title_canvas,
        text= 'Aşağıya sonucun kaydedilmesini istediginiz ismi giriniz:',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    save_name_text = Text(
        title_canvas,
        height=1,
        border=0,
        fg=canvas2_text_color,
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')
    dc_name_label = Label(
        title_canvas,
        text= 'DC KODU:',
        background=color,
        font=("JetBrainsMonoRoman Regular", 12),
        fg=canvas2_text_color,
    )
    dc_name_text = Text(
        title_canvas,
        height=1,
        border=0,
        fg=canvas2_text_color,
        bg=line_color,
        font=("JetBrainsMonoRoman Regular", 12),
        pady=4,
        insertbackground='#c0c0c0')
    settings_height = 250
    settings_label = Label(items_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)
    if 'shipment_settings.txt' not in os.listdir('Settings'):
        settings('Settings/shipment_settings.txt', shipment_settings_var)
    shipment_settings = Text(
        items_canvas,
        border=0,
        wrap= WORD,
        width=int(width_f(650)),
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )
    shipment_settings.bind('<Enter>',lambda e: on_text_enter(e))
    shipment_settings.bind('<Leave>',lambda e: on_text_leave(e))
    with open('Settings/shipment_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        shipment_settings.insert(tk.END, readed)
        shipment_settings.see(tk.END)


    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('shi', path_text)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', path_text, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))

    browse_button.pack(side=RIGHT, fill=Y, padx=(8,0))
    save_button.pack(side=RIGHT, fill=Y, padx=(8,0))
    path_text.pack(side=LEFT, fill=X,expand=True, padx=0, pady=0)
    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('shi', path_text, placeholder, save_name_text)
    window.unbind("<Button-1>")
    path_text.bind("<Button-1>", lambda e: on_focus_in(e, path_text, placeholder, canvas2_text_color))
    path_text.bind("<FocusOut>", lambda e: on_focus_out(e, path_text, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, path_text, placeholder, canvas2_text_color))

    def resize(e,a):
        scale = main_frame_resize()
        items = items_canvas.winfo_children()
        shipment_inner_frame.config(width=resize_dictionary[shipment_inner_frame]['width']*scale)
        for item in items:
            if type(item) == Label and item != settings_label:
                item.config(font=("JetBrainsMonoRoman Regular", round(9*scale)))
            elif type(item) == Frame:
                item.config(height=round(175*scale))
        k=20
        if a == 1:
            shipment_output.pack_configure(padx=(canvas.winfo_width(), 0))
            shipment_output_line.pack_configure(padx=(canvas.winfo_width(), 0))
            k = 300
        p = items_canvas.winfo_y()+items_canvas.winfo_height()+k
        if p>=canvas2.winfo_height():
            shipment_inner_frame.config(height=p)
        else:
            shipment_inner_frame.config(height=canvas2.winfo_height())


    def output(path):
        shipment_output.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        shipment_output_line.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        shipment_ayarlar = shipment_settings.get("1.0", tk.END)
        shipment_ayarlar = shipment_ayarlar.rstrip("\n")
        settings("Settings/shipment_settings.txt", shipment_ayarlar)
        save_name = save_name_text.get(1.0, tk.END).strip('\n')
        save_location_saver('shi', save_name_text)
        dc_name = dc_name_text.get(1.0, tk.END)
        dc_name = dc_name.strip('\n')
        if path == "Example: C:/Users/Username/Desktop/sonuc":
            text_print(shipment_output, 'Maalesef path degeri algilanamadi. Lutfen dogru bir dosya yolu belirttiginizden emin olup tekrar deneyiniz')
        elif dc_name == "":
            text_print(shipment_output, 'Maalesef dc kod degeri algilanamadi. Lutfen dogru bir kod belirttiginizden emin olup tekrar deneyiniz')
        else:
            text_print(shipment_output, path)
            text_print(shipment_output, str(dosyalar_dictionary['order_form']))
            text_print(shipment_output, str(dosyalar_dictionary['invoice']))
            text_print(shipment_output, str(dosyalar_dictionary['restock']))
            def shipmentcreater_script_starter(path, output_text, save_name, dc_name):
                t = Thread(target=shipmentcreater_script, args=(path, output_text, save_name, dc_name), daemon=True)
                t.start()
            shipmentcreater_script_starter(path, shipment_output, save_name, dc_name)
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize(e, 1))
    shipment_submit_button = MyButton(
        items_canvas,
        round=15,
        width=100,
        height=50,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    def color_change(e,c,t):
        shipment_submit_button.config(background=c, text_color=t)
    def shipment_submit_click(e,c,t):
        shipment_submit_button.config(background=c, text_color=t)
        path = path_text.get(1.0, tk.END)
        path = path.strip("\n")
        output(path)
    shipment_submit_button.bind("<Button-1>", lambda e: shipment_submit_click(e,'#8AB4F8','black'))
    shipment_submit_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white'))
    shipment_submit_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color))
    shipment_submit_button.bind("<Leave>", lambda e: color_change(e,line_color,'white'))


    #YERLESIM

    Shipment_Title.grid(column=0, row=0, sticky='w', padx=(25,0), pady=(25,0))
    title_line.grid(column=0,row=1,sticky='we', padx=(20,0))
    path_label.grid(column=0, row=2, pady=(20,0), padx=(25,0), sticky='w')
    path_frame.grid(column=0, row=3,pady=(0,20), padx=(25,5), sticky='we')
    save_name_label.grid(column=0, row=4, pady=(0,0), padx=(25,0), sticky='w')
    save_name_text.grid(column=0, row=5,pady=(0,20), padx=(25,5), sticky='we')
    dc_name_label.grid(column=0, row=6, pady=(0,0), padx=(25,0), sticky='w')
    dc_name_text.grid(column=0, row=7,pady=(0,20), padx=(25,5), sticky='we')
    drag_drop(0,1,0,'invoice','Invoice excelini aşağıya sürükleyip bırakın:', items_canvas)
    drag_drop(2,3,0,'order_form','OrderForm excelini aşağıya sürükleyip bırakın:', items_canvas)
    drag_drop(4,5,0,'restock','Restock excelini aşağıya sürükleyip bırakın:', items_canvas)
    settings_label.grid(column=0, row=6, columnspan=2, sticky = 'w', padx=25, pady=3)
    shipment_settings.grid(column=0, row=7, sticky='we', padx=25, pady=4,)
    shipment_submit_button.grid(column=0, row=8, sticky='e', padx=(0,25), pady=(25,0))

    #KUTUPHANEYE EKLEME
    canvas2.update_idletasks()
    resize_dictionary[shipment_inner_frame] = {'width': 750, 'height': shipment_inner_frame.winfo_height()}

    #RESIZE
    p = items_canvas.winfo_y()+items_canvas.winfo_height()+20
    if p>=canvas2.winfo_height():
        shipment_inner_frame.config(height=p)
    else:
        shipment_inner_frame.config(height=canvas2.winfo_height())

    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    shipment_inner_frame.bind("<Configure>", lambda e: canvas2.config(scrollregion=canvas2.bbox('all')))
    resize(1,0)
    window.bind("<Configure>", lambda e: resize(e,0))


def invoice_script(path, output_text, delzero):
    def settings_writer():
        text_print(output_text, 'Creating settings file')
        with open('Settings/invoice_settings.txt', 'w', encoding='utf-8') as file:
            file.write('remove = Status, QuantityNotShipped, InvalidReason\n'
                       'shipquantity = ShipQuantity\n'
                       'date = InvoiceDate')
            file.close()
    def settings_reader():
        text_print(output_text, 'Reading settings')
        columns_dict={
            'remove': [],
            'shipquantity': [],
            'date': []
        }
        with open('Settings/invoice_settings.txt', 'r', encoding='utf-8') as file:
            lines = file.readlines()
            for line in lines:
                sp = line.split('=')
                if sp[0] == 'remove' or sp[0] == 'remove ':
                    sp[1] = sp[1].replace('\n', '')
                    degerler = sp[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        columns_dict['remove'].append(deger)
                if sp[0] == 'shipquantity' or sp[0] == 'shipquantity ':
                    sp[1] = sp[1].replace('\n', '')
                    degerler = sp[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        columns_dict['shipquantity'].append(deger)
                if sp[0] == 'date' or sp[0] == 'date ':
                    sp[1] = sp[1].replace('\n', '')
                    degerler = sp[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        columns_dict['date'].append(deger)
        text_print(output_text, str(columns_dict))
        return columns_dict
    def dir_creater():
        try:
            os.mkdir(path + '/invoice_sonuc_excel')
        except FileExistsError:pass
    def csv_reader(columns_dict):


        # Dizin içindeki tüm CSV dosyalarının isimlerini al
        csv_files = dosyalar_dictionary['invoice_csv']
        dataframes = []
        # Tüm CSV dosyalarını bir listeye oku
        def column_checker(columns_dict, dosya):
            def wait_for_enter():
                window.bind("<Return>", on_enter)
                window.wait_variable(wait_var)

            def on_enter(event):
                wait_var.set(1)

            wait_var = tk.IntVar()
            for key in columns_dict.keys():
                columns = columns_dict[key]
                boolenfirst = False
                while True:
                    df = pd.read_csv(file)
                    for col in columns:
                        try:
                            ship_quantity_list = df[col].tolist()
                            right_column = col
                            break
                        except:
                            if col == columns[-1]:
                                boolenfirst = True
                                text_print(output_text, f'{dosya} icin {key} sutunu bulunamadi lutfen tekrar denemek icin ENTER\'a basiniz...')
                                output_text.config(state=tk.NORMAL)
                                wait_for_enter()
                            continue
                    if boolenfirst == True:
                        boolenfirst = False
                        continue
                    break
            return df
        check_dict = {
            'shipquantity': columns_dict['shipquantity'],
            'date': columns_dict['date']
        }
        for file in csv_files:
            df = column_checker(check_dict, file)
            dataframes.append(df)


        # Tüm DataFrame'leri birleştir
        df_merged = pd.concat(dataframes, ignore_index=True)
        # Birleştirilmiş DataFrame'i kontrol et
        return df_merged

        # Eğer istersen birleştirilmiş DataFrame'i yeni bir CSV dosyasına kaydedebilirsin
        #df_merged.to_csv('birlesmis_dosya.csv', index=False)
        #df.to_excel('invoice_sonuc_excel/'+ file, index=False)
    def remove(df, columns):
        df.drop(columns, axis=1, inplace=True, errors='ignore')
        return df
    def ship_quantity_remove(df,columns):
        def wait_for_enter():
            window.bind("<Return>", on_enter)
            window.wait_variable(wait_var)

        def on_enter(event):
            wait_var.set(1)

        wait_var = tk.IntVar()
        boolenfirst= False
        while True:
            for col in columns:
                try:
                    ship_quantity_list = df[col].tolist()
                    right_column = col
                    break
                except:
                    if col == columns[-1]:
                        boolenfirst = True
                        text_print(output_text, 'ShipQuantity sutunu bulunamadi lutfen tekrar denemek icin ENTER\'a basiniz...')
                        output_text.config(state=tk.NORMAL)
                        wait_for_enter()
                    continue
            if boolenfirst == True:
                boolenfirst = False
                continue
            break
        if delzero != 0:
            df = df[df[right_column] != 0]
        return df
    def date_converter(df,columns):
        def wait_for_enter():
            window.bind("<Return>", on_enter)
            window.wait_variable(wait_var)

        def on_enter(event):
            wait_var.set(1)

        wait_var = tk.IntVar()
        boolenfirst= False
        while True:
            for col in columns:
                try:
                    date_list = df[col].tolist()
                    right_column = col
                    break
                except:
                    if col == columns[len(columns) - 1]:
                        boolenfirst = True
                        text_print(output_text, 'date sutunu bulunamadi lutfen tekrar denemek icin ENTER\'a basiniz...')
                        output_text.config(state=tk.NORMAL)
                        wait_for_enter()
                    continue
            if boolenfirst == True:
                boolenfirst = False
                continue
            break
        new_date_list = []
        for date in date_list:
            if '/' in date:
                date = date.split('/')
                tr_date = date[1]+'/'+date[0]+'/'+date[2]
                new_date_list.append(tr_date)
            elif ',' in date:
                date = date.split(',')
                tr_date = date[1]+'/'+date[0]+'/'+date[2]
                new_date_list.append(tr_date)
            elif '-' in date:
                date = date.split('-')
                tr_date = date[1]+'/'+date[0]+'/'+date[2]
                new_date_list.append(tr_date)
            else:
                new_date_list.append('#HATA')

        df[col] = new_date_list
        return df
    def noktavirgul(df):
        df = df.map(lambda x: str(x).replace('.', ',') if isinstance(x, (str, float)) and x != 0 and x != 0.0 else x)
        return df
    def excel_writer(df, klasor):
        df.to_excel(path + '/' + klasor +'/'+'toplu.xlsx', index=False)
    def main():
        try:
            if 'invoice_settings.txt' not in os.listdir('Settings'):
                settings_writer()
            columns_dict = settings_reader()
            dir_creater()
            df = csv_reader(columns_dict)
            df = remove(df, columns_dict['remove'])
            df = ship_quantity_remove(df, columns_dict['shipquantity'])
            df = date_converter(df, columns_dict['date'])
            df = noktavirgul(df)
            excel_writer(df, klasor='invoice_sonuc_excel')
            text_print(output_text, 'islem basariyla tamamlandi...')
            open_folder_in_explorer(path)
            sys.exit()
        except Exception:
            text_print(output_text, 'bir hata meydana geldi')
            text_print(output_text, traceback.format_exc(), color='red')
            sys.exit()
    main()


def button_invoice(canvas2):
    def resize(e, a):
        scale = main_frame_resize()
        drag_frame.config(height=175*scale)
        height = bottom_canvas.winfo_y()+bottom_canvas.winfo_height()+20
        if a:
            output_text.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                invoice_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                invoice_main_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                invoice_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                invoice_main_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox("all"))
    invoice_active_dictionary= {
        '0': 1,
    }
    def invoice_builder():
        invoice_active_dictionary['0'] = 1
        invoice_yes.configure(image=selected_image)
        invoice_no.configure(image=not_selected_image)


    def invoice_destroyer():
        invoice_active_dictionary['0'] = 0
        invoice_yes.configure(image=not_selected_image)
        invoice_no.configure(image=selected_image)

    invoice_main_frame = Frame(
        canvas2,
        background=color,
        width=750,
        height=canvas2.winfo_height()
    )
    invoice_main_frame.grid_propagate(False)
    invoice_main_frame.grid_columnconfigure(0, weight=1)
    canvas2.create_window((0,0), window=invoice_main_frame, anchor='nw')
    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    invoice_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=invoice_scrollbar.set, scrollregion=canvas2.bbox('all'))
    invoice_scrollbar.pack(side=RIGHT, fill=Y)
    top_canvas = Canvas(
        invoice_main_frame,
        border=0,
        highlightthickness=0,
        background=color
    )
    top_canvas.grid_columnconfigure(0, weight=1)
    bottom_canvas = Canvas(
        invoice_main_frame,
        border=0,
        highlightthickness=0,
        background=color
    )
    bottom_canvas.grid_columnconfigure(0, weight=1)
    invoice_title = Label(
        top_canvas,
        background=color,
        text="Invoice Program",
        font=(("JetBrainsMonoRoman Regular", 24 * -1)),
        fg=canvas2_text_color
    )
    invoice_title_line = Frame(
        top_canvas,
        height=2,
        background=line_color
    )
    invoice_cevap = Frame(top_canvas, bg=color)
    invoice_yes = Button(
        invoice_cevap,
        image = selected_image,
        relief='sunken',
        border = 0,
        background=color,
        activebackground=color,
        text='Evet',
        compound='left',
        fg=canvas2_text_color,
        activeforeground=canvas2_text_color,
        cursor='hand2',
        padx=5,
        font=("JetBrainsMonoRoman Regular", 12),
        command= lambda: invoice_builder()
    )
    #restock_yes.image = not_selected_image

    invoice_no = Button(
        invoice_cevap,
        image = not_selected_image,
        relief='sunken',
        background=color,
        activebackground=color,
        border = 0,
        text='Hayır',
        compound='left',
        fg=canvas2_text_color,
        activeforeground=canvas2_text_color,
        cursor='hand2',

        font=("JetBrainsMonoRoman Regular", 12),
        padx=5,
        command= lambda: invoice_destroyer()
    )
    invoice_yes.pack(side=LEFT, padx=15)
    invoice_no.pack(side=LEFT)
    #restock_no.image = selected_image
    invoice_question = Label(
        top_canvas,
        text='0\'lari silmek istiyor musun?',
        background=color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    save_path_label = Label(
        top_canvas,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        top_canvas,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item , w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('inv', save_path)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))

    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('inv', save_path, placeholder)
    window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, save_path, placeholder, canvas2_text_color))

    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    save_path.pack(side=LEFT, fill=X, expand=True)

    return_list = drag_drop(0,1,0,'invoice_csv',
              'Aşağıya invoice csv dosyalarını sürükleyip bırakınız:',
              bottom_canvas, padx=0, bg_image=csv_drag_drop_image, file_image=csv_icon_image, file_type='.csv')
    drag_frame = return_list[0]

    settings_label = Label(bottom_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)

    settings_height=150
    if 'invoice_settings.txt' not in os.listdir('Settings'):
        settings('Settings/invoice_settings.txt', invoice_settings_var)
    settings_text = Text(
        bottom_canvas,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )
    settings_text.bind('<Enter>',lambda e: on_text_enter(e))
    settings_text.bind('<Leave>',lambda e: on_text_leave(e))
    with open('Settings/invoice_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        settings_text.insert(tk.END, readed)
        settings_text.see(tk.END)


    baslat_button = MyButton(
        bottom_canvas,
        round=15,
        width=100,
        height=50,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    def color_change(e,c,t):
        baslat_button.config(background=c, text_color=t)
    def baslat_click(e,c,t):
        color_change(e,c,t)
        path = save_path.get(1.0, END)
        path = path.rstrip("\n")
        output(path)
    baslat_button.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black'))
    baslat_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white'))
    baslat_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color))
    baslat_button.bind("<Leave>", lambda e: color_change(e,line_color,'white'))

    output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    output_text.bind("<Enter>", on_text_enter)
    output_text.bind("<Leave>", on_text_leave)


    top_canvas.grid(column=0, row=0, sticky='we', padx=(25,0), pady=(25,0))
    bottom_canvas.grid(column=0, row=1, sticky='we', padx=(25,0), pady=(25,0))
    invoice_title.grid(column=0, row=0, sticky='w')
    invoice_title_line.grid(column=0, row=1, sticky='we')
    invoice_question.grid(column=0, row=2, sticky='w', padx=(0,0))
    invoice_cevap.grid(column=0, row=3, sticky='w', padx=(0,0), pady=(5, 0))
    save_path_label.grid(column=0, row=4, sticky='w', pady=(25,0))
    path_frame.grid(column=0, row=5, sticky='we')
    settings_label.grid(column=0, row=2, sticky='w', pady=4)
    settings_text.grid(column=0, row=3, sticky='we')
    baslat_button.grid(column=0, row=4, sticky='e', pady=(20,0))
    def invoice_script_starter(path, output_text, delzero):
        t = Thread(target=invoice_script, args=(path, output_text, delzero), daemon=True)
        t.start()
    def output(path):
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize(e, True))
        invoice_ayarlar = settings_text.get("1.0", tk.END)
        invoice_ayarlar = invoice_ayarlar.rstrip("\n")
        settings('Settings/invoice_settings.txt', invoice_ayarlar)
        delzero = invoice_active_dictionary["0"]
        if path == "Example: C:/Users/Username/Desktop/sonuc":
            output_text.insert(END, "path degeri algilanamadi, lutfen dogru bir deger girdiginizden emin olup tekrar deneyiniz.\n")
            output_text.see(END)
        else:
            invoice_script_starter(path, output_text, delzero)
    window.bind("<Configure>", lambda e: resize(e, False))


def button_converter(canvas2):
    def resize_converter(e, a):
        scale = main_frame_resize()
        for item in bottom_canvas.winfo_children():
            if type(item) == Frame:
                item.config(height=175*scale)
        height = bottom_canvas.winfo_y()+bottom_canvas.winfo_height()+20
        if a:
            convert_output_text.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                converter_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                converter_main_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                converter_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                converter_main_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox('all'))
    converter_main_frame = Frame(
        canvas2,
        background=color,
        width=750,
        height=canvas2.winfo_height()
    )
    canvas2.create_window((0,0), window=converter_main_frame, anchor='nw')
    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    converter_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=converter_scrollbar.set, scrollregion=canvas2.bbox('all'))
    converter_scrollbar.pack(side=RIGHT, fill=Y)

    converter_main_frame.grid_columnconfigure(0, weight=1)
    converter_main_frame.grid_propagate(False)

    #creating the top and bottom canvas:

    top_canvas = Canvas(
        converter_main_frame,
        background=color,
        highlightthickness=0,
        border=0
    )
    bottom_canvas = Canvas(
        converter_main_frame,
        background=color,
        highlightthickness=0,
        border=0
    )

    #top and bottom canvaslarin yerlesimi:

    top_canvas.grid(column=0, row=0, sticky='ew', padx=(25,0), pady=(20,0))
    top_canvas.grid_columnconfigure(0, weight=1)
    bottom_canvas.grid(column=0, row=1, sticky='ew', padx=(25,0), pady=0)

    #widgets:

    title = Label(
        top_canvas,
        background=color,
        fg=canvas2_text_color,
        text="Converter",
        font=(("JetBrainsMonoRoman Regular", 24 * -1))
    )
    title_line = Frame(
        top_canvas,
        height = 2,
        background=line_color,
    )
    down_arrow = PhotoImage(file=relative_to_assets('arrow_down1.png'))
    var1 = StringVar()
    var1.set('csv')
    var2 = StringVar()
    var2.set('xlsx')
    convert_choose_frame = ConvertChooser(window, top_canvas, down_arrow, var1, var2)
    def var1_changed(*args):
        items = bottom_canvas.winfo_children()
        for item in items:
            if item != convert_button.canvas:
                item.destroy()
        file_type_dictionary = {
            '.csv': {
                'bg_image': csv_drag_drop_image,
                'file_image': csv_icon_image
            },
            '.xlsx': {
                'bg_image': 0,
                'file_image': 0
            },
            '.txt': {
                'bg_image': txt_drag_drop_image,
                'file_image': txt_icon_image
            }
        }
        file_type = '.' + var1.get()
        bg_image = file_type_dictionary[file_type]['bg_image']
        file_image = file_type_dictionary[file_type]['file_image']
        drag_drop(0,1,0,'convert',
                  'Aşağıya dönüştürmek istediğiniz dosyaları sürükleyip bırakınız:',
                  bottom_canvas, padx=0, bg_image=bg_image, file_image=file_image, file_type=file_type)
        '''if var1.get() == 'csv':
            drag_drop(0,1,0,'convert',
                      'Aşağıya dönüştürmek istediğiniz dosyaları sürükleyip bırakınız:',
                      bottom_canvas, padx=0, bg_image=csv_drag_drop_image, file_image=csv_icon_image, file_type=".csv")
        elif var1.get() == 'xlsx':
            drag_drop(0,1,0,'convert',
                      'Aşağıya dönüştürmek istediğiniz dosyaları sürükleyip bırakınız:',
                      bottom_canvas, padx=0, file_type=".xlsx")'''
    var1.trace_add('write', var1_changed)

    save_path_label = Label(
        top_canvas,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        top_canvas,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item , w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('converter', save_path)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))
    bottom_canvas.grid_columnconfigure(0, weight=1)

    converter_return = drag_drop(0,1,0,'convert',
                                 'Aşağıya dönüştürmek istediğiniz dosyaları sürükleyip bırakınız:',
                                 bottom_canvas, padx=0, bg_image=csv_drag_drop_image, file_image=csv_icon_image, file_type=".csv")
    converter_dd_text = converter_return[0]
    converter_dd_frame = converter_return[1]
    convert_button = MyButton(
        bottom_canvas,
        round=15,
        width=100,
        height=50,
        text='Dönüştür',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    def convert_color_change(e,c,t):
        convert_button.config(background=c, text_color=t)
    def convert_click(e,c,t):
        convert_color_change(e, c, t)
        path = save_path.get(1.0, tk.END)
        path = path.strip('\n')
        input_type = var1.get()
        output_type = var2.get()
        output(path, input_type, output_type)

    convert_button.bind("<Button-1>", lambda e: convert_click(e, '#8AB4F8', 'black'))
    convert_button.bind("<ButtonRelease-1>", lambda e: convert_color_change(e, '#727478', 'white'))
    convert_button.bind("<Enter>", lambda e: convert_color_change(e, '#727478', canvas2_text_color))
    convert_button.bind("<Leave>", lambda e: convert_color_change(e, line_color, 'white'))

    convert_output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    #yerlesim:

    title.grid(column=0, row=0, sticky='w')
    title_line.grid(column=0, row=1, sticky='ew')
    convert_choose_frame.grid(column=0, row=2, sticky='ew')
    save_path_label.grid(column=0, row=3, sticky='w', pady=(0,0))
    path_frame.grid(column=0, row=4, sticky='we')
    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('converter', save_path, placeholder)
    window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, save_path, placeholder, canvas2_text_color))
    save_path.pack(side=LEFT, fill=X, expand=True)
    convert_button.grid(column=0, row=2, sticky='e', padx=0, pady=(15,0))

    def output(path, input_type, output_type):
        convert_output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0))
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize_converter(e, True))
        if path == 'Example: C:/Users/Username/Desktop/sonuc':
            text_print(convert_output_text, 'yanlis path')
        else:
            def converter_script_starter(path, output_text, input_type, output_type):
                t = Thread(target=converter_script, args=(path, output_text, input_type, output_type,), daemon=True)
                t.start()
            text_print(convert_output_text, path)
            text_print(convert_output_text, input_type)
            text_print(convert_output_text, output_type)
            converter_script_starter(path, convert_output_text, input_type, output_type)

    window.bind("<Configure>", lambda e: resize_converter(e, False))


def button_tsv(canvas2):
    def tsv_resize(e, a):
        scale = main_frame_resize()
        tvs_drop_frame.config(height=175*scale)
        height = alt_canvas.winfo_y()+alt_canvas.winfo_height()+20
        if a:
            tsv_output.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                tvs_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                tvs_main_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                tvs_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                tvs_main_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox('all'))
    tvs_main_frame = Frame(
        canvas2,
        bg=color,
        height=canvas2.winfo_height(),
        width=750
    )
    canvas2.create_window((0,0), anchor='nw', window=tvs_main_frame)
    canvas2.config(scrollregion=canvas2.bbox("all"))
    #tvs_main_frame.pack(side=LEFT,fill=BOTH, expand=True)
    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    tsv_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=tsv_scrollbar.set, scrollregion=canvas2.bbox('all'))
    tsv_scrollbar.pack(side=RIGHT, fill=Y)
    tvs_main_frame.grid_columnconfigure(0, weight=1)
    tvs_main_frame.grid_propagate(False)
    ust_canvas = Canvas(
        tvs_main_frame,
        border=0,
        highlightthickness=0,
        bg=color
    )
    ust_canvas.grid(column=0, row=0, sticky='we', padx=(25,0), pady=(25,0))
    ust_canvas.grid_columnconfigure(0, weight=1)
    alt_canvas = Canvas(
        tvs_main_frame,
        border=0,
        highlightthickness=0,
        bg=color,
    )
    alt_canvas.grid(column=0, row=1, sticky='we', padx=(0,0))
    alt_canvas.grid_columnconfigure(0, weight=1)
    title = Label(
        ust_canvas,
        text="TSV PROGRAMI",
        bg=color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 24 * -1)
    )
    title.grid(column=0, row=0, sticky="w")
    tsv_title_line = Frame(
        ust_canvas,
        height=2,
        bg=line_color,
        border=0,
        highlightthickness=0
    )
    tsv_title_line.grid(column=0, row=1, sticky="we")
    tsv_path_label = Label(
        ust_canvas,
        text="Aşağıya sonuçların kaydedilmesini istediğiniz dosya yolunu giriniz:",
        background=color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )
    tsv_path_label.grid(column=0, row=2, sticky="w", pady=(25,0))

    tsv_path_frame = Frame(ust_canvas, bg=color, height=30)
    tsv_path_frame.grid(column=0, row=3, sticky="we", pady=(0,25))
    tsv_path_text = Text(
        tsv_path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        bg=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    tsv_browse_button = MyButton(
        tsv_path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        tsv_path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item , w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('tsv', tsv_path_text)
    tsv_browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', tsv_path_text, tsv_browse_button))
    tsv_browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', tsv_browse_button))
    tsv_browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, tsv_browse_button))
    tsv_browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', tsv_browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))
    tsv_browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('tsv', tsv_path_text, placeholder)
    window.unbind("<Button-1>")
    tsv_path_text.pack(side=LEFT, fill=X, expand=True)
    tsv_path_text.bind("<Button-1>", lambda e: on_focus_in(e, tsv_path_text, placeholder, canvas2_text_color))
    tsv_path_text.bind("<FocusOut>", lambda e: on_focus_out(e, tsv_path_text, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, tsv_path_text, placeholder, canvas2_text_color))




    tvs_bg_image = PhotoImage(
        file=relative_to_assets('tvs_bg_rs.png')
    )
    tvs_file_image = PhotoImage(
        file=relative_to_assets('TVS_file.png')
    )

    tvs_drop_return = drag_drop(row1=0,row=1,column=0,dict_name="tsv",
                                text=".tsv uzantili dosyalarinizi asagiya surukleyip birakiniz...", parent=alt_canvas,
                                file_image=tvs_file_image, bg_image=tvs_bg_image, file_type=".tsv", padx=(25,0))
    tvs_drop_frame = tvs_drop_return[0]
    tvs_surukle_text = tvs_drop_return[1]
    baslat_button = MyButton(
        alt_canvas,
        round=15,
        width=100,
        height=50,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )

    settings_height=100
    tsv_settings_label = Label(alt_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)
    if 'tsv_settings.txt' not in os.listdir('Settings'):
        settings('Settings/tsv_settings.txt', tsv_settings_var)
    tsv_settings_text = Text(
        alt_canvas,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )
    tsv_settings_text.bind('<Enter>',lambda e: on_text_enter(e))
    tsv_settings_text.bind('<Leave>',lambda e: on_text_leave(e))
    with open('Settings/tsv_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        tsv_settings_text.insert(tk.END, readed)
        tsv_settings_text.see(tk.END)
    tsv_settings_label.grid(column=0, row=2, columnspan=2, sticky = 'w', padx=25, pady=3)
    tsv_settings_text.grid(column=0, row=3, columnspan=2, sticky = 'we', padx=(25,0), pady=5)


    baslat_button.grid(column=0, row=4, sticky='e', padx=(0,0), pady=(15,0))
    def color_change(e,c,t):
        baslat_button.config(background=c, text_color=t)
    def baslat_click(e,c,t):
        color_change(e,c,t)
        path = tsv_path_text.get(1.0, END)
        path = path.rstrip("\n")
        output(path)
    baslat_button.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black'))
    baslat_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white'))
    baslat_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color))
    baslat_button.bind("<Leave>", lambda e: color_change(e,line_color,'white'))
    tsv_liste = [canvas, canvas2, button_1, button_2, button_3, button_4, button_5, tvs_main_frame]
    tsv_output = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )

    def tsv_script_starter(path, output_text):
        t = Thread(target=tsv_script, args=(path, output_text,), daemon=True)
        t.start()
    def output(path):
        tsv_output.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(),0))
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: tsv_resize(e, True))
        tsv_ayarlar = tsv_settings_text.get("1.0", tk.END)
        tsv_ayarlar = tsv_ayarlar.rstrip("\n")
        settings('Settings/tsv_settings.txt', tsv_ayarlar)
        if path == "Example: C:/Users/Username/Desktop/sonuc":
            tsv_output.insert(END, "path degeri algilanamadi, lutfen dogru bir deger girdiginizden emin olup tekrar deneyiniz.\n")
            tsv_output.see(END)
        else:
            tsv_output.insert(END, str(dosyalar_dictionary["tsv"])+"\n")
            tsv_output.insert(END, path+"\n")
            tsv_output.see(END)
            tsv_script_starter(path, tsv_output)


    canvas2.config(scrollregion=canvas2.bbox('all'))
    window.bind("<Configure>", lambda e: tsv_resize(e, False))

def button_costupdater(canvas2):
    def resize(e, a):
        scale = main_frame_resize()
        height = bottom_canvas.winfo_y()+bottom_canvas.winfo_height()+20
        drag_frame.config(height=175*scale)
        if a:
            output_text.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                cost_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                cost_main_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                cost_main_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                cost_main_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox('all'))
    def new_active():
        if 'costupdater2_settings.txt' not in os.listdir('Settings'):
            settings('Settings/costupdater2_settings.txt', costupdater2_settings_var)
        with open('Settings/costupdater2_settings.txt', 'r', encoding='utf-8') as file:
            readed = file.read()
            settings_text.delete(1.0, tk.END)
            settings_text.insert(tk.END, readed)
            settings_text.see(tk.END)
        baslat_button.bind("<Button-1>", lambda e: baslat2_click(e,'#8AB4F8','black'))
    def new_deactive():
        if 'costupdater_settings.txt' not in os.listdir('Settings'):
            settings('Settings/costupdater_settings.txt', costupdater_settings_var)
        with open('Settings/costupdater_settings.txt', 'r', encoding='utf-8') as file:
            readed = file.read()
            settings_text.delete(1.0, tk.END)
            settings_text.insert(tk.END, readed)
            settings_text.see(tk.END)
        baslat_button.bind("<Button-1>", lambda e: baslat_click(e, '#8AB4F8', 'black'))
    cost_main_frame = Frame(
        canvas2,
        bg=color,
        height=canvas2.winfo_height(),
        width=750
    )
    canvas2.create_window((0,0), window=cost_main_frame, anchor='nw')

    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)
    costupdater_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=costupdater_scrollbar.set, scrollregion=canvas2.bbox('all'))
    costupdater_scrollbar.pack(side=RIGHT, fill=Y)

    cost_main_frame.grid_columnconfigure(0, weight=1)
    cost_main_frame.grid_propagate(False)


    #creating the top and bottom canvas:

    top_canvas = Canvas(
        cost_main_frame,
        background=color,
        highlightthickness=0,
        border=0
    )
    bottom_canvas = Canvas(
        cost_main_frame,
        background=color,
        highlightthickness=0,
        border=0
    )

    #top and bottom canvaslarin yerlesimi:

    top_canvas.grid(column=0, row=0, sticky='ew', padx=(25,0), pady=(20,0))
    top_canvas.grid_columnconfigure(0, weight=1)
    bottom_canvas.grid(column=0, row=1, sticky='ew', padx=(25,0), pady=0)
    bottom_canvas.grid_columnconfigure(0, weight=1)

    #widgets:
    title_frame = Frame(
        top_canvas,
        background=color
    )
    title = Label(
        title_frame,
        background=color,
        fg=canvas2_text_color,
        text="Cost Updater",
        font=(("JetBrainsMonoRoman Regular", 24 * -1))
    )
    new_switch = SwitchButton(
        parent=title_frame,
        border=0,
        highlightthickness=0,
        active_function=lambda: new_active(),
        pasif_function=lambda: new_deactive(),
        f='red',
        s='green',
        status=True
    )
    title_line = Frame(
        top_canvas,
        height = 2,
        background=line_color,
    )

    save_path_label = Label(
        top_canvas,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        top_canvas,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('cos', save_path)
    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))

    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('cos', save_path, placeholder)
    window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    window.bind("<Button-1>", lambda e: on_click_outside(e, save_path, placeholder, canvas2_text_color))

    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    save_path.pack(side=LEFT, fill=X, expand=True)

    return_list = drag_drop(0,1,0,'costupdater',
                            'Aşağıya ilgili csv dosyasini surukleyip birakiniz:',
                            bottom_canvas, padx=0, bg_image=csv_drag_drop_image, file_image=csv_icon_image, file_type='.csv')
    drag_frame = return_list[0]

    settings_label = Label(bottom_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)

    settings_height=225
    if 'costupdater2_settings.txt' not in os.listdir('Settings'):
        settings('Settings/costupdater2_settings.txt', costupdater2_settings_var)
    settings_text = Text(
        bottom_canvas,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )
    settings_text.bind('<Enter>',lambda e: on_text_enter(e))
    settings_text.bind('<Leave>',lambda e: on_text_leave(e))
    with open('Settings/costupdater2_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        settings_text.insert(tk.END, readed)
        settings_text.see(tk.END)


    baslat_button = MyButton(
        bottom_canvas,
        round=15,
        width=100,
        height=50,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    def color_change(e,c,t):
        baslat_button.config(background=c, text_color=t)
    def baslat2_click(e,c,t):
        color_change(e,c,t)
        path = save_path.get(1.0, END)
        path = path.rstrip("\n")
        print("baslat2 calisti")
        output2(path)
    def baslat_click(e,c,t):
        color_change(e,c,t)
        path = save_path.get(1.0, END)
        path = path.rstrip("\n")
        print("baslat calisti")
        output(path)
    baslat_button.bind("<Button-1>", lambda e: baslat2_click(e,'#8AB4F8','black'))
    baslat_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white'))
    baslat_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color))
    baslat_button.bind("<Leave>", lambda e: color_change(e,line_color,'white'))

    output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    output_text.bind("<Enter>", on_text_enter)
    output_text.bind("<Leave>", on_text_leave)

    title.pack(side='left')
    new_switch.pack(side='right')
    title_frame.grid(column=0, row=0, sticky='ew')
    title_line.grid(column=0, row=1, sticky='ew')

    top_canvas.grid(column=0, row=0, sticky='we', padx=(25,0), pady=(25,0))
    bottom_canvas.grid(column=0, row=1, sticky='we', padx=(25,0), pady=(25,0))
    save_path_label.grid(column=0, row=2, sticky='w', pady=(25,0))
    path_frame.grid(column=0, row=3, sticky='we')
    settings_label.grid(column=0, row=2, sticky='w', pady=4)
    settings_text.grid(column=0, row=3, sticky='we')
    baslat_button.grid(column=0, row=4, sticky='e', pady=(20,0))
    def costupdater2_script_starter(path, output_text):
        t = Thread(target=costupdater2_script, args=(path, output_text), daemon=True)
        t.start()
    def costupdater_script_starter(path, output_text):
        t = Thread(target=costupdater_script, args=(path, output_text), daemon=True)
        t.start()
    def output2(path):
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize(e, True))
        
        costupdater_ayarlar = settings_text.get("1.0", tk.END).rstrip("\n")
        settings('Settings/costupdater2_settings.txt', costupdater_ayarlar)
        
        if path == "Example: C:/Users/Username/Desktop/sonuc" or path == "":
            text_print(output_text, "Hata: Dosya yolu algılanamadı, lütfen geçerli bir klasör seçin.", color="red")
            return
            
        csv_files = dosyalar_dictionary.get('costupdater', [])
        if not csv_files:
            text_print(output_text, "Hata: İşlenecek CSV dosyası sürüklemediniz.", color="red")
            return
            
        input_file = csv_files[0]
        
        def update_progress(msg: str):
            output_text.after(0, lambda: text_print(output_text, msg))

        def run_in_thread():
            try:
                result = process_costupdater2(
                    input_file, 
                    path, 
                    costupdater_ayarlar, 
                    progress_callback=update_progress
                )
                output_text.after(0, lambda: text_print(output_text, result["message"], color='#90EE90'))
                output_text.after(0, lambda: open_folder_in_explorer(path))
            except Exception as e:
                output_text.after(0, lambda: text_print(output_text, f"Hata: {str(e)}", color='red'))

        conversion_thread = threading.Thread(target=run_in_thread, daemon=True)
        conversion_thread.start()

    def output(path):
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize(e, True))
        
        costupdater_ayarlar = settings_text.get("1.0", tk.END).rstrip("\n")
        settings('Settings/costupdater_settings.txt', costupdater_ayarlar)
        
        if path == "Example: C:/Users/Username/Desktop/sonuc" or path == "":
            text_print(output_text, "Hata: Dosya yolu algılanamadı, lütfen geçerli bir klasör seçin.", color="red")
            return
            
        csv_files = dosyalar_dictionary.get('costupdater', [])
        if not csv_files:
            text_print(output_text, "Hata: İşlenecek CSV dosyası sürüklemediniz.", color="red")
            return
            
        input_file = csv_files[0]
        
        def update_progress(msg: str):
            output_text.after(0, lambda: text_print(output_text, msg))

        def run_in_thread():
            try:
                result = process_costupdater(
                    input_file, 
                    path, 
                    costupdater_ayarlar, 
                    progress_callback=update_progress
                )
                output_text.after(0, lambda: text_print(output_text, result["message"], color='#90EE90'))
                output_text.after(0, lambda: open_folder_in_explorer(path))
            except Exception as e:
                output_text.after(0, lambda: text_print(output_text, f"Hata: {str(e)}", color='red'))

        conversion_thread = threading.Thread(target=run_in_thread, daemon=True)
        conversion_thread.start()

    window.bind("<Configure>", lambda e: resize(e, False))


def button_updater(canvas2):
    checkforupdates_label = Label(
        canvas2,
        text='Güncellemeleri kontrol etmek için Check For Updates butonuna tıklayın:',
        font=("JetBrainsMonoRoman Regular", 12),
        bg=color,
        fg=canvas2_text_color
    )
    checkforupdates = MyButton(
        canvas2,
        round=5,
        width=150,
        height=45,
        text='Check For Updates',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    doyouwanna_frame = Frame(
        canvas2,
        bg=color,
    )
    doyouwanna_label = Label(
        doyouwanna_frame,
        bg=color,
        fg=canvas2_text_color,
        text="Yeni bir güncelleme bulundu! Yüklemek istiyor musun?",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    release_notes = MyButton(
        doyouwanna_frame,
        round=5,
        width=125,
        height=30,
        text="Release Notes",
        background=line_color,
        text_color='white',
        align_text='center'
    )
    yes_button = MyButton(
        doyouwanna_frame,
        round=5,
        width=75,
        height=30,
        text='Yükle',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    yes_button.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', yes_button))
    yes_button.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, yes_button))
    yes_button.bind("<Leave>", lambda e: color_change(e,line_color,'white', yes_button))
    release_notes.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', release_notes))
    release_notes.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, release_notes))
    release_notes.bind("<Leave>", lambda e: color_change(e,line_color,'white', release_notes))
    doyouwanna_frame.grid_columnconfigure(0, weight=1)
    doyouwanna_label.grid(column=0, row=0, stick='w')
    release_notes.grid(column=0, row=1, stick='e', padx=(0, 5))
    yes_button.grid(column=1, row=1, stick='e')
    progress_label = Label(window, bg=color, fg=canvas2_text_color, text="İndiriliyor...")
    progress_bar = ttk.Progressbar(window, orient="horizontal", mode='determinate')
    def is_connected_starter(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, doyouwanna_label, yes_button, release_notes):
        t = Thread(target=is_connected, args=(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, doyouwanna_label, yes_button, release_notes), daemon=True)
        t.start()
    def color_change(e,c,t,i):
        i.config(background=c, text_color=t)
    def baslat_click(e,c,t,i):
        color_change(e,c,t,i)
        is_connected_starter(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, doyouwanna_label, yes_button, release_notes)
    checkforupdates.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black', checkforupdates))
    checkforupdates.bind("<ButtonRelease-1>", lambda e: color_change(e,'#727478','white', checkforupdates))
    checkforupdates.bind("<Enter>", lambda e: color_change(e,'#727478',canvas2_text_color, checkforupdates))
    checkforupdates.bind("<Leave>", lambda e: color_change(e,line_color,'white', checkforupdates))
    canvas2.grid_columnconfigure(0, weight=1)
    checkforupdates_label.grid(column=0, row=0, sticky='w')
    canvas2.update_idletasks()
    checkforupdates.grid(column=0, row=1, padx=(checkforupdates_label.winfo_width()-150,0), pady=(10,0), sticky='w')




def invoicefinder_script(path, invoice_folder, user_input_date, output_text, allinvoices):
    def indexFinder(item, liste):
        index_list = []
        for a, z in enumerate(liste):
            if z == item:
                index_list.append(a)
        return index_list
    def resource_excel_reader():
        text_print(output_text, 'Sağlanan excel dosyası okunuyor...')
        file_path = dosyalar_dictionary['invoice_finder'][0]
        df = pd.read_excel(file_path, header=None)
        temporary_lines = df.values.tolist()
        lines = []
        dictionary = {}
        text_print(output_text, 'Dosyadan Sku ve Quantity değerleri ayrıştırılıyor...')
        for line in temporary_lines:
            lines.append(str(line[0]))
        for i, line in enumerate(lines):
            if line.count('_') >=3:
                number_list = []
                a = i+1
                for item in lines[i:len(lines)]:
                    if 'FNSKU' in item:
                        number_list = [lines[a], lines[a+1]]
                        break
                    a+=1
                dictionary[line] = number_list
        upc_list = {}
        return_dictionary = {}
        excel_write_dictionary = {}
        atlanan_asinler = []
        for asin in dictionary.keys():
            number_list = dictionary[asin]
            number = '#YOK'
            if '-' in number_list[1]:
                number = number_list[0]
            elif '+' in number_list[1]:
                number = number_list[1].split('+')[0]

            split_asin = asin.split('_')
            upc = split_asin[1]
            pk = split_asin[2]
            pka = pk.replace('PK', '')
            pka = int(pka)
            deger = int(number) * pka
            excel_write_dictionary[asin] = {
                'upc': upc,
                'pk': pk,
                'amazonshipquantity': int(deger),
                'invoice quantity': '',
                'item number': '',
                'invoice number': '',
                'invoice each': '',
                'invoice date': '',
                'Yapildi/Yapilmadi': '',
                'Fark': ''
            }
            if upc not in upc_list.keys():
                upc_list[upc] = asin
                return_dictionary[asin] = {
                    'upc': float(upc),
                    'pk': pk,
                    'amazonshipquantity': int(deger)
                }
            else:
                atlanan_asinler.append(asin)
                asin = upc_list[upc]
                return_dictionary[asin]['amazonshipquantity'] = int(return_dictionary[asin]['amazonshipquantity']) + deger

        return [return_dictionary, excel_write_dictionary, atlanan_asinler, upc_list]
    def allinvoices_excel_reader(source_dictionary, excel_dictionary, atlanan_asinler, upc_list):
        text_print(output_text, 'ALL INVOICES excel dosyası okunuyor...')
        df = pd.read_excel(allinvoices)
        upcs = df['Upc'].tolist()
        shipquantity = df['ShipQuantity'].tolist()
        shipitem = df['ShipItem'].tolist()
        invoice_number = df['InvoiceNumber'].tolist()
        date = df['Date'].tolist()

        user_date = pd.to_datetime(user_input_date, format='%d.%m.%Y')
        invoices = os.listdir(invoice_folder)
        def file_finder(a):
            if temporary_dictionary:
                max_date_key = max(temporary_dictionary, key=lambda x: temporary_dictionary[x]['date'])
                invoice_number_tosearch = str(temporary_dictionary[max_date_key]['invoice_number'])
                for file in invoices:
                    if invoice_number_tosearch in file:
                        shutil.copy2(f'{invoice_folder}/{file}', f'{path}/{file}')
                        text_print(output_text, f'{file}: {upc}')
                        dict_a['invoice'].append(invoice_number_tosearch)
                b = int(temporary_dictionary[max_date_key]['shipquantity'])
                a = b + a
                dict_a['itemid'].append(temporary_dictionary[max_date_key]['shipitem'])
                date_temp = temporary_dictionary[max_date_key]['date']
                formatted_date = date_temp.strftime('%d-%m-%Y')
                dict_a['date'].append(formatted_date)
                temporary_dictionary.pop(max_date_key)
                amazonshipquantity = int(source_dictionary[SKU]['amazonshipquantity'])
                #text_print(output_text, f'amazonshipquantity: {amazonshipquantity}, toplam invoice: {a}, bu invoice: {b}')
                dict_a['a'] = a
                dict_a['b'].append(b)
                if a < amazonshipquantity and temporary_dictionary.keys() != []:
                    file_finder(a)
        text_print(output_text, 'Her bir UPC için Invoice numarası bulunuyor ve pdf dosyalarında aratılıyor...')
        for SKU in source_dictionary.keys():
            upc = source_dictionary[SKU]['upc']
            index_list = indexFinder(upc, upcs)
            temporary_dictionary = {}
            if index_list:
                for index in index_list:
                    if date[index] <= user_date:
                        temporary_dictionary[index] = {
                            'shipquantity': shipquantity[index],
                            'shipitem': shipitem[index],
                            'invoice_number': invoice_number[index],
                            'date': date[index],
                        }

                a = 0
                dict_a = {}
                dict_a['invoice'] = []
                dict_a['itemid'] = []
                dict_a['date'] = []
                dict_a['b'] = []
                file_finder(a)

                a = dict_a['a']
                b_list = dict_a['b']
                invoicelar = dict_a['invoice']
                itemid = dict_a['itemid']
                invoicedate = dict_a['date']


                invoice_each_string = ''
                for oge in b_list:
                    if invoice_each_string == '':
                        invoice_each_string = str(oge)
                    else:
                        invoice_each_string = invoice_each_string+ ', ' +str(oge)


                invoice_date_string = ''
                for date_temp in invoicedate:
                    if invoice_date_string == '':
                        invoice_date_string = date_temp
                    else:
                        invoice_date_string = invoice_date_string+ ', ' +date_temp


                invoice_item_string = ''
                for id in itemid:
                    if invoice_item_string == '':
                        invoice_item_string = str(int(id))
                    else:
                        invoice_item_string = invoice_item_string+ ', ' +str(int(id))


                invoice_string = ''
                for invoice in invoicelar:
                    if invoice_string == '':
                        invoice_string = invoice
                    else:
                        invoice_string = invoice_string+ ', ' +invoice
                excel_dictionary[SKU]['invoice quantity'] = a
                excel_dictionary[SKU]['invoice number'] = invoice_string
                excel_dictionary[SKU]['item number'] = invoice_item_string
                excel_dictionary[SKU]['invoice date'] = invoice_date_string
                excel_dictionary[SKU]['Yapildi/Yapilmadi'] = 'Yapildi'
                excel_dictionary[SKU]['invoice each'] = invoice_each_string
                fark = int(excel_dictionary[SKU]['invoice quantity']) - int(excel_dictionary[SKU]['amazonshipquantity'])
                if fark > 0:
                    excel_dictionary[SKU]['Fark'] = f'+{fark}'
                else:
                    excel_dictionary[SKU]['Fark'] = fark
            else:
                excel_dictionary[SKU]['Yapildi/Yapilmadi'] = 'Yapilmadi'
            if excel_dictionary[SKU]['invoice number'] == '':
                excel_dictionary[SKU]['Yapildi/Yapilmadi'] = 'Yapilmadi'
        for asin in atlanan_asinler:
            gercek_asin = upc_list[excel_dictionary[asin]['upc']]
            excel_dictionary[asin]['invoice each'] = excel_dictionary[gercek_asin]['invoice each']
            excel_dictionary[asin]['item number'] = excel_dictionary[gercek_asin]['item number']
            excel_dictionary[asin]['invoice number'] = excel_dictionary[gercek_asin]['invoice number']
            excel_dictionary[asin]['invoice quantity'] = excel_dictionary[gercek_asin]['invoice quantity']
            excel_dictionary[asin]['invoice date'] = excel_dictionary[gercek_asin]['invoice date']
            excel_dictionary[asin]['Yapildi/Yapilmadi'] = excel_dictionary[gercek_asin]['Yapildi/Yapilmadi']
            excel_dictionary[asin]['Fark'] = excel_dictionary[gercek_asin]['Fark']
        text_print(output_text, 'Son excel dosyası oluşturuluyor ve yazdırılıyor...')
        df = pd.DataFrame.from_dict(excel_dictionary, orient='index').reset_index()

        # "index" sütununu "SKU" olarak yeniden adlandırma
        df = df.rename(columns={'index': 'SKU'})
        df.to_excel(f'{path}/sonexcel.xlsx', index=False)
        text_print(output_text, 'excel dosyasi basariyla kaydedildi!')

    def main():
        try:
            return_list = resource_excel_reader()
            source_dictionary = return_list[0]
            excel_write_dictionary = return_list[1]
            atlanan_asinler = return_list[2]
            upc_list = return_list[3]
            try:
                allinvoices_excel_reader(source_dictionary, excel_write_dictionary, atlanan_asinler, upc_list)
                text_print(output_text, 'Operasyon Tamamlandı!')
                open_folder_in_explorer(path)
            except:
                text_print(output_text, 'all invoice excelini okurken ve gerekli islemleri yaparken bir hata meydana geldi :/')
                text_print(output_text, traceback.format_exc(), color='red')
        except:
            text_print(output_text, 'sağlanan kaynak dosyasını okurken bir hata meydana geldi :/')
            text_print(output_text, traceback.format_exc(), color='red')
    main()


def invoicefinderupc_script(path, invoice_folder, user_input_upc, user_input_month, output_text, allinvoices):
    def indexFinder(item, liste):
        index_list = []
        for a, z in enumerate(liste):
            if z == item:
                index_list.append(a)
        return index_list
    def allinvoices_excel_reader():
        text_print(output_text, 'ALL INVOICES excel dosyası okunuyor...')
        df = pd.read_excel(allinvoices)
        upcs = df['Upc'].tolist()
        invoice_number = df['InvoiceNumber'].tolist()
        date = df['Date'].tolist()
        now_date = pd.Timestamp.now()
        int_month = int(user_input_month)
        if int_month != 0:
            before_fourteen_months = now_date - pd.DateOffset(months=int_month)
            text_print(output_text, f'{int_month} ay öncesine kadar olan invoice numberlar bulunuyor ve dizinde aratılarak kaydediliyor...')
        else:
            text_print(output_text, f'tüm invoice numberlar bulunuyor ve dizinde aratılarak kaydediliyor...')


        split_upc = user_input_upc.split(',')
        if len(split_upc) == 1:
            upc = split_upc[0]
            upc = upc.replace(' ', '')
            f_upc = float(upc)
            print(f_upc)
            indexes = indexFinder(f_upc, upcs)
            for index in indexes:
                if int_month != 0 :
                    if date[index] > before_fourteen_months:
                        print(invoice_number[index])
                        for file in os.listdir(invoice_folder):
                            if str(invoice_number[index]) in file:
                                shutil.copy2(f"{invoice_folder}/{file}", f"{path}/{file}")
                else:
                    print(invoice_number[index])
                    for file in os.listdir(invoice_folder):
                        if str(invoice_number[index]) in file:
                            shutil.copy2(f"{invoice_folder}/{file}", f"{path}/{file}")
        else:
            for upc in split_upc:
                upc = upc.replace(' ', '')
                f_upc = float(upc)
                print(f_upc)
                indexes = indexFinder(f_upc, upcs)
                for index in indexes:
                    if int_month != 0 :
                        if date[index] > before_fourteen_months:
                            print(invoice_number[index])
                            for file in os.listdir(invoice_folder):
                                if str(invoice_number[index]) in file:
                                    try:
                                        os.mkdir(f"{path}/{str(upc)}")
                                    except FileExistsError:pass
                                    shutil.copy2(f"{invoice_folder}/{file}", f"{path}/{str(upc)}/{file}")
                    else:
                        print(invoice_number[index])
                        for file in os.listdir(invoice_folder):
                            if str(invoice_number[index]) in file:
                                try:
                                    os.mkdir(f"{path}/{str(upc)}")
                                except FileExistsError:pass
                                shutil.copy2(f"{invoice_folder}/{file}", f"{path}/{str(upc)}/{file}")
    def main():
        try:
            allinvoices_excel_reader()
        except:
            text_print(output_text, 'All invoice excel dosyasi okunurken veya pdf dosyalari bulunmaya calisirken bir hata meydana geldi!')
            text_print(output_text, traceback.format_exc(), color='red')
        text_print(output_text, 'işlem tamamlandı!')
        open_folder_in_explorer(path)
    main()
def button_invoicefinder(canvas2):
    def invoicefinder_resize(e, a):
        scale = main_frame_resize()
        invoice_finder_drop_frame.config(height=175*scale)
        height = bottom_frame.winfo_y()+bottom_frame.winfo_height()+20
        if a:
            output_text.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                inner_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                inner_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                inner_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                inner_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox('all'))
    inner_frame = Canvas(
        canvas2,
        width=750,
        background=color,
        border=0,
        height=canvas2.winfo_height(),
        highlightthickness=0
    )
    canvas2.create_window((0,0), window=inner_frame, anchor='nw')
    invoicefinder_scrollbar = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.config(yscrollcommand=invoicefinder_scrollbar.set)
    invoicefinder_scrollbar.pack(side=RIGHT, fill=Y)
    top_frame = Frame(
        inner_frame,
        background=color,
    )
    bottom_frame = Frame(
        inner_frame,
        background=color
    )
    title_frame = Frame(
        top_frame,
        background=color
    )
    title = Label(
        title_frame,
        background=color,
        fg=canvas2_text_color,
        text="Invoice Finder",
        font=("JetBrainsMonoRoman Regular", 24 * -1)
    )
    upc_switch = SwitchButton(
        parent=title_frame,
        border=0,
        highlightthickness=0,
        active_function=lambda: upc_active(),
        pasif_function=lambda: upc_deactive(),
        f='red',
        s='green',
        status=True
    )
    title.pack(side='left')
    upc_switch.pack(side='right')
    title_line = Frame(
        top_frame,
        height=2,
        bg=line_color
    )
    save_path_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Sonuçların kaydedilmesini istediğiniz klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    path_frame = Frame(
        top_frame,
        background=color,
        height=30
    )
    save_path = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    invoice_path_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Invoice Pdf'lerinin bulunduğu klasörün yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    invoice_path_frame = Frame(
        top_frame,
        background=color,
        height=30
    )
    invoice_save_path = Text(
        invoice_path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    invoice_browse_button = MyButton(
        invoice_path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    invoice_save_button = MyButton(
        invoice_path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    allinvoices_path_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Butun invoiceleri iceren excel dosyasinin yolunu giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    allinvoices_path_frame = Frame(
        top_frame,
        background=color,
        height=30
    )
    allinvoices_save_path = Text(
        allinvoices_path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg='#747474',
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    allinvoices_browse_button = MyButton(
        allinvoices_path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    allinvoices_save_button = MyButton(
        allinvoices_path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    invoice_date_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Bir tarih degeri giriniz:",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    invoice_date_text = Text(
        top_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg="#747474",
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    invoice_upc_text = Text(
        top_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg="#747474",
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    invoice_month_label = Label(
        top_frame,
        background=color,
        fg=canvas2_text_color,
        text="Kaç ay öncesinin invoiceları çekilsin giriniz (hepsi için 0 yazınız):",
        font=("JetBrainsMonoRoman Regular", 12),
    )
    invoice_month_text = Text(
        top_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg="#747474",
        background=line_color,
        border=0,
        pady=4,
        insertbackground='#c0c0c0'
    )
    buttons_frame = Frame(
        bottom_frame,
        bg=color,
    )
    baslat_button = MyButton(
        buttons_frame,
        round=12,
        width=100,
        height=40,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    yonerge_button = MyButton(
        buttons_frame,
        round=12,
        width=100,
        height=40,
        text='Yönerge',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )

    def upc_deactive():
        invoice_date_text.grid_forget()
        invoice_upc_text.grid(column=0, row=9, sticky='we')
        invoice_month_label.grid(column=0, row=10, sticky='w')
        invoice_month_text.grid(column=0, row=11, sticky='we')
        invoice_date_text.config(state=NORMAL)
        invoice_upc_text.config(state=NORMAL)
        invoice_date_label.config(text='Upc değer(ler)ini giriniz:')
        invoice_finder_drop_frame.grid_forget()
        invoice_finder_surukle_text.grid_forget()

    def upc_active():
        invoice_upc_text.grid_forget()
        invoice_month_label.grid_forget()
        invoice_month_text.grid_forget()
        invoice_date_text.grid(column=0, row=9, sticky='we')
        invoice_date_text.config(state=NORMAL)
        invoice_upc_text.config(state=NORMAL)
        invoice_date_label.config(text="Bir tarih degeri giriniz:")
        invoice_finder_surukle_text.grid(column=0, row=0, sticky='w', pady=10)
        invoice_finder_drop_frame.grid(column=0, row=1, sticky='we')


    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_click_excel(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_excel(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b, name, save_path):
        browse_color_change(event,c,t,b)
        placeholder_saver(name, save_path)
    def baslat_click(event, c, t, b):
        b.config(background=c, text_color=t)
        path = save_path.get(1.0, END).strip('\n')
        invoice_folder = invoice_save_path.get(1.0, END).strip('\n')
        date = invoice_date_text.get(1.0, END).strip('\n')
        upc = invoice_upc_text.get(1.0, END).strip('\n')
        allinvoices = allinvoices_save_path.get(1.0, END).strip('\n')
        month = invoice_month_text.get(1.0, END).strip('\n')
        output(path, invoice_folder, date, upc, month, allinvoices)
    def yonerge_click(event, c, t, b):
        b.config(background=c, text_color=t)
        yonerge_window = Tk()
        yonerge_window.geometry('600x400')
        try:
            yonerge_window.iconbitmap('assets/icon.ico')
        except:pass
        yonerge_window.title('Invoice Finder Programı Yönergeleri!')
        content_canvas = Canvas(
            yonerge_window,
            highlightthickness=0,
            border=0,
            bg=color
        )
        content_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        main_text = Text(
            content_canvas,
            bg=color,
            fg=canvas2_text_color,
            font=('JetBrainsMonoRoman Regular', 12),
            wrap='word',
            border=0
        )
        main_text.pack(side=LEFT, fill=BOTH, expand=True, padx=25, pady=25)
        with open('Settings/invoicefinder_yonergeler.txt', encoding='UTF-8') as file:
            z = file.read()
            main_text.insert(tk.END, z)
            main_text.config(state=DISABLED)


    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', save_path, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))


    invoice_browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', invoice_save_path, invoice_browse_button))
    invoice_browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', invoice_browse_button))
    invoice_browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, invoice_browse_button))
    invoice_browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', invoice_browse_button))

    allinvoices_browse_button.bind("<Button-1>", lambda e: browse_click_excel(e,'#8AB4F8','black', allinvoices_save_path, allinvoices_browse_button))
    allinvoices_browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', allinvoices_browse_button))
    allinvoices_browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, allinvoices_browse_button))
    allinvoices_browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', allinvoices_browse_button))

    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button, 'fin', save_path))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))

    invoice_save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', invoice_save_button, 'invoice_folder', invoice_save_path))
    invoice_save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', invoice_save_button))
    invoice_save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, invoice_save_button))
    invoice_save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', invoice_save_button))

    allinvoices_save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', allinvoices_save_button, 'all_invoices', allinvoices_save_path))
    allinvoices_save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', allinvoices_save_button))
    allinvoices_save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, allinvoices_save_button))
    allinvoices_save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', allinvoices_save_button))

    baslat_button.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black', baslat_button))
    baslat_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', baslat_button))
    baslat_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, baslat_button))
    baslat_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', baslat_button))

    yonerge_button.bind("<Button-1>", lambda e: yonerge_click(e,'#8AB4F8','black', yonerge_button))
    yonerge_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', yonerge_button))
    yonerge_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, yonerge_button))
    yonerge_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', yonerge_button))


    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    date_placeholder = "GG.AA.YYYY"
    upc_placeholder = 'Example: 000000000000, 111111111111'
    month_placeholder = 'Example: 14'
    path_text_function('fin', save_path, placeholder)
    path_text_function('invoice_folder', invoice_save_path, placeholder)
    path_text_function('all_invoices', allinvoices_save_path, placeholder)
    window.unbind("<Button-1>")
    save_path.bind("<Button-1>", lambda e: on_focus_in(e, save_path, placeholder, canvas2_text_color))
    save_path.bind("<FocusOut>", lambda e: on_focus_out(e, save_path, placeholder, canvas2_text_color))
    invoice_save_path.bind("<Button-1>", lambda e: on_focus_in(e, invoice_save_path, placeholder, canvas2_text_color))
    invoice_save_path.bind("<FocusOut>", lambda e: on_focus_out(e, invoice_save_path, placeholder, canvas2_text_color))
    allinvoices_save_path.bind("<Button-1>", lambda e: on_focus_in(e, allinvoices_save_path, placeholder, canvas2_text_color))
    allinvoices_save_path.bind("<FocusOut>", lambda e: on_focus_out(e, allinvoices_save_path, placeholder, canvas2_text_color))
    invoice_date_text.bind("<Button-1>", lambda e: on_focus_in(e, invoice_date_text, date_placeholder, canvas2_text_color))
    invoice_date_text.bind("<FocusOut>", lambda e: on_focus_out(e, invoice_date_text, date_placeholder, canvas2_text_color))
    invoice_upc_text.bind("<Button-1>", lambda e: on_focus_in(e, invoice_upc_text, upc_placeholder, canvas2_text_color))
    invoice_upc_text.bind("<FocusOut>", lambda e: on_focus_out(e, invoice_upc_text, upc_placeholder, canvas2_text_color))
    invoice_month_text.bind("<Button-1>", lambda e: on_focus_in(e, invoice_month_text, month_placeholder, canvas2_text_color))
    invoice_month_text.bind("<FocusOut>", lambda e: on_focus_out(e, invoice_month_text, month_placeholder, canvas2_text_color))
    invoice_month_text.insert(END, '14')
    invoice_month_text.config(fg=canvas2_text_color)
    invoice_date_text.insert(END, date_placeholder)
    all_placeholders = [
        [save_path, placeholder], [invoice_save_path, placeholder],
        [allinvoices_save_path, placeholder], [invoice_date_text, date_placeholder],
        [invoice_upc_text, upc_placeholder], [invoice_month_text, month_placeholder]
    ]
    window.bind("<Button-1>", lambda e: on_click_outside(e, all_placeholders, placeholder, canvas2_text_color))


    browse_button.pack(side=RIGHT, padx=(8,0))
    save_button.pack(side=RIGHT, padx=(8,0))
    save_path.pack(side=LEFT, fill=X, expand=True)

    invoice_browse_button.pack(side=RIGHT, padx=(8,0))
    invoice_save_button.pack(side=RIGHT, padx=(8,0))
    invoice_save_path.pack(side=LEFT, fill=X, expand=True)

    allinvoices_browse_button.pack(side=RIGHT, padx=(8,0))
    allinvoices_save_button.pack(side=RIGHT, padx=(8,0))
    allinvoices_save_path.pack(side=LEFT, fill=X, expand=True)

    inner_frame.grid_columnconfigure(0, weight=1)
    inner_frame.grid_propagate(False)
    top_frame.grid(column=0, row=0, sticky='we', padx=(25,0), pady=(20,0))
    bottom_frame.grid(column=0, row=1, sticky='we', padx=(25,0))
    top_frame.grid_columnconfigure(0, weight=1)
    bottom_frame.grid_columnconfigure(0, weight=1)
    title_frame.grid(column=0, row=0, sticky='we')
    title_line.grid(column=0, row=1, sticky='we')
    save_path_label.grid(column=0, row=2, sticky='w', pady=(20, 0))
    path_frame.grid(column=0, row=3, sticky='we')
    invoice_path_label.grid(column=0, row=4, sticky='w', pady=(20, 0))
    invoice_path_frame.grid(column=0, row=5, sticky='we')
    allinvoices_path_label.grid(column=0, row=6, sticky='w', pady=(20, 0))
    allinvoices_path_frame.grid(column=0, row=7, sticky='we')
    invoice_date_label.grid(column=0, row=8, sticky='w', pady=(20, 0))
    invoice_date_text.grid(column=0, row=9, sticky='we')
    return_list = drag_drop(row1=0, row=1, column=0, dict_name='invoice_finder', text='Aşağıya siteden aldığınız verileri içeren excel dosyasını sürükleyip bırakınız:', parent=bottom_frame, padx=0)
    invoice_finder_drop_frame = return_list[0]
    invoice_finder_surukle_text = return_list[1]
    buttons_frame.grid(column=0, row=2, sticky='e', pady=(20,0))
    baslat_button.pack(side='right', padx=(10,0))
    yonerge_button.pack(side='right', padx=(10,0))
    canvas2.config(scrollregion=canvas2.bbox('all'))
    def invoicefinder_script_starter(path, invoice_folder, date, output_text, allinvoices):
        t = Thread(target=invoicefinder_script, args=(path, invoice_folder, date, output_text, allinvoices), daemon=True)
        t.start()
    def invoicefinderupc_script_starter(path, invoice_folder, upc, month, output_text, allinvoices):
        t = Thread(target=invoicefinderupc_script, args=(path, invoice_folder, upc, month, output_text, allinvoices), daemon=True)
        t.start()
    def output(path, invoice_folder, date, upc, month, allinvoices):
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0), anchor='w')
        if path == placeholder or path == '':
            text_print(output_text, 'Dosyalarin kaydedilecegi dosya yolu algilanamadi lutfen girdiginiz yolu kontrol edip tekrar deneyiniz.')
        elif invoice_folder == placeholder or invoice_folder == '':
            text_print(output_text, 'Invoice Pdf\'lerinin oldugu dosya yolu algilanamadi lutfen girdiginiz yolu kontrol edip tekrar deneyiniz.')
        elif allinvoices == '' or allinvoices == placeholder:
            text_print(output_text, 'ALL INVOICE excel dosyasinin oldugu dosya yolu algilanamadi lutfen girdiginiz yolu kontrol edip tekrar deneyiniz.')
        if upc_switch.status == True:
            if date == '' or date == date_placeholder:
                text_print(output_text, 'Girdiginiz tarih degeri dogru gozukmuyor lutfen kontrol edip tekrar deneyiniz.')
            else:
                invoicefinder_script_starter(path, invoice_folder, date, output_text, allinvoices)
        else:
            if upc == '' or upc == upc_placeholder:
                text_print(output_text, 'Girdiginiz upc degeri dogru gozukmuyor lutfen kontrol edip tekrar deneyiniz.')
            elif month == '' or month == month_placeholder:
                text_print(output_text, 'Girdiginiz Ay degeri dogru gozukmuyor lutfen kontrol edip tekrar deneyiniz.')
            else:
                invoicefinderupc_script_starter(path, invoice_folder, upc, month, output_text, allinvoices)
        window.unbind("<Configure>")
        window.bind('<Configure>', lambda e: invoicefinder_resize(e, 1))
    window.bind('<Configure>', lambda e: invoicefinder_resize(e, 0))
    canvas2.bind_all('<MouseWheel>', on_mouse_wheel)


def order_create_script(path, template, restock_excel, orderform_excel, output_text):
    def settings_writer():
        if 'shipment_settings.txt' not in os.listdir('Settings'):
            with open('Settings/shipment_settings.txt', 'w', encoding='utf-8') as settings:
                text_print(output_text, 'ayarlar dosyası oluşturuluyor...')
                settings.write(ordercreate_settings_var)
                settings.close()
    def settings_reader():
        text_print(output_text, 'ayarlar yükleniyor...')
        sutunlar_dict = {
            'restock_upc': [],
            'restock_pcs': [],
            'restock_suplier': [],
            'restock_notes': [],
            'orderform_upc': [],
            'orderform_pcs': [],
            'orderform_suplier': [],
        }

        with open('Settings/ordercreate_settings.txt', 'r', encoding='utf-8') as settings:
            settings = settings.readlines()
            #RESTOCK SETTINGS
            for line in settings:
                if '=====' in line:
                    break
                line = line.replace('\n', '')
                line = line.split('=')
                if line[0] == 'upc ' or line[0] == 'upc':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_upc'].append(deger)
                elif line[0] == 'pcs ' or line[0] == 'pcs':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_pcs'].append(deger)
                elif line[0] == 'suplier ' or line[0] == 'suplier':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_suplier'].append(deger)
                elif line[0] == 'notes ' or line[0] == 'notes':
                    degerler = line[1].split(',')
                    for deger in degerler:
                        deger = deger.replace(' ', '', 1)
                        sutunlar_dict['restock_notes'].append(deger)


                #ORDER FORM SETTINGS
            a = 0
            for line in settings:
                if '=====' in line:
                    a+=1
                if a == 1:
                    line = line.replace('\n', '')
                    line = line.split('=')
                    if line[0] == 'upc ' or line[0] == 'upc':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_upc'].append(deger)
                    elif line[0] == 'pcs ' or line[0] == 'pcs':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_pcs'].append(deger)
                    elif line[0] == 'suplier ' or line[0] == 'suplier':
                        degerler = line[1].split(',')
                        for deger in degerler:
                            deger = deger.replace(' ', '', 1)
                            sutunlar_dict['orderform_suplier'].append(deger)

        for key in sutunlar_dict.keys():
            text_print(output_text, str(key) + ': ' + str(sutunlar_dict[key]))
        text_print(output_text, 'Ayarlar başarıyla çekİldİ.'.upper())
        return sutunlar_dict
    def dir_creater():
        try:
            os.mkdir(f'{path}/ORDERS')
        except FileExistsError:
            pass

    def restock_reader(restock_excel, sutunlar_dict):
        text_print(output_text, 'Restock excel dosyasi okunuyor...')
        df = pd.read_excel(restock_excel[0])
        df = df.fillna(0)
        try:
            pcs_values = df[sutunlar_dict['restock_pcs'][0]].tolist()
        except:
            text_print(output_text, 'restock dosyasinda PCS sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        try:
            upc_values = df[sutunlar_dict['restock_upc'][0]].tolist()
        except:
            text_print(output_text, 'restock dosyasinda UPC sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        try:
            suplier_values = df[sutunlar_dict['restock_suplier'][0]].tolist()
        except:
            text_print(output_text, 'restock dosyasinda SUPLIER sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        try:
            notes = df[sutunlar_dict['restock_notes'][0]].tolist()
        except:
            text_print(output_text, 'restock dosyasinda NOTES sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        output_dictionary = {}
        for i, a in enumerate(upc_values):
            if pcs_values[i] != 0:
                try:
                    output_dictionary[suplier_values[i]]
                except:
                    output_dictionary[suplier_values[i]] = {}
                try:
                    output_dictionary[suplier_values[i]][a]
                except:
                    output_dictionary[suplier_values[i]][a] = 0
                output_dictionary[suplier_values[i]][a] = output_dictionary[suplier_values[i]][a] + pcs_values[i]
                if notes[i] != 0:
                    try:
                        output_dictionary[notes[i]]
                    except:
                        output_dictionary[notes[i]] = {}
                    try:
                        output_dictionary[notes[i]][a]
                    except:
                        output_dictionary[notes[i]][a] = 0
                    output_dictionary[notes[i]][a] = output_dictionary[notes[i]][a] + pcs_values[i]

        return output_dictionary

    def orderform_reader(orderform_excel, output_dictionary, sutunlar_dict):
        text_print(output_text, 'Order Form excel dosyasi okunuyor...')
        df = pd.read_excel(orderform_excel[0])
        df = df.fillna(0)
        try:
            pcs_values = df[sutunlar_dict['orderform_pcs'][0]].tolist()
        except:
            text_print(output_text, 'orderform dosyasinda PCS sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        try:
            upc_values = df[sutunlar_dict['orderform_upc'][0]].tolist()
        except:
            text_print(output_text, 'orderform dosyasinda UPC sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        try:
            suplier_values = df[sutunlar_dict['orderform_suplier'][0]].tolist()
        except:
            text_print(output_text, 'orderform dosyasinda SUPLIER sutunu bulunamadi lutfen kontrol edip tekrar deneyiniz!', color='red')
            return None
        for i, a in enumerate(upc_values):
            if pcs_values[i] != 0:
                try:
                    output_dictionary[suplier_values[i]]
                except:
                    output_dictionary[suplier_values[i]] = {}
                try:
                    output_dictionary[suplier_values[i]][a]
                except:
                    output_dictionary[suplier_values[i]][a] = 0
                output_dictionary[suplier_values[i]][a] = output_dictionary[suplier_values[i]][a] + pcs_values[i]
        return output_dictionary

    def template_writer(output_dictionary):
        start_row = 2
        text_print(output_text, 'Bulunan değerler yazdırılıyor...')
        for suplier in output_dictionary.keys():
            wb = load_workbook(template)
            ws = wb.active
            for i, upc in enumerate(output_dictionary[suplier].keys()):
                ws.cell(row=start_row + i, column=1, value=upc)
                ws.cell(row=start_row + i, column=3, value=output_dictionary[suplier][upc])
                ws[f'A{start_row+i}'].number_format = '000000000000'
            output_path = f'{path}/ORDERS/{suplier.upper()}.xlsx'
            wb.save(output_path)

    def main():
        try:
            dir_creater()
            settings_writer()
            sutunlar_dict = settings_reader()
            output_dictionary = restock_reader(restock_excel, sutunlar_dict)
            if output_dictionary:
                output_dictionary = orderform_reader(orderform_excel, output_dictionary, sutunlar_dict)
                if output_dictionary:
                    pass
                else:
                    return None
            else:
                return None
            template_writer(output_dictionary)
            open_folder_in_explorer(path)
            text_print(output_text, 'İşlem başarıyla tamamlandı!', color='#90EE90')
        except:
            text_print(output_text, 'İşlem sırasında bir hata meydana geldi!')
            text_print(output_text, traceback.format_exc())
            text_print(output_text, "lütfen kontrol edip tekrar deneyiniz.", color='red')
    main()

def button_order_create(canvas2):

    #RESIZE
    def resize(e, a):
        scale = main_frame_resize()
        restock_dragframe.config(height=175*scale)
        orderform_dragframe.config(height=175*scale)
        height = items_canvas.winfo_height() + items_canvas.winfo_y() + 20
        if a:
            output_text.pack_configure(padx=(canvas.winfo_width(), 0))
            if height < canvas2.winfo_height()-200:
                order_inner_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                order_inner_frame.config(width=750*scale, height=height+200)
        else:
            if height < canvas2.winfo_height():
                order_inner_frame.config(width=750*scale, height=canvas2.winfo_height())
            else:
                order_inner_frame.config(width=750*scale, height=height)
        canvas2.config(scrollregion=canvas2.bbox('all'))


    #MOUSE SCROLL
    def on_mouse_wheel(event):
        canvas2.yview_scroll(int(-1*(event.delta/120)), "units")
    #SCROLLBAR VE FRAME OLUSUMU
    order_inner_frame = Frame(canvas2, width = 0, height=0, bg=color)
    canvas2.create_window((0, 0), anchor='nw', window=order_inner_frame)
    canvas2.config(scrollregion=canvas2.bbox("all"))
    order_scrollbar_y = MyScrollbar(window, target=canvas2, command=canvas2.yview, thumb_thickness=8, thumb_color='#888888', thickness=18, line_color=line_color)
    canvas2.configure(yscrollcommand=order_scrollbar_y.set)
    order_scrollbar_y.pack(side=RIGHT, fill=Y)


    #INNER FRAME GRID SETTINGS
    order_inner_frame.grid_propagate(False)
    order_inner_frame.grid_columnconfigure(0, weight=1)

    #OGELERI GURUPLAMAK ICIN CANVAS OLUSUMU

    title_canvas = Canvas(order_inner_frame, bg=color, highlightthickness=0)
    title_canvas.grid(column=0, row=0, sticky='nwes')
    items_canvas = Canvas(order_inner_frame, bg=color, highlightthickness=0)

    items_canvas.grid(column=0, row=1, sticky='nwes')


    title_canvas.grid_columnconfigure(0, weight=1)
    items_canvas.grid_columnconfigure(0, weight=1)
    #CANVASLAR ICINEKI OGELERIN OLUSUMU

    Order_Title = Label(
        title_canvas,
        text="Order Creater",
        font=("JetBrainsMonoRoman Regular", 24 * -1),
        bg=color,
        fg=canvas2_text_color
    )
    title_line = Frame(title_canvas, height=2, bg=line_color)
    shipment_output = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    shipment_output.bind('<Enter>',lambda e: on_text_enter(e))
    shipment_output.bind('<Leave>',lambda e: on_text_leave(e))
    shipment_output_line = Frame(
        window,
        height=2,
        bg='#787a7e'
    )

    path_label = Label(
        title_canvas,
        text="Aşağıya sonuçların kaydedilmesini istediğiniz dosya yolunu giriniz:",
        background=color,
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 12)
    )


    path_frame = Frame(title_canvas, bg=color, height=30)
    path_text = Text(
        path_frame,
        height=1,
        font=("JetBrainsMonoRoman Regular", 12),
        fg="#747474",
        border=0,
        pady=4,
        bg=line_color,
        insertbackground='#c0c0c0'
    )
    browse_button = MyButton(
        path_frame,
        text='Browse',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )
    save_button = MyButton(
        path_frame,
        text='Kaydet',
        background=line_color,
        text_color='white',
        width=100,
        height=25,
        round=0,
        align_text="center",
        font=("Helvatica", 9)
    )


    restock_return = drag_drop(0, 1, 0, 'order_create_restock', 'RESTOCK excel dosyasini asagiya surukleyip birakiniz.', items_canvas,)
    orderform_return = drag_drop(2, 3, 0, 'order_create_orderform', 'ORDER FORM excel dosyasini asagiya surukleyip birakiniz.', items_canvas,)

    restock_dragframe = restock_return[0]
    orderform_dragframe = orderform_return[0]
    buttons_frame = Frame(
        items_canvas,
        border=0,
        highlightthickness=0,
        bg=color
    )
    baslat_button = MyButton(
        buttons_frame,
        round=12,
        width=100,
        height=40,
        text='Başlat',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    template_location_button = MyButton(
        buttons_frame,
        round=12,
        width=100,
        height=40,
        text='Template',
        background=line_color,
        text_color='white',
        align_text='center'
    )
    output_text = Text(
        window,
        border=0,
        wrap= WORD,
        bg=line_color,
        fg='#c0c0c0',
        height = 10,
        font=("JetBrainsMonoRoman Regular", 13),
        insertbackground='#c0c0c0'
    )
    settings_height = 250
    settings_label = Label(items_canvas, text='Settings:', font=("JetBrainsMonoRoman Regular", 12), background=color, fg=canvas2_text_color)
    if 'ordercreate_settings.txt' not in os.listdir('Settings'):
        settings('Settings/ordercreate_settings.txt', ordercreate_settings_var)
    order_create_settings = Text(
        items_canvas,
        border=0,
        wrap= WORD,
        width=int(width_f(650)),
        bg=line_color,
        fg='#c0c0c0',
        height = int(settings_height/15),
        font=("JetBrainsMonoRoman Regular", 10),
        insertbackground='#c0c0c0'
    )
    order_create_settings.bind('<Enter>', lambda e: on_text_enter(e))
    order_create_settings.bind('<Leave>', lambda e: on_text_leave(e))
    with open('Settings/ordercreate_settings.txt', 'r', encoding='utf-8') as file:
        readed = file.read()
        order_create_settings.insert(tk.END, readed)
        order_create_settings.see(tk.END)
    items_canvas.grid_columnconfigure(0, weight=1)

    baslat_button.pack(side=RIGHT)
    template_location_button.pack(side=RIGHT, padx=(0, 15))
    buttons_frame.grid(column=0, row=6, padx=(0,25), pady=(20,0), sticky='e')
    settings_label.grid(column=0, row=4, columnspan=2, sticky = 'w', padx=25, pady=3)
    order_create_settings.grid(column=0, row=5, sticky='we', padx=25, pady=4,)

    def browse_click(event, c, t, text_item, b):
        browse_color_change(event,c,t,b)
        browse_directory(text_item, w=window)
    def browse_color_change(e,c,t,b):
        b.config(background=c, text_color=t)
    def save_click(event, c, t, b):
        browse_color_change(event,c,t,b)
        placeholder_saver('order_create', path_text)
    def color_changer(event, c, t, b):
        b.config(background=c, text_color=t)
    def template_location_button_click(event, c,t,b):
        b.config(background=c, text_color=t)
        path = os.getcwd()
        open_folder_in_explorer(f'{path}/Settings/Template')
    def baslat_button_click(event, c,t,b):
        b.config(background=c, text_color=t)
        path = path_text.get(1.0, tk.END).strip()
        maindir = os.getcwd()
        template = f'{maindir}/Settings/Template/Template.xlsx'
        restock_excel = dosyalar_dictionary['order_create_restock']
        orderform_excel = dosyalar_dictionary['order_create_orderform']
        output(path, template, restock_excel, orderform_excel)

    browse_button.bind("<Button-1>", lambda e: browse_click(e,'#8AB4F8','black', path_text, browse_button))
    browse_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', browse_button))
    browse_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, browse_button))
    browse_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', browse_button))
    save_button.bind("<Button-1>", lambda e: save_click(e,'#8AB4F8','black', save_button))
    save_button.bind("<ButtonRelease-1>", lambda e: browse_color_change(e,'#727478','white', save_button))
    save_button.bind("<Enter>", lambda e: browse_color_change(e,'#727478',canvas2_text_color, save_button))
    save_button.bind("<Leave>", lambda e: browse_color_change(e,line_color,'white', save_button))
    baslat_button.bind("<Button-1>", lambda e: baslat_button_click(e,'#8AB4F8','black', baslat_button))
    baslat_button.bind("<ButtonRelease-1>", lambda e: color_changer(e,'#727478','white', baslat_button))
    baslat_button.bind("<Enter>", lambda e: color_changer(e,'#727478',canvas2_text_color, baslat_button))
    baslat_button.bind("<Leave>", lambda e: color_changer(e,line_color,'white', baslat_button))
    template_location_button.bind("<Button-1>", lambda e: template_location_button_click(e,'#8AB4F8','black', template_location_button))
    template_location_button.bind("<ButtonRelease-1>", lambda e: color_changer(e,'#727478','white', template_location_button))
    template_location_button.bind("<Enter>", lambda e: color_changer(e,'#727478',canvas2_text_color, template_location_button))
    template_location_button.bind("<Leave>", lambda e: color_changer(e,line_color,'white', template_location_button))

    browse_button.pack(side=RIGHT, fill=Y, padx=(8,0))
    save_button.pack(side=RIGHT, fill=Y, padx=(8,0))
    path_text.pack(side=LEFT, fill=X,expand=True, padx=0, pady=0)
    placeholder = "Example: C:/Users/Username/Desktop/sonuc"
    path_text_function('order_create', path_text, placeholder)
    window.unbind("<Button-1>")
    path_text.bind("<Button-1>", lambda e: on_focus_in(e, path_text, placeholder, canvas2_text_color))
    path_text.bind("<FocusOut>", lambda e: on_focus_out(e, path_text, placeholder, canvas2_text_color))

    canvas2.update_idletasks()
    resize_dictionary[order_inner_frame] = {'width': 750, 'height': order_inner_frame.winfo_height()}

    window.bind("<Button-1>", lambda e: on_click_outside(e, path_text, placeholder, canvas2_text_color))

    Order_Title.grid(column=0, row=0, sticky='w', padx=(25,0), pady=(25,0))
    title_line.grid(column=0,row=1,sticky='we', padx=(20,0))
    path_label.grid(column=0, row=2, pady=(20,0), padx=(25,0), sticky='w')
    path_frame.grid(column=0, row=3,pady=(0,20), padx=(25,5), sticky='we')

    def order_create_script_starter(path, template, restock_excel, orderform_excel, output_text):
        t = Thread(target=order_create_script, args=(path, template, restock_excel, orderform_excel, output_text), daemon=True)
        t.start()


    def output(path, template, restock_excel, orderform_excel):
        output_text.pack(side=BOTTOM, fill=X, padx=(canvas.winfo_width(), 0))
        ordercreate_ayarlar = order_create_settings.get("1.0", tk.END)
        ordercreate_ayarlar = ordercreate_ayarlar.rstrip("\n")
        settings("Settings/ordercreate_settings.txt", ordercreate_ayarlar)
        window.unbind("<Configure>")
        window.bind("<Configure>", lambda e: resize(e, True))
        if path == "Example: C:/Users/Username/Desktop/sonuc":
            output_text.insert(END, "path degeri algilanamadi, lutfen dogru bir deger girdiginizden emin olup tekrar deneyiniz.\n")
            output_text.see(END)
        else:
            order_create_script_starter(path, template, restock_excel, orderform_excel, output_text)

    canvas2.bind_all("<MouseWheel>", on_mouse_wheel)
    window.bind("<Configure>", lambda e: resize(e, False))




def button(canvas2, button):
    canvas2.delete("all")
    canvas2.unbind_all("<MouseWheel>")
    window.unbind('<Configure>')
    silici()
    scrollbar = Scrollbar()
    canvas2.config(scrollregion=canvas2.bbox("all"))  # Canvas'ın scroll bölgesini güncelle
    scrollbar.config(command=canvas2.yview)
    canvas2.config(yscrollcommand=scrollbar.set)
    canvas2.yview_moveto(0)

    image_dictionary = {
        button_1: program_icon_selected,
        button_2: program_icon_selected,
        button_3: program_icon_selected,
        button_4: program_icon_selected,
        button_5: home_icon_selected,
        button_6: program_icon_selected,
        button_7: program_icon_selected,
        button_8: program_icon_selected,
        button_9: program_icon_selected,
        button_10: program_icon_selected,
        button_11: program_icon_selected,
    }
    button_1.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_2.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_3.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_4.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_5.config(background=color, text_color=canvas2_text_color, image=home_icon_notselected)
    button_6.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_7.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_8.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_9.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_10.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button_11.config(background=color, text_color=canvas2_text_color, image=program_icon_notselected)
    button.config(background='#8AB4F8', text_color='black', image=image_dictionary[button])
    def dictionary_update(button):
        dictionary[button_1] = 0
        dictionary[button_2] = 0
        dictionary[button_3] = 0
        dictionary[button_4] = 0
        dictionary[button_5] = 0
        dictionary[button_6] = 0
        dictionary[button_7] = 0
        dictionary[button_8] = 0
        dictionary[button_9] = 0
        dictionary[button_10] = 0
        dictionary[button_11] = 0
        dictionary[button] = 1
    if button == button_1:
        dictionary_update(button_1)
        canvas2.unbind_all('<MouseWheel>')
        button_expration(canvas2)
    if button == button_2:
        window.unbind("<Configure>")
        dictionary_update(button_2)
        shipmentCreater(canvas2)
    if button == button_3:
        dictionary_update(button_3)
        canvas2.unbind_all('<MouseWheel>')
        button_tsv(canvas2)
    if button == button_4:
        dictionary_update(button_4)
        canvas2.config(height=window.winfo_height())
        restock(canvas2)
    if button == button_5:
        dictionary_update(button_5)
        canvas2.unbind_all('<MouseWheel>')

        anasayfa_canvas = Canvas(
            canvas2,
            background=color,
            highlightthickness=0,
            border=0
        )
        anasayfa_canvas.pack(anchor='center', expand=True, side=LEFT)

        line = Frame(
            anasayfa_canvas,
            height=4,
            background=line_color
        )

        hello = Label(
            anasayfa_canvas,
            background=color,
            fg=canvas2_text_color,
            text="KWIEK LLC TOPLU İŞLEM PLATFORMUNA HOŞGELDİNİZ!",
            font=("JetBrainsMonoRoman Regular", 24 * -1)
        )

        islem = Label(
            anasayfa_canvas,
            background=color,
            text="Bir işlem yapmak için lütfen sol menüdeki işlemlerden birini seçiniz...",
            fg=canvas2_text_color,
            font=("JetBrainsMonoRoman Regular", 15 * -1)
        )

        hello.grid(column=0, row=0, sticky='ew', padx=40)
        line.grid(column=0, row=1, sticky='ew', pady=15)
        islem.grid(column=0, row=2, sticky='ew')
        liste= [canvas, canvas2, button_1, button_2, button_3, button_4, button_5]
        window.bind("<Configure>", lambda e: main_resize(e,liste, hello, islem))
    if button == button_6:

        dictionary_update(button_6)
        canvas2.unbind_all('<MouseWheel>')
        button_invoice(canvas2)
    if button == button_7:
        dictionary_update(button_7)
        button_converter(canvas2)
    if button == button_8:
        dictionary_update(button_8)
        button_costupdater(canvas2)
    if button == button_9:
        dictionary_update(button_9)
        button_updater(canvas2)
    if button == button_10:
        dictionary_update(button_10)
        button_invoicefinder(canvas2)
    if button == button_11:
        dictionary_update(button_11)
        button_order_create(canvas2)


def drag_drop(row1,row,column,dict_name,text,parent,win=0, bg_image=0, file_image=0, file_type=".xlsx", padx=25, pady=25, ):
    if win == 0:
        win = window
    dosyalar_dictionary[dict_name] = []
    button_list = []
    image_dictionary = {
        'sil_button_image': PhotoImage(file=relative_to_assets("image_3.png"), width=35, height=25),
        'excel_dosya_image': PhotoImage(file=relative_to_assets("image_5.png")),
        'drag_drop_image': ''
    }
    if file_image == 0:
        file_image = image_dictionary['excel_dosya_image']
    surukle_text = Label(parent, text=text, background=color, fg=canvas2_text_color, font=("JetBrainsMonoRoman Regular", 12))
    surukle_text.grid(column=column, row=row1, columnspan=2, padx=padx, pady=10, sticky='w')
    def on_frame_enter(event):
        canvas2.unbind_all("<MouseWheel>")
        drop_canvas.bind_all("<MouseWheel>", on_mouse_wheel_frame)

    def on_frame_leave(event):
        canvas2.bind_all("<MouseWheel>", on_mouse_wheel)
    def drag():
        main_canvas.yview_scroll(10, "units")
    def on_mouse_wheel_frame(event):
        drop_canvas.yview_scroll(int(-1*(event.delta/120)), "units")



    def drop(event):
        main_frame.bind("<Enter>", on_frame_enter)
        main_frame.bind("<Leave>", on_frame_leave)
        drop_canvas.bind_all("<MouseWheel>", on_mouse_wheel_frame)
        drop_canvas.config(scrollregion=drop_canvas.bbox('all'))
        k = 1
        if type(event) == tuple or event == '':
            file_path = list(event)
        else:
            file_path = event.data
            if file_type not in file_path:
                k = 0
                inner_frame.config(height=main_canvas.winfo_height(), width=main_canvas.winfo_width(), bg='#616161')
                inner_frame.pack_propagate(False)
                label = Label(
                    inner_frame,
                    bg='#616161',
                    text="Yanlış dosya tipi algılandı!",
                    fg='white'
                )
                label.pack(side=LEFT,fill=BOTH, expand=True)
                def destroy():
                    label.destroy()
                    inner_frame.config(height=0, bg='#D9D9D9')
                    inner_frame.pack_propagate(True)
                win.after(1000, destroy)
        z = 0
        if type(event) == tuple or event == '':
            file_path = list(event)
            z=1
        else:
            file_path = event.data.strip().split(file_type)
            z=0
        for i in file_path:
            if '{' or '}' in i:
                i = i.replace('{', "")
                i = i.replace('}', "")
            if i != '':
                if i[0] == ' ':
                    i = i[1:]
                if z == 0:
                    i = i + file_type
                if i not in dosyalar_dictionary[dict_name] and k == 1:
                    inner_frame.grid_propagate(True)
                    dosyalar_dictionary[dict_name].append(i)
                    buttons_frame = Frame(inner_frame, highlightbackground='black', highlightthickness=1, height=30, padx=0, pady=0, width= drop_canvas.winfo_width())
                    excel_image = Label(buttons_frame, image=file_image)
                    excel_image.pack(side=LEFT)
                    button = Button(buttons_frame, text=i, font=("JetBrainsMonoRoman Regular", 15 * -1), height=1, border= 0, anchor='w')
                    inner_frame.columnconfigure(0, weight=1, minsize=drop_canvas.winfo_width())
                    buttons_frame.grid(column=0, row=dosyalar_dictionary[dict_name].index(i), columnspan=2, padx=0, pady=0, sticky='nswe')
                    button.pack(side=RIGHT, fill=BOTH, expand=True)
                    parent.update_idletasks()
                    button_width = button.winfo_width()
                    inner_frame.grid_propagate(False)
                    buttons_frame.config(height=30)
                    buttons_frame.pack_propagate(False)


                    button_tik = Button(buttons_frame,image=image_dictionary['sil_button_image'], border=0, width=35, command=lambda b= button, db = None, bl = None: delete_button_func(b, db, bl))
                    button_tik.config(command=lambda bf = buttons_frame, b=button, bl=button_list: delete_button_func(bf, b, bl))

                    def ustunde(event, buttons_frame, button, button_tik):

                        button_tik.place(x=0,y=1)
                        buttons_frame.update()
                        index = dosyalar_dictionary[dict_name].index(button.cget('text'))
                        buttons_frame.update_idletasks()
                        buttons_frame.bind('<Leave>', lambda e: degil(e, button_tik))
                    def degil(event, button_tik):
                        try:
                            button_tik.place_forget()
                        except:pass
                        buttons_frame.update_idletasks()
                    def delete_button_func(buttons_frame, button, button_list):
                        for button1 in button_list:
                            if button1[0] == button:
                                button_list.remove(button1)
                        i = button.cget('text')
                        dosyalar_dictionary[dict_name].remove(i)
                        button.destroy()
                        buttons_frame.destroy()

                        if len(dosyalar_dictionary[dict_name]) == 0:
                            inner_frame.config(height=0)
                        else:
                            update_buttons()
                        # Tuşlar silindiği için frame güncellenmeli
                    def update_buttons():
                        for button in button_list:
                            i = button[0].cget('text')
                            button[2].grid_configure(row=dosyalar_dictionary[dict_name].index(i))
                            #parent.update_idletasks()
                            button_width = button[0].winfo_width()
                            '''
                            if len(dosyalar_dictionary[dict_name]) == 0:
                                inner_frame.config(height=0)
                            elif 20+30*len(dosyalar_dictionary[dict_name]) > drop_canvas.winfo_height():
                                inner_frame.config(height=20+30*len(dosyalar_dictionary[dict_name]))
                            else:
                                inner_frame.config(height=drop_canvas.winfo_height())
                            button[2].config(width= inner_frame.winfo_width())
                            '''
                            #parent.update_idletasks()
                    def update_size(button_list, inner_frame):
                        inner_frame.update()
                        for i in button_list:
                            i[2].config(width= inner_frame.winfo_width())
                            parent.update_idletasks()

                    button_list.append([button, 'h', buttons_frame])

                    #button_tik.place(x=button_width + 15, y=2+30*(dosyalar_dictionary[dict_name].index(i)))
                    #button_tik.place(x=button_width + 15, y=0)
                    buttons_frame.bind('<Enter>', lambda e, bf = buttons_frame, b = button, bt=button_tik : ustunde(e, bf, b, bt))
                    parent.update_idletasks()
                    inframe_width = drop_canvas.winfo_width()
                    buttons_frame.configure(width=button.winfo_width()+30)
                    parent.update_idletasks()
                    toplam = buttons_frame.winfo_width()
                    if (10+(30*(len(dosyalar_dictionary[dict_name])))) > drop_canvas.winfo_height():
                        inner_frame.config(height=10+(30*(len(dosyalar_dictionary[dict_name]))))
                    else:
                        inner_frame.config(height=drop_canvas.winfo_height()-5)
                    if toplam > inframe_width:
                        scrollbar_h.grid(column=0, row=1, sticky='ew')
                        inner_frame.config(width=toplam)

                    #update_size(button_list, inner_frame)
                    #parent.update_idletasks()
    main_frame = Frame(parent,background='#3F4042', borderwidth=0, relief="solid", highlightcolor='#3F4042', highlightthickness=6, highlightbackground='#3F4042')
    main_frame.grid(column = column, row = row, columnspan=2, sticky='nwes', padx=padx)
    main_canvas= Canvas(main_frame, bg='white')
    main_canvas.pack(side=LEFT, fill=BOTH, expand=True)
    drop_canvas = Canvas(main_canvas, bg='white', height= 150)

    if bg_image == 0:
        image_dictionary['drag_drop_image'] = PhotoImage(file=relative_to_assets("image_6.png"))
        bg_image = image_dictionary['drag_drop_image']
    else:pass
    main_canvas.grid_columnconfigure(0, weight=1)
    main_canvas.grid_rowconfigure(0, weight=1)
    drop_canvas.grid(column = 0, row = 0, sticky='nsew')
    drop_canvas.pack_propagate(False)
    drag_drop_label = Label(drop_canvas, bg='#D9D9D9')
    drag_drop_label.background_image = bg_image
    drag_drop_label.config(image=bg_image)
    drag_drop_label.pack(side=LEFT, fill=BOTH, expand=True)


    scrollbar_v = Scrollbar(main_canvas, orient=VERTICAL, command=drop_canvas.yview)
    scrollbar_v.grid(column = 1, row = 0, sticky='ns')
    drop_canvas.config(yscrollcommand=scrollbar_v.set)

    scrollbar_h = Scrollbar(main_canvas, orient=HORIZONTAL, command=drop_canvas.xview)

    drop_canvas.config(xscrollcommand=scrollbar_h.set)

    inner_frame = Frame(drop_canvas, bg='#D9D9D9', height=0, width=main_canvas.winfo_width())
    drop_canvas.create_window((0, 0), window=inner_frame, anchor="nw")
    def config(e):
        inner_frame.config(width=e.width)
    inner_frame.bind("<Configure>", lambda e: drop_canvas.config(scrollregion=drop_canvas.bbox('all')))
    main_canvas.bind("<Configure>", lambda e: config(e))
    a=0
    # Sürükle bırak işlemi için hedef belirleme
    def drop_canvas_click(event):
        if file_type == ".xlsx":
            file_path = filedialog.askopenfilename(
                parent=win,
                title="Bir Excel dosyası seçin",
                filetypes=[("Excel Files", "*.xlsx *.xls")],  # Sadece Excel dosyalarını filtreler
                multiple=True
            )
        else:
            file_path = filedialog.askopenfilename(
                parent=win,
                title="Bir Excel dosyası seçin",
                filetypes=[("Excel Files", "*{}".format(file_type))],  # Sadece Excel dosyalarını filtreler
                multiple=True
            )
        drop(file_path)
    drag_drop_label.bind('<Button-1>', drop_canvas_click)
    drop_canvas.drop_target_register(DND_FILES)

    drop_canvas.dnd_bind('<<Drop>>', lambda e: drop(e))
    return [main_frame, surukle_text]


def ham_drag_drop2(row1,row,column,dict_name,text,parent):
    dosyalar_dictionary[dict_name] = []
    button_list = []
    sil_button_image = PhotoImage(
        file=relative_to_assets("image_3.png"),
        width=35,
        height=25
    )
    excel_dosya_image = PhotoImage(
        file=relative_to_assets("image_5.png")
    )
    surukle_text = Label(parent, text=text, background=color, fg=canvas2_text_color,)
    surukle_text.grid(column=column, row=row1, columnspan=2, padx=25, pady=10, sticky='w')
    def on_frame_enter(event):
        canvas2.unbind_all("<MouseWheel>")
        drop_canvas.bind_all("<MouseWheel>", on_mouse_wheel_frame)

    def on_frame_leave(event):
        canvas2.bind_all("<MouseWheel>", on_mouse_wheel)
    def drag():
        main_canvas.yview_scroll(10, "units")
    def on_mouse_wheel_frame(event):
        drop_canvas.yview_scroll(int(-1*(event.delta/120)), "units")



    def drop(event):
        main_frame.bind("<Enter>", on_frame_enter)
        main_frame.bind("<Leave>", on_frame_leave)
        drop_canvas.bind_all("<MouseWheel>", on_mouse_wheel_frame)
        drop_canvas.config(scrollregion=drop_canvas.bbox('all'))

        k = 1
        if type(event) == tuple or event == '':
            file_path = event
        else:
            file_path = event.data
            if '.xlsx' not in file_path:
                k = 0
                inner_frame.config(height=main_canvas.winfo_height(), width=main_canvas.winfo_width(), bg='#616161')
                inner_frame.pack_propagate(False)
                label = Label(
                    inner_frame,
                    bg='#616161',
                    text="Yanlış dosya tipi algılandı!",
                    fg='white'
                )
                label.pack(side=LEFT,fill=BOTH, expand=True)
                def destroy():
                    label.destroy()
                    inner_frame.config(height=0, bg='#D9D9D9')
                    inner_frame.pack_propagate(True)
                window.after(1000, destroy)
        z = 0
        if type(event) == tuple or event == '':
            file_path = event
            z = 1
        else:
            file_path = event.data.strip().split('.xlsx')
            z = 0
        for i in file_path:
            if '{' or '}' in i:
                i = i.replace('{', "")
                i = i.replace('}', "")
            if i != '':
                if i[0] == ' ':
                    i = i[1:]
                if z == 0:
                    i = i + ".xlsx"
                if i not in dosyalar_dictionary[dict_name] and k == 1:
                    def button_click(event):
                        var.set('1')
                        vary.set(event.y)
                    def button_release(event):
                        var.set('0')
                        to = int(surukle_line.grid_info()['row']/2)
                        index = dosyalar_dictionary[dict_name].index(event.widget.cget('text'))
                        print(index)
                        tasi(dosyalar_dictionary[dict_name], index, to)
                        surukle_line.grid_forget()


                    def button_motion(event):
                        if var.get() == '1':
                            y = event.y
                            if y >= 0:
                                which_file = int((y+vary.get())/30)
                            else:
                                which_file = int((y-vary.get())/30)
                            #print(which_file)
                            which_row = (2*which_file)

                            row = event.widget.master.grid_info()['row']+which_row - 1
                            if row < 0:
                                row =0
                            surukle_line.grid(column=0, row=row, sticky='ew')
                            '''elif y > 15 and y <=30:
                                row = event.widget.master.grid_info()['row']+which_row + 1
                                print(row)
                                surukle_line.grid(column=0, row=row, sticky='ew')'''
                    inner_frame.grid_propagate(True)
                    dosyalar_dictionary[dict_name].append(i)
                    buttons_frame = Frame(inner_frame, highlightbackground='black', highlightthickness=1, height=30, padx=0, pady=0, width= drop_canvas.winfo_width())
                    excel_image = Label(buttons_frame, image=excel_dosya_image)
                    excel_image.pack(side=LEFT)
                    button = Label(buttons_frame, text=i, font=("JetBrainsMonoRoman Regular", 15 * -1), height=1, border= 0, anchor='w')
                    inner_frame.columnconfigure(0, weight=1, minsize=drop_canvas.winfo_width())
                    buttons_frame.grid(column=0, row=2*dosyalar_dictionary[dict_name].index(i)+1, columnspan=2, padx=0, pady=0, sticky='nswe')
                    button.pack(side=RIGHT, fill=BOTH, expand=True)
                    button.bind("<Button-1>", button_click)
                    button.bind("<ButtonRelease-1>", button_release)
                    button.bind("<B1-Motion>", lambda e: button_motion(e))

                    #button.bind("<<ButtonRelease-1>>", button_release())

                    parent.update_idletasks()
                    button_width = button.winfo_width()
                    inner_frame.grid_propagate(False)
                    buttons_frame.config(height=30)
                    buttons_frame.pack_propagate(False)

                    var = tk.StringVar()
                    var.set('0')
                    vary = tk.IntVar()

                    def tasi(lst, from_index, to_index):
                        if from_index < 0 or from_index >= len(lst):
                            raise IndexError("from_index is out of bounds")
                        if to_index < 0 or to_index >= len(lst):
                            raise IndexError("to_index is out of bounds")

                        # Öğeyi çıkart ve yeni konuma ekle
                        item = lst.pop(from_index)
                        lst.insert(to_index, item)

                        update_buttons()
                    button_tik = Button(buttons_frame,image=sil_button_image, border=0, width=35, command=lambda b= button, db = None, bl = None: delete_button_func(b, db, bl))
                    button_tik.config(command=lambda bf = buttons_frame, b=button, bl=button_list: delete_button_func(bf, b, bl))
                    yukari_button = Button(
                        buttons_frame,
                        text='yukari',
                    )
                    asagi_button = Button(
                        buttons_frame,
                        text='asagi',
                    )
                    def ustunde(event, buttons_frame, button, button_tik, yukari_button, asagi_button):

                        button_tik.place(x=0,y=1)
                        buttons_frame.update()
                        index = dosyalar_dictionary[dict_name].index(button.cget('text'))
                        yukari_button.config(command= lambda: tasi(dosyalar_dictionary[dict_name], index, index-1))
                        asagi_button.config(command= lambda: tasi(dosyalar_dictionary[dict_name], index, index+1))
                        yukari_button.place(
                            x=button_tik.winfo_width()+5,
                            y = 0
                        )

                        asagi_button.place(
                            x=button_tik.winfo_width()+55,
                            y = 0
                        )

                        buttons_frame.update_idletasks()
                        buttons_frame.bind('<Leave>', lambda e: degil(e, button_tik, asagi_button, yukari_button))
                    def degil(event, button_tik, asagi_button, yukari_button):
                        try:
                            button_tik.place_forget()
                        except:
                            pass
                        try:
                            yukari_button.place_forget()
                        except:
                            pass
                        try:
                            asagi_button.place_forget()
                        except:
                            pass
                        buttons_frame.update_idletasks()
                    def delete_button_func(buttons_frame, button, button_list):
                        for button1 in button_list:
                            if button1[0] == button:
                                button_list.remove(button1)
                        i = button.cget('text')
                        dosyalar_dictionary[dict_name].remove(i)
                        button.destroy()
                        buttons_frame.destroy()

                        if len(dosyalar_dictionary[dict_name]) == 0:
                            inner_frame.config(height=0)
                        else:
                            update_buttons()
                        # Tuşlar silindiği için frame güncellenmeli
                    def update_buttons():
                        for button in button_list:
                            i = button[0].cget('text')
                            button[2].grid_configure(row=2*(dosyalar_dictionary[dict_name].index(i))+1)
                            #parent.update_idletasks()
                            button_width = button[0].winfo_width()
                            '''
                            if len(dosyalar_dictionary[dict_name]) == 0:
                                inner_frame.config(height=0)
                            elif 20+30*len(dosyalar_dictionary[dict_name]) > drop_canvas.winfo_height():
                                inner_frame.config(height=20+30*len(dosyalar_dictionary[dict_name]))
                            else:
                                inner_frame.config(height=drop_canvas.winfo_height())
                            button[2].config(width= inner_frame.winfo_width())
                            '''
                            #parent.update_idletasks()
                    def update_size(button_list, inner_frame):
                        inner_frame.update()
                        for i in button_list:
                            i[2].config(width= inner_frame.winfo_width())
                            parent.update_idletasks()

                    button_list.append([button, 'h', buttons_frame])

                    #button_tik.place(x=button_width + 15, y=2+30*(dosyalar_dictionary[dict_name].index(i)))
                    #button_tik.place(x=button_width + 15, y=0)
                    buttons_frame.bind('<Enter>', lambda e, bf = buttons_frame, b = button, bt=button_tik, y = yukari_button, a = asagi_button: ustunde(e, bf, b, bt, y, a))
                    parent.update_idletasks()
                    inframe_width = drop_canvas.winfo_width()
                    buttons_frame.configure(width=button.winfo_width()+30)
                    parent.update_idletasks()
                    toplam = buttons_frame.winfo_width()
                    if (10+(30*(len(dosyalar_dictionary[dict_name])))) > drop_canvas.winfo_height():
                        inner_frame.config(height=10+(30*(len(dosyalar_dictionary[dict_name]))))
                    else:
                        inner_frame.config(height=drop_canvas.winfo_height()-5)
                    if toplam > inframe_width:
                        scrollbar_h.grid(column=0, row=1, sticky='ew')
                        inner_frame.config(width=toplam)


                    #update_size(button_list, inner_frame)
                    #parent.update_idletasks()
    main_frame = Frame(parent, background='#3F4042', border=0, relief="solid", highlightcolor='#3F4042', highlightthickness=6, highlightbackground='#3F4042')
    main_frame.grid(column=column, row=row, columnspan=2, sticky='nwes', padx=25)
    main_canvas = Canvas(main_frame, bg='white')
    main_canvas.pack(side=LEFT, fill=BOTH, expand=True)
    drop_canvas = Canvas(main_canvas, bg='white', height= 150)
    drag_drop_image = PhotoImage(
        file=relative_to_assets("image_6.png")
    )
    main_canvas.grid_columnconfigure(0, weight=1)
    main_canvas.grid_rowconfigure(0, weight=1)
    drop_canvas.grid(column = 0, row = 0, sticky='nsew')
    drop_canvas.pack_propagate(False)
    drag_drop_label = Label(drop_canvas, bg='#D9D9D9')
    drag_drop_label.background_image= drag_drop_image
    drag_drop_label.config(image=drag_drop_image)
    drag_drop_label.pack(side=LEFT, fill=BOTH, expand=True)


    scrollbar_v = Scrollbar(main_canvas, orient=VERTICAL, command=drop_canvas.yview)
    scrollbar_v.grid(column = 1, row = 0, sticky='ns')
    drop_canvas.config(yscrollcommand=scrollbar_v.set)

    scrollbar_h = Scrollbar(main_canvas, orient=HORIZONTAL, command=drop_canvas.xview)

    drop_canvas.config(xscrollcommand=scrollbar_h.set)

    inner_frame = Frame(drop_canvas, bg='#D9D9D9', height=0, width=main_canvas.winfo_width(), border=0, highlightthickness=0)
    surukle_line = Frame(
        inner_frame,
        bg='black',
        height=5
    )
    drop_canvas.create_window((0, 0), window=inner_frame, anchor="nw")
    def config(e):
        inner_frame.config(width=e.width)
    inner_frame.bind("<Configure>", lambda e: drop_canvas.config(scrollregion=drop_canvas.bbox('all')))
    main_canvas.bind("<Configure>", lambda e: config(e))
    a=0
    def drop_canvas_click(event):
        file_path = filedialog.askopenfilename(
            parent=window,
            title="Bir Excel dosyası seçin",
            filetypes=[("Excel Files", "*.xlsx *.xls")],  # Sadece Excel dosyalarını filtreler
            multiple=True
        )
        drop(file_path)
    drag_drop_label.bind('<Button-1>', drop_canvas_click)
    # Sürükle bırak işlemi için hedef belirleme
    drop_canvas.drop_target_register(DND_FILES)

    drop_canvas.dnd_bind('<<Drop>>', lambda e: drop(e))

    return [main_frame, surukle_text]

def is_connected_whenstart():
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=5)
        api_url = "https://api.github.com/repos/hasali2603/KWIEKLLC/releases/latest"
        response = requests.get(api_url)
        release_data = response.json()
        latest_version = release_data['tag_name']
        if latest_version > CURRENT_VERSION:
            version.config(text=f"{CURRENT_VERSION} yeni version({latest_version}) mevcut!", fg='yellow')
            tk.messagebox.showinfo("Version", f"New version available: {latest_version}!")
    except:
        pass
def is_connected(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, label, yuklebutton, releasebutton):

    try:
        socket.create_connection(("8.8.8.8", 53), timeout=5)
        application_updater(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, label, yuklebutton, releasebutton)
    except OSError:
        messagebox.showwarning("Uyarı", "İnternet bağlantınızı kontrol edip daha sonra tekrar deneyin.")

def download_file_with_progress(url, save_path, progress_bar, progress_label):
    response = requests.head(url, allow_redirects=True)
    total_size = int(response.headers.get('content-length', 0))
    progress_bar.pack(side=BOTTOM, fill=X)
    progress_label.pack(side=BOTTOM, anchor='w')
    progress_bar['maximum'] = total_size
    response = requests.get(url, stream=True)
    with open(save_path, 'wb') as file:
        downloaded_size = 0
        for chunk in response.iter_content(1024):
            if chunk:
                file.write(chunk)
                downloaded_size += len(chunk)
                progress_bar['value'] = downloaded_size
                window.update_idletasks()

def start_download(url, progress_bar, progress_label):
    temp_dir = tempfile.gettempdir()
    update_save_path = os.path.join(temp_dir, "KWIEKLLC_update.exe")
    download_file_with_progress(url, update_save_path, progress_bar, progress_label)


def application_updater(CURRENT_VERSION, progress_bar, progress_label, doyouwanna_frame, label:tk.Label, yuklebutton, releasebutton):
    def baslat_click(e,c,t, i):
        i.config(background=c, text_color=t)
        update_starter()
    def release_click(e,c,t,i,text, vs):
        i.config(background=c, text_color=t)
        release_window = Tk()
        release_window.geometry("500x300")
        release_window.title(f'Release Notes of {vs}')
        releasenotes_text = Text(
            release_window,
            border=0,
            wrap= WORD,
            bg=line_color,
            fg='#c0c0c0',
            #height = int(settings_height/15),
            font=("JetBrainsMonoRoman Regular", 10),
            insertbackground='#c0c0c0'
        )
        releasenotes_text.insert(tk.END, text)
        releasenotes_text.see(tk.END)
        releasenotes_text.config(state=tk.DISABLED)
        releasenotes_text.pack(side=LEFT, fill=BOTH, expand=True, anchor='nw')
    def update_starter():
        t = Thread(target=update, daemon=True)
        t.start()
    def update():
        asset_url = release_data['assets'][1]['browser_download_url']
        start_download(asset_url, progress_bar, progress_label)
        temp_dir = tempfile.gettempdir()
        batch_file_path = os.path.join(temp_dir, "run_update.bat")
        with open(batch_file_path, "w") as batch_file:
            batch_file.write(
                f'@echo off\n'
                f'set "update_path=%~dp0KWIEKLLC_update.exe"\n'
                f'start "" /b "%update_path%"\n'
                f':wait_loop\n'
                f'tasklist /FI "IMAGENAME eq KWIEKLLC_update.exe" 2>NUL | find /I /N "KWIEKLLC_update.exe">NUL\n'
                f'if "%ERRORLEVEL%"=="0" (\n'
                f'    timeout /t 5 > NUL\n'
                f'    goto wait_loop\n'
                f')\n'
                f'echo Update process finished >> "%~dp0update_log.txt"\n'
                f'echo Update process finished >> "%~dp0update_log.txt"\n'
                f'del /f /q "%~dp0KWIEKLLC_update.exe"\n'
                f'del /f /q "%~f0" & exit\n'
            )

        # Run the batch file
        subprocess.Popen([batch_file_path], creationflags=subprocess.CREATE_NO_WINDOW)
        exit()
    def force_exit():
        subprocess.run(["taskkill", "/F", "/IM", "KWIEK LLC.exe"])
        sys.exit()
    def exit():
        os._exit(0)
        window.quit()
        try:
            force_exit()
        except:
            pass
        pid = os.getpid()  # Geçerli işlem kimliğini alır
        os.kill(pid, signal.SIGTERM)  # İşlemi zorla sonlandırır
        sys.exit()



    api_url = "https://api.github.com/repos/hasali2603/KWIEKLLC/releases/latest"

    response = requests.get(api_url) #headers=headers)
    release_data = response.json()
    latest_version = release_data['tag_name']
    if latest_version > CURRENT_VERSION:
        release_notes = release_data['body']
        version.config(text=f"{CURRENT_VERSION} yeni version({latest_version}) mevcut!", fg='yellow')

        label.config(text=f'Yeni bir versiyon bulundu({latest_version})! Yüklemek istiyor musun?')

        yuklebutton.bind("<Button-1>", lambda e: baslat_click(e,'#8AB4F8','black', yuklebutton))
        releasebutton.bind("<Button-1>", lambda e: release_click(e,'#8AB4F8','black', releasebutton, release_notes, latest_version))

        doyouwanna_frame.grid(column=0, row=2, sticky='w')
    else:
        items = doyouwanna_frame.winfo_children()
        for item in items:
            if type(item) == Label:
                item.config(text='Uygulama güncel!')
            else:
                item.grid_forget()
        doyouwanna_frame.grid(column=0, row=2, sticky='w')


if __name__ == '__main__':
    freeze_support()

    CURRENT_VERSION = "v1.2.2"

    OUTPUT_PATH = Path(__file__).resolve().parent
    ASSETS_PATH = OUTPUT_PATH / "assets" / "frame0"


    dosyalar_dictionary = {}

    color = '#202124'
    window = TkinterDnD.Tk()

    #bg='#ADD8E6'

    canvas2_text_color = '#E3E3E3'
    line_color='#3F4042'
    active_dictionary= {
        'restock': 1,
        'export': 1
    }

    #865x519
    wr=float(865/1920)
    hr=float(519/1080)
    m = get_monitors()[0]

    scale = 1
    window.configure(bg = color)
    screenwidth = int(scale*wr * 2560)
    screen_height = int(scale*hr * 1600)
    #print(screenwidth,screen_height)
    window.geometry("{}x{}".format(screenwidth, screen_height))
    window.title('KWIEK LLC')
    try:
        dark_title_bar(window)
    except:pass
    try:
        window.iconbitmap('assets/icon.ico')
    except:pass

    original_selected_image = HASAN.open(relative_to_assets('selected.png'))
    original_notselected_image = HASAN.open(relative_to_assets('not_selected.png'))
    selected_resized_image = original_selected_image.resize((15, 15))
    notselected_resized_image = original_notselected_image.resize((15, 15))
    selected_image = ImageTk.PhotoImage(selected_resized_image)
    not_selected_image = ImageTk.PhotoImage(notselected_resized_image)
    csv_drag_drop_image = PhotoImage(file=relative_to_assets('csv_drag_drop_rs.png'))
    csv_icon_image = PhotoImage(file=relative_to_assets('csv_icon_rs.png'))
    txt_drag_drop_image = PhotoImage(file=relative_to_assets('txt_drag_drop_rs.png'))
    txt_icon_image = PhotoImage(file=relative_to_assets('txt_icon_rs.png'))
    settings_var = (
        "upc = UPC, upc, Upc, UPC #\n"
        "brand = BRAND, Brand, brand\n"
        "price = NET_AMOUNT, Price, price\n"
        "case = CASEPACK, Size, Case, case, size\n"
        "Quantity on hand = Qty on Hand, Quantity Available\n"
        "pk = PK\n"
        "======================================\n"
        "41 cost: 0.78\n"
        "41 standart: 0.78\n"
        "45 cost: 0.78\n"
        "45 standart: 0.78\n"
        "19 cost: 0.78\n"
        "19 standart: 0.78\n"
        "27 cost: 1.10\n"
        "27 standart: 1.10\n"
        "18 cost: 1.10\n"
        "18 standart: 1.10\n"
        "01 cost: 1.10\n"
        "01 standart: 1.10\n"
        "NF: 0.78")
    shipment_settings_var = (
        'RESTOCK:\n'
        'upc = Upc\n'
        'pcs = PCS\n'
        'asin = ASIN\n'
        'pk = PK\n'
        'price = Price\n'
        'suplier = suplier\n'
        '=====================================================\n'
        'ORDER FORM:\n'
        'upc = UPC\n'
        'pcs = PCS\n'
        'asin = ASIN 1, ASIN 2, ASIN 3, ASIN 4\n'
        'SKU = ASIN1_SKU, ASIN2_SKU, ASIN3_SKU, ASIN4_SKU\n'
        'pk = PK\n'
        'price = price\n'
        'suplier = suplier\n'
        '=====================================================\n'
        'INVOICE:\n'
        'shipquantity = ShipQuantity\n'
        'upc = Upc\n'
        'price = NetEach2\n'
        'packsize = PackSize\n'
        'brand = Brand\n'
        'description = Description\n')
    tsv_settings_var = (
        "columns = Merchant SKU, Title, ASIN, FNSKU, external-id, Condition, Shipped"
    )
    expration_settings_var = (
        'login_button_id = mainForm:j_idt23, mainForm:j_idt13\n'
        'email_id = mainForm:email\n'
        'password_id = mainForm:password\n'
        'default_email = sales@buyable.net\n'
        'default_password = hasali2603\n'
    )
    invoice_settings_var = (
        'remove = Status, QuantityNotShipped, InvalidReason\n'
        'shipquantity = ShipQuantity\n'
        'date = InvoiceDate'
    )
    costupdater2_settings_var = (
        'cost = cost\n'
        'sku = sku\n'
        'additional cost = additional_cost\n'
        'business pricing = business_pricing\n'
        'bp strategy = bp_strategy\n'
        'qd strategy = qd_strategy\n'
        'pkg volume = pkg_volume\n'
        'pkg weight = pkg_weight\n'
        '====================================\n'
        'DC_NAME: ADDITIONAL_COST EQUATION_NUMBER DEPOSIT_COST\n'
        'BX: 0 2 0.70\n'
        'CANDY: 0 2 0.70\n'
        'COS: 0 2 0.70\n'
        'CS: 0 2 0.70\n'
        'CSC: 0 2 0.70\n'
        'DL: 0 2 0.70\n'
        'FC: 0 2 0.70\n'
        'FD: 0 2 0.70\n'
        'FL: 0 1 0.70\n'
        'FOUR: 0 2 0.70\n'
        'FR: 0 2 0.70\n'
        'GEMCO: 0 2 0.70\n'
        'IL: 0 1 0.70\n'
        'JC: 0 2 0.70\n'
        'KH: 0 2 0.70\n'
        'LR: 0 2 0.70\n'
        'MD: 0 1 0.70\n'
        'MONIN PUMP SL: 0 2 0.70\n'
        'NC: 0 2 0.70\n'
        'NF: 0 2 0.70\n'
        'NJ: 0 2 0.70\n'
        'NK: 0 2 0.70\n'
        'NT: 0 2 0.70\n'
        'SN: 0 2 0.70\n'
        'UC: 0 2 0.70\n'
        'UD: 0 2 0.70\n'
        'UN: 0 2 0.70\n'
        'UPC: 0 2 0.70\n'
        'WB: 0 2 0.70\n'
        'WEBS: 0 2 0.70\n'
        'TD: 0 2 0.70\n'
        'IN: 0 1 0.70\n'
        'BL: 0 2 0.70\n'
        'YT: 0 1 0.70\n'
        'BZ: 0 1 0.70\n'
        'MI: 0 2 0.70'
    )
    costupdater_settings_var = (
        'cost = cost\n'
        'sku = sku\n'
        'additional cost = additional_cost\n'
        'business pricing = business_pricing\n'
        'bp strategy = bp_strategy\n'
        'qd strategy = qd_strategy\n'
        '====================================\n'
        'BX: 0.3\n'
        'CANDY: 0.3\n'
        'COS: 0.3\n'
        'CS: 0.3\n'
        'CSC: 0.3\n'
        'DL: 0.3\n'
        'FC: 0.3\n'
        'FD: 0.3\n'
        'FL: 0.75\n'
        'FOUR: 0.3\n'
        'FR: 0.3\n'
        'GEMCO: 0.3\n'
        'IL: 0.75\n'
        'JC: 0.3\n'
        'KH: 0.3\n'
        'LR: 0.3\n'
        'MD: 0.75\n'
        'MONIN PUMP SL: 0.3\n'
        'NC: 0.3\n'
        'NF: 0.3\n'
        'NJ: 0.3\n'
        'NK: 0.3\n'
        'NT: 0.3\n'
        'SN: 0.3\n'
        'UC: 0.3\n'
        'UD: 0.3\n'
        'UN: 0.3\n'
        'UPC: 0.3\n'
        'WB: 0.3\n'
        'WEBS: 0.3\n'
    )

    ordercreate_settings_var = (
        'RESTOCK:\n'
        'upc = Upc\n'
        'pcs = PCS\n'
        'suplier = suplier\n'
        'notes = Notes\n'
        '=====================================================\n'
        'ORDER FORM:\n'
        'upc = UPC\n'
        'pcs = PCS(TOTAL)\n'
        'suplier = suplier')



    global last_scale
    last_scale = 1
    def main_frame_resize():
        global last_scale
        new_width = window.winfo_width()
        new_height = window.winfo_height()
        scale = (new_width*new_height)/(screen_height*screenwidth)
        scale = round(scale,1)
        if scale <= 1:
            scale = 1
        elif scale >= 1.60:
            scale = 1.60
        canvas.place_configure(height=new_height)
        canvas2.place_configure(height=new_height, width=new_width-canvas.winfo_width()+10)
        if scale != last_scale:
            canvas.place_configure(width=resize_dictionary[canvas]['width']*scale)
            canvas2.place_configure(x=resize_dictionary[canvas2]['x']*scale, y=resize_dictionary[canvas2]['y']*scale)
            button_list=[button_1, button_2, button_3, button_4, button_5, button_6, button_7, button_8, button_9, button_10, button_11]
            for button in button_list:
                width = resize_dictionary[button]['width']*scale
                height = resize_dictionary[button]['height']*scale
                button.config(width=width, height=height, round=20*scale)
            last_scale = scale
        return scale

    def main_resize(event, liste, hello, islem):
        scale = main_frame_resize()
        hello.config(font=("JetBrainsMonoRoman Regular", round(24*scale)*-1))
        islem.config(font=("JetBrainsMonoRoman Regular", round(15*scale)*-1))
    window.update()

    canvas2_height = 519
    canvas2_width = 763
    canvas_widht = 175
    canvas2 = Canvas(
        window,
        height = int((window.winfo_height())),
        width = int((window.winfo_width()-canvas_widht)),
        bd = 0,
        highlightthickness = 0,
        relief = "ridge",
        background=color
    )
    canvas2.place(x = int(canvas_widht*scale), y = 0)

    canvas2.pack_propagate(False)
    canvas = Canvas(
        window,
        bg = "#FFFFFF",
        height = int((window.winfo_height())),
        width = int(canvas_widht*scale)+3,
        border=0,
        bd = 0,
        highlightthickness = 0,
        relief = "ridge",
        background=color
    )
    canvas.place(x = 0, y = 0)
    canvas.grid_propagate(False)
    canvas.grid_columnconfigure(0, weight=1)
    canvas.grid_columnconfigure(1, weight=2)

    anasayfa_canvas = Canvas(
        canvas2,
        background=color,
        highlightthickness=0,
        border=0
    )
    anasayfa_canvas.pack(anchor='center', expand=True, side=LEFT)

    line = Frame(
        anasayfa_canvas,
        height=4,
        background=line_color
    )

    hello = Label(
        anasayfa_canvas,
        background=color,
        fg=canvas2_text_color,
        text="KWIEK LLC TOPLU İŞLEM PLATFORMUNA HOŞGELDİNİZ!",
        font=("JetBrainsMonoRoman Regular", 24 * -1)
    )

    islem = Label(
        anasayfa_canvas,
        background=color,
        text="Bir işlem yapmak için lütfen sol menüdeki işlemlerden birini seçiniz...",
        fg=canvas2_text_color,
        font=("JetBrainsMonoRoman Regular", 15 * -1)
    )

    hello.grid(column=0, row=0, sticky='ew', padx=40)
    line.grid(column=0, row=1, sticky='ew', pady=15)
    islem.grid(column=0, row=2, sticky='ew')

    home_icon_selected = PhotoImage(file=relative_to_assets('home_icon_selected_rs.png'))
    home_icon_hover = PhotoImage(file=relative_to_assets('home_icon_hover_rs.png'))
    home_icon_notselected = PhotoImage(file=relative_to_assets('home_icon_notselected_rs.png'))
    program_icon_selected = PhotoImage(file=relative_to_assets('program_icon_selected_rs.png'))
    program_icon_hover = PhotoImage(file=relative_to_assets('program_icon_hover_rs.png'))
    program_icon_notselected = PhotoImage(file=relative_to_assets('program_icon_notselected_rs.png'))
    pad = 5
    button_1 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Expration Date",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_1.grid(column=0, row=6)

    button_1.bind('<Enter>', lambda event: button_hover(event, button_1))
    button_1.bind('<Leave>', lambda event: button_leave(event, button_1))
    button_1.bind("<Button-1>", lambda e: button(canvas2, button_1))
    button_1_line = Frame(canvas, height=2, bg=line_color)
    button_1_line.grid(column=0, row=7, sticky='ew')
    button_2 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Shipment Creater",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_2.grid(column=0, row=8)
    button_2.bind('<Enter>', lambda event: button_hover(event, button_2))
    button_2.bind('<Leave>', lambda event: button_leave(event, button_2))
    button_2.bind("<Button-1>", lambda e: button(canvas2, button_2))
    button_2_line = Frame(canvas, height=2, bg=line_color)
    button_2_line.grid(column=0, row=9, sticky='ew')
    button_3 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="TSV PROGRAMI",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_3.grid(column=0, row=10)
    button_3.bind('<Enter>', lambda event: button_hover(event, button_3))
    button_3.bind('<Leave>', lambda event: button_leave(event, button_3))
    button_3.bind("<Button-1>", lambda e: button(canvas2, button_3))
    button_3_line = Frame(canvas, height=2, bg=line_color)
    button_4 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="RESTOCK",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_4.grid(column=0, row=4)
    button_4.bind('<Enter>', lambda event: button_hover(event, button_4))
    button_4.bind('<Leave>', lambda event: button_leave(event, button_4))
    button_4.bind("<Button-1>", lambda e: button(canvas2, button_4))
    button_4_line = Frame(canvas, height=2, bg=line_color)
    button_4_line.grid(column=0, row=5, sticky='ew')

    button_5 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color='black',
        text="Ana Sayfa",
        align_text="west",
        round=20,
        background='#8AB4F8',
        corners=[0,1,0,1],
        image=home_icon_selected,
        text_pad=pad
    )
    button_5.grid(column=0, row=0, pady=(30,0))
    button_5.bind('<Enter>', lambda event: button_hover(event, button_5))
    button_5.bind('<Leave>', lambda event: button_leave(event, button_5))
    button_5.bind("<Button-1>", lambda e: button(canvas2, button_5))
    button_5_line1 = Frame(canvas, height=2, bg=line_color)
    button_5_line2 = Frame(canvas, height=2, bg=line_color)
    button_5_line3 = Frame(canvas, height=2, bg=line_color)
    button_5_line1.grid(column=0, row=1, sticky='ew', pady=(20, 1))
    button_5_line2.grid(column=0, row=2, sticky='ew', pady=(1, 1))
    button_5_line3.grid(column=0, row=3, sticky='ew', pady=(1, 20))

    button_6_line = Frame(canvas, height=2, bg=line_color)
    button_6_line.grid(column=0, row=11, sticky='ew')
    button_6 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Invoice",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_6.grid(column=0, row=12)
    button_6.bind('<Enter>', lambda event: button_hover(event, button_6))
    button_6.bind('<Leave>', lambda event: button_leave(event, button_6))
    button_6.bind("<Button-1>", lambda e: button(canvas2, button_6))

    button_7_line = Frame(canvas, height=2, bg=line_color)
    button_7_line.grid(column=0, row=13, sticky='ew')
    button_7 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Converter",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_7.grid(column=0, row=14)
    button_7.bind('<Enter>', lambda event: button_hover(event, button_7))
    button_7.bind('<Leave>', lambda event: button_leave(event, button_7))
    button_7.bind("<Button-1>", lambda e: button(canvas2, button_7))

    button_8_line = Frame(canvas, height=2, bg=line_color)
    button_8_line.grid(column=0, row=15, sticky='ew')

    button_8 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Cost Updater",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_8.grid(column=0, row=16)
    button_8.bind('<Enter>', lambda event: button_hover(event, button_8))
    button_8.bind('<Leave>', lambda event: button_leave(event, button_8))
    button_8.bind("<Button-1>", lambda e: button(canvas2, button_8))

    button_9_line = Frame(canvas, height=2, bg=line_color)
    button_9_line.grid(column=0, row=21, sticky='ew')
    button_9 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Update",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_9.grid(column=0, row=22)
    button_9.bind('<Enter>', lambda event: button_hover(event, button_9))
    button_9.bind('<Leave>', lambda event: button_leave(event, button_9))
    button_9.bind("<Button-1>", lambda e: button(canvas2, button_9))

    button_10_line = Frame(canvas, height=2, bg=line_color)
    button_10_line.grid(column=0, row=17, sticky='ew')
    button_10 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Invoice Finder",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_10.grid(column=0, row=18)
    button_10.bind('<Enter>', lambda event: button_hover(event, button_10))
    button_10.bind('<Leave>', lambda event: button_leave(event, button_10))
    button_10.bind("<Button-1>", lambda e: button(canvas2, button_10))

    button_11_line = Frame(canvas, height=2, bg=line_color)
    button_11_line.grid(column=0, row=19, sticky='ew')
    button_11 = MyButton(
        canvas,
        width=canvas_widht,
        height=45,
        text_color=canvas2_text_color,
        text="Order Create",
        align_text="west",
        round=20,
        background=color,
        corners=[0,1,0,1],
        image=program_icon_notselected,
        text_pad=pad
    )
    button_11.grid(column=0, row=20)
    button_11.bind('<Enter>', lambda event: button_hover(event, button_11))
    button_11.bind('<Leave>', lambda event: button_leave(event, button_11))
    button_11.bind("<Button-1>", lambda e: button(canvas2, button_11))
    dictionary = {
        button_1: 0,
        button_2: 0,
        button_3: 0,
        button_4: 0,
        button_5: 1,
        button_6: 0,
        button_7: 0,
        button_8: 0,
        button_9: 0,
        button_10: 0,
        button_11: 0,
    }
    version = Label(canvas, fg=canvas2_text_color, bg=color, text=CURRENT_VERSION, font=('Helvatica', 8))
    version.place(x=0, y=0)
    window.after(1, is_connected_whenstart)
    liste = [canvas, canvas2, button_1, button_2, button_3, button_4, button_5, button_6, button_7, button_8]

    window.update_idletasks()
    resize_dictionary = {
        canvas: {'width': canvas.winfo_width(), 'height': canvas.winfo_height(), 'x': canvas.winfo_x(), 'y': canvas.winfo_y()},
        canvas2: {'width': canvas2.winfo_width(), 'height': canvas2.winfo_height(), 'x': canvas2.winfo_x(), 'y': canvas2.winfo_y()},
        button_1: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_1.winfo_x(), 'y': button_1.winfo_y()},
        button_2: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_2.winfo_x(), 'y': button_2.winfo_y()},
        button_3: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_3.winfo_x(), 'y': button_3.winfo_y()},
        button_4: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_4.winfo_x(), 'y': button_4.winfo_y()},
        button_5: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_5.winfo_x(), 'y': button_5.winfo_y()},
        button_6: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_6.winfo_x(), 'y': button_6.winfo_y()},
        button_7: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_7.winfo_x(), 'y': button_7.winfo_y()},
        button_8: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_8.winfo_x(), 'y': button_8.winfo_y()},
        button_9: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_9.winfo_x(), 'y': button_9.winfo_y()},
        button_10: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_10.winfo_x(), 'y': button_10.winfo_y()},
        button_11: {'width': canvas_widht*scale, 'height': 45*scale, 'x': button_11.winfo_x(), 'y': button_11.winfo_y()},
    }
    window.bind('<Configure>', lambda e: main_resize(e, liste, hello, islem))
    #888888
    window.mainloop()


